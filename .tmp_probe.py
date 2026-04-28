from openpyxl import load_workbook
from pathlib import Path
p=Path(r"C:/Users/gupta/SumoCallibration/ProcessedVideoOutput/Krishna Marg (Kupondole Busstop)/20260222/9/approach_directional_09.20.00-09.25.00_R__0_0__0.xlsx")
wb=load_workbook(p,data_only=True)
for ws in wb.worksheets:
    if ws.title.strip().lower() not in {"direction counts","events"}: continue
    print("Sheet",ws.title)
    found=0
    for r in range(1, ws.max_row+1):
      for c in range(1, ws.max_column+1):
        v=ws.cell(r,c).value
        if isinstance(v,str):
          s=v.strip()
          if s in {"East-to-West","North-to-West"} or s.upper() in {"EW","NW"}:
            print(r,c,repr(s))
            found+=1
            if found>=20: break
      if found>=20: break
    print('count>=',found)
