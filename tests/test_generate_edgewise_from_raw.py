"""Unit tests for generate_edgewise_from_raw.py utility functions."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest

import generate_edgewise_from_raw as gen


class TestNormalizeDateToken:
    def test_yyyymmdd_format(self):
        assert gen.normalize_date_token("20260222") == "2026-02-22"

    def test_ddmmyyyy_format(self):
        assert gen.normalize_date_token("22022026") == "2026-02-22"

    def test_iso_format(self):
        assert gen.normalize_date_token("2026-02-22") == "2026-02-22"

    def test_slash_separated(self):
        assert gen.normalize_date_token("2026/02/22") == "2026-02-22"

    def test_underscore_separated(self):
        assert gen.normalize_date_token("2026_02_22") == "2026-02-22"

    def test_empty_string(self):
        assert gen.normalize_date_token("") is None

    def test_whitespace_only(self):
        assert gen.normalize_date_token("   ") is None

    def test_invalid_date(self):
        assert gen.normalize_date_token("abc") is None

    def test_ddmmyyyy_dash_separated(self):
        assert gen.normalize_date_token("22-02-2026") == "2026-02-22"


class TestParseIntervalFromFilename:
    def test_time_range_in_filename(self):
        path = Path("approach_directional_09.00.00-09.05.00_foo.xlsx")
        result = gen.parse_interval_from_filename(path)
        assert result == "09:00:00-09:05:00"

    def test_time_range_colon_separated(self):
        path = Path("data_08:30:00-08:35:00.xlsx")
        result = gen.parse_interval_from_filename(path)
        assert result == "08:30:00-08:35:00"

    def test_numeric_stem_four_digit(self):
        path = Path("0905.xlsx")
        result = gen.parse_interval_from_filename(path)
        assert result == "09:15:00-09:20:00"

    def test_numeric_stem_three_digit(self):
        path = Path("905.xlsx")
        result = gen.parse_interval_from_filename(path)
        assert result == "09:15:00-09:20:00"

    def test_no_interval_detectable(self):
        path = Path("some_random_file.xlsx")
        result = gen.parse_interval_from_filename(path)
        assert result is None

    def test_time_range_without_seconds(self):
        path = Path("data_08.30-08.35.xlsx")
        result = gen.parse_interval_from_filename(path)
        assert result == "08:30:00-08:35:00"


class TestClassifyFilenamePriority:
    def test_approach_directional(self):
        assert gen.classify_filename_priority(Path("approach_directional_09.xlsx")) == 4

    def test_zone_map(self):
        assert gen.classify_filename_priority(Path("zone_map_1.xlsx")) == 3

    def test_direction(self):
        assert gen.classify_filename_priority(Path("direction_counts.xlsx")) == 2

    def test_numeric_stem(self):
        assert gen.classify_filename_priority(Path("0905.xlsx")) == 1

    def test_other(self):
        assert gen.classify_filename_priority(Path("random.xlsx")) == 0


class TestParseTrailingTimestamp:
    def test_valid_timestamp(self):
        path = Path("file_20260222_091500.xlsx")
        result = gen.parse_trailing_timestamp(path)
        assert result > 0

    def test_no_timestamp(self):
        path = Path("file.xlsx")
        result = gen.parse_trailing_timestamp(path)
        assert result == -1

    def test_invalid_timestamp(self):
        path = Path("file_99999999_999999.xlsx")
        result = gen.parse_trailing_timestamp(path)
        assert result == -1


class TestToSeconds:
    def test_midnight(self):
        assert gen.to_seconds("00:00:00") == 0

    def test_one_hour(self):
        assert gen.to_seconds("01:00:00") == 3600

    def test_mixed(self):
        assert gen.to_seconds("09:15:30") == 9 * 3600 + 15 * 60 + 30

    def test_end_of_day(self):
        assert gen.to_seconds("23:59:59") == 23 * 3600 + 59 * 60 + 59

    def test_invalid_format(self):
        assert gen.to_seconds("abc") is None

    def test_two_parts(self):
        assert gen.to_seconds("09:15") is None

    def test_out_of_range_hours(self):
        assert gen.to_seconds("25:00:00") is None

    def test_out_of_range_minutes(self):
        assert gen.to_seconds("09:61:00") is None

    def test_out_of_range_seconds(self):
        assert gen.to_seconds("09:00:60") is None

    def test_whitespace_handling(self):
        assert gen.to_seconds("  09:00:00  ") == 9 * 3600


class TestIntervalInWindow:
    def test_inside_window(self):
        begin = gen.to_seconds("08:00:00")
        end = gen.to_seconds("10:00:00")
        assert gen.interval_in_window("09:00:00-09:05:00", begin, end) is True

    def test_outside_window(self):
        begin = gen.to_seconds("08:00:00")
        end = gen.to_seconds("10:00:00")
        assert gen.interval_in_window("07:00:00-07:05:00", begin, end) is False

    def test_at_window_boundary(self):
        begin = gen.to_seconds("08:00:00")
        end = gen.to_seconds("10:00:00")
        assert gen.interval_in_window("08:00:00-10:00:00", begin, end) is True

    def test_partially_outside(self):
        begin = gen.to_seconds("08:00:00")
        end = gen.to_seconds("10:00:00")
        assert gen.interval_in_window("09:55:00-10:05:00", begin, end) is False

    def test_invalid_interval(self):
        begin = gen.to_seconds("08:00:00")
        end = gen.to_seconds("10:00:00")
        assert gen.interval_in_window("invalid", begin, end) is False


class TestCanonicalizeTo5minInterval:
    def test_aligned_interval(self):
        assert gen.canonicalize_to_5min_interval("09:00:00-09:05:00") == "09:00:00-09:05:00"

    def test_unaligned_interval(self):
        result = gen.canonicalize_to_5min_interval("09:01:00-09:06:00")
        assert result == "09:00:00-09:05:00"

    def test_at_midnight(self):
        result = gen.canonicalize_to_5min_interval("00:00:00-00:05:00")
        assert result == "00:00:00-00:05:00"

    def test_invalid_start(self):
        result = gen.canonicalize_to_5min_interval("invalid-09:05:00")
        assert result is None

    def test_near_end_of_day(self):
        result = gen.canonicalize_to_5min_interval("23:55:00-00:00:00")
        assert result == "23:55:00-24:00:00"


class TestNormalizeDirection:
    def test_basic(self):
        assert gen.normalize_direction("North-to-South") == "North-to-South"

    def test_extra_spaces(self):
        assert gen.normalize_direction("  North  to  South  ") == "North to South"

    def test_none_value(self):
        assert gen.normalize_direction(None) == ""

    def test_empty_string(self):
        assert gen.normalize_direction("") == ""


class TestResolveLocationLabel:
    def test_known_location(self):
        assert gen.resolve_location_label("Thapathali") == "Thapathali"

    def test_known_with_long_name(self):
        assert gen.resolve_location_label("Krishna Marg (Kupondole Busstop)") == "Kupondole Busstop"

    def test_case_insensitive(self):
        assert gen.resolve_location_label("thapathali") == "Thapathali"

    def test_unknown_location(self):
        assert gen.resolve_location_label("Unknown Place") is None


class TestIsTotalText:
    def test_total(self):
        assert gen.is_total_text("Total") is True

    def test_total_lowercase(self):
        assert gen.is_total_text("total") is True

    def test_total_with_spaces(self):
        assert gen.is_total_text("  Total  ") is True

    def test_not_total(self):
        assert gen.is_total_text("North") is False

    def test_none_value(self):
        assert gen.is_total_text(None) is False


class TestApproachFromDirection:
    def test_with_to_separator(self):
        assert gen.approach_from_direction("North-to-South") == "North"

    def test_without_separator(self):
        assert gen.approach_from_direction("North") == "North"

    def test_multiple_to(self):
        assert gen.approach_from_direction("North-to-South-to-East") == "North"

    def test_whitespace(self):
        assert gen.approach_from_direction("  North-to-South  ") == "North"


class TestChooseBestCandidate:
    def test_prefers_primary_root(self):
        candidates = [
            {"root_idx": 1, "path": Path("a/b/c/approach_directional_09.xlsx")},
            {"root_idx": 0, "path": Path("a/b/c/approach_directional_09.xlsx")},
        ]
        best = gen.choose_best_candidate(candidates)
        assert best["root_idx"] == 0

    def test_prefers_higher_priority_filename(self):
        candidates = [
            {"root_idx": 0, "path": Path("a/b/c/0905.xlsx")},
            {"root_idx": 0, "path": Path("a/b/c/approach_directional_09.xlsx")},
        ]
        best = gen.choose_best_candidate(candidates)
        assert "approach_directional" in str(best["path"])

    def test_single_candidate(self):
        candidates = [{"root_idx": 0, "path": Path("a/b/c/file.xlsx")}]
        best = gen.choose_best_candidate(candidates)
        assert best["root_idx"] == 0


class TestGetSheetCaseInsensitive:
    def test_exact_match(self):
        from openpyxl import Workbook
        wb = Workbook()
        wb.create_sheet("Direction Counts")
        sheet = gen.get_sheet_case_insensitive(wb, "Direction Counts")
        assert sheet is not None

    def test_case_insensitive_match(self):
        from openpyxl import Workbook
        wb = Workbook()
        wb.create_sheet("direction counts")
        sheet = gen.get_sheet_case_insensitive(wb, "Direction Counts")
        assert sheet is not None

    def test_no_match(self):
        from openpyxl import Workbook
        wb = Workbook()
        sheet = gen.get_sheet_case_insensitive(wb, "NonExistent")
        assert sheet is None


class TestBuildOutputSheets:
    def test_empty_dataframe(self):
        df = pd.DataFrame()
        result = gen.build_output_sheets(df)
        assert "Bus_5MinCols" in result
        assert "Car_5MinCols" in result
        assert "Motorcycle_5MinCols" in result
        assert "Total_Volume_5MinCols" in result
        for key, frame in result.items():
            assert isinstance(frame, pd.DataFrame)

    def test_with_data(self):
        rows = [
            {
                "Location": "Thapathali",
                "Date": "2026-02-22",
                "Approach": "North",
                "Direction": "North-to-South",
                "VehicleClass": "Car",
                "Count": 10.0,
                "Interval": "09:00:00-09:05:00",
                "SourceWorkbook": "test.xlsx",
            },
            {
                "Location": "Thapathali",
                "Date": "2026-02-22",
                "Approach": "North",
                "Direction": "North-to-South",
                "VehicleClass": "Bus",
                "Count": 5.0,
                "Interval": "09:00:00-09:05:00",
                "SourceWorkbook": "test.xlsx",
            },
            {
                "Location": "Thapathali",
                "Date": "2026-02-22",
                "Approach": "North",
                "Direction": "North-to-South",
                "VehicleClass": "Motorcycle",
                "Count": 3.0,
                "Interval": "09:00:00-09:05:00",
                "SourceWorkbook": "test.xlsx",
            },
        ]
        df = pd.DataFrame(rows).astype({"Direction": "object", "Approach": "object"})
        result = gen.build_output_sheets(df)
        assert len(result["Car_5MinCols"]) == 1
        assert len(result["Bus_5MinCols"]) == 1
        assert len(result["Total_Volume_5MinCols"]) == 1

    def test_total_rows_excluded(self):
        rows = [
            {
                "Location": "Thapathali",
                "Date": "2026-02-22",
                "Approach": "Total",
                "Direction": "Total",
                "VehicleClass": "Car",
                "Count": 100.0,
                "Interval": "09:00:00-09:05:00",
                "SourceWorkbook": "test.xlsx",
            },
        ]
        df = pd.DataFrame(rows).astype({"Direction": "object", "Approach": "object"})
        result = gen.build_output_sheets(df)
        assert len(result["Car_5MinCols"]) == 0
