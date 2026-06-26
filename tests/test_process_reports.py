"""Unit tests for process_20260222_reports.py utility functions."""
from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import pandas as pd
import pytest
from openpyxl import Workbook

import process_20260222_reports as proc


class TestParseIntervalStart:
    def test_four_digit_code(self):
        assert proc.parse_interval_start("0905") == "09:05:00"

    def test_three_digit_code(self):
        assert proc.parse_interval_start("905") == "09:05:00"

    def test_midnight(self):
        assert proc.parse_interval_start("0000") == "00:00:00"

    def test_afternoon(self):
        assert proc.parse_interval_start("1430") == "14:30:00"


class TestParseIntervalEnd:
    def test_basic(self):
        result = proc.parse_interval_end("09:05:00")
        assert result == "09:10:00"

    def test_hour_boundary(self):
        result = proc.parse_interval_end("09:55:00")
        assert result == "10:00:00"

    def test_midnight_wrap(self):
        result = proc.parse_interval_end("23:55:00")
        assert result == "00:00:00"


class TestToInt:
    def test_integer(self):
        assert proc.to_int(42) == 42

    def test_float(self):
        assert proc.to_int(42.7) == 42

    def test_string_number(self):
        assert proc.to_int("42") == 42

    def test_none(self):
        assert proc.to_int(None) == 0

    def test_empty_string(self):
        assert proc.to_int("") == 0

    def test_whitespace_string(self):
        assert proc.to_int("   ") == 0


class TestAutoSizeColumns:
    def test_basic(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Short", "A much longer column header value"])
        ws.append(["x", "y"])
        proc.autosize_columns(ws)
        assert ws.column_dimensions["A"].width > 0
        assert ws.column_dimensions["B"].width > ws.column_dimensions["A"].width


class TestWriteDataframe:
    def test_writes_sheet(self):
        wb = Workbook()
        df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        proc.write_dataframe(wb, "TestSheet", df)
        assert "TestSheet" in wb.sheetnames
        sheet = wb["TestSheet"]
        assert sheet.cell(1, 1).value == "A"
        assert sheet.cell(1, 2).value == "B"
        assert sheet.cell(2, 1).value == 1

    def test_replaces_existing_sheet(self):
        wb = Workbook()
        wb.create_sheet("TestSheet")
        df = pd.DataFrame({"X": [10]})
        proc.write_dataframe(wb, "TestSheet", df)
        sheet = wb["TestSheet"]
        assert sheet.cell(1, 1).value == "X"


class TestRemoveEventSheets:
    def test_removes_event_sheets(self, tmp_path):
        wb = Workbook()
        wb.create_sheet("Frame Events")
        wb.create_sheet("Summary")
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = proc.remove_event_sheets(path)
        assert result is True
        wb2 = Workbook()
        from openpyxl import load_workbook
        wb2 = load_workbook(path)
        assert "Frame Events" not in wb2.sheetnames
        assert "Summary" in wb2.sheetnames

    def test_no_event_sheets(self, tmp_path):
        wb = Workbook()
        wb.create_sheet("Summary")
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = proc.remove_event_sheets(path)
        assert result is False


class TestRowsFromApproachWise:
    def test_basic(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Approach-Wise"
        ws.append(["Approach", "Direction", "Total", "Bus", "Car", "Motorcycle"])
        ws.append(["North", "North-to-South", 100, 10, 60, 30])
        ws.append(["South", "South-to-North", 80, 5, 50, 25])
        rows = proc.rows_from_approach_wise(ws)
        assert len(rows) == 2
        assert rows[0]["direction"] == "North-to-South"
        assert rows[0]["total"] == 100
        assert rows[0]["bus"] == 10

    def test_skips_empty_direction(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Approach", "Direction", "Total", "Bus", "Car", "Motorcycle"])
        ws.append(["North", None, 100, 10, 60, 30])
        rows = proc.rows_from_approach_wise(ws)
        assert len(rows) == 0


class TestRowsFromSummary:
    def test_basic(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Direction", "Total", "Bus", "Car", "Motorcycle"])
        ws.append(["North-to-South", 100, 10, 60, 30])
        ws.append(["TOTAL", 200, 20, 120, 60])
        rows = proc.rows_from_summary(ws)
        assert len(rows) == 1
        assert rows[0]["direction"] == "North-to-South"

    def test_skips_total_row(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Direction", "Total", "Bus", "Car", "Motorcycle"])
        ws.append(["TOTAL", 200, 20, 120, 60])
        rows = proc.rows_from_summary(ws)
        assert len(rows) == 0


class TestRowsFromDirectionCounts:
    def test_basic(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Direction", "Class", "Count"])
        ws.append(["North-to-South", "car", 50])
        ws.append(["North-to-South", "bus", 10])
        ws.append(["North-to-South", "motorcycle", 30])
        rows = proc.rows_from_direction_counts(ws)
        assert len(rows) == 1
        assert rows[0]["car"] == 50
        assert rows[0]["bus"] == 10
        assert rows[0]["motorcycle"] == 30
        assert rows[0]["total"] == 90

    def test_missing_columns(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Foo", "Bar", "Baz"])
        ws.append(["a", "b", "c"])
        rows = proc.rows_from_direction_counts(ws)
        assert len(rows) == 0


class TestExtractDirectionRows:
    def test_approach_wise_priority(self, tmp_path):
        wb = Workbook()
        ws_aw = wb.create_sheet("Approach-Wise")
        ws_aw.append(["Approach", "Direction", "Total", "Bus", "Car", "Motorcycle"])
        ws_aw.append(["North", "North-to-South", 100, 10, 60, 30])
        ws_dc = wb.create_sheet("Direction Counts")
        ws_dc.append(["Direction", "Class", "Count"])
        ws_dc.append(["North-to-South", "car", 999])
        del wb["Sheet"]
        path = tmp_path / "test.xlsx"
        wb.save(path)
        rows = proc.extract_direction_rows(path)
        assert len(rows) == 1
        assert rows[0]["total"] == 100

    def test_falls_back_to_direction_counts(self, tmp_path):
        wb = Workbook()
        ws_dc = wb.create_sheet("Direction Counts")
        ws_dc.append(["Direction", "Class", "Count"])
        ws_dc.append(["North-to-South", "car", 50])
        del wb["Sheet"]
        path = tmp_path / "test.xlsx"
        wb.save(path)
        rows = proc.extract_direction_rows(path)
        assert len(rows) == 1
        assert rows[0]["car"] == 50

    def test_no_matching_sheets(self, tmp_path):
        wb = Workbook()
        path = tmp_path / "test.xlsx"
        wb.save(path)
        rows = proc.extract_direction_rows(path)
        assert len(rows) == 0


class TestBuildFolderReport:
    def test_returns_empty_when_no_interval_files(self, tmp_path):
        five_min, fifteen_min = proc.build_folder_report(tmp_path, [])
        assert five_min.empty
        assert fifteen_min.empty

    def test_processes_interval_files(self, tmp_path):
        wb = Workbook()
        ws = wb.create_sheet("Summary")
        ws.append(["Direction", "Total", "Bus", "Car", "Motorcycle"])
        ws.append(["North-to-South", 100, 10, 60, 30])
        del wb["Sheet"]
        path = tmp_path / "0905.xlsx"
        wb.save(path)
        five_min, fifteen_min = proc.build_folder_report(tmp_path, [path])
        assert not five_min.empty
        assert "date" in five_min.columns
        assert "direction" in five_min.columns
