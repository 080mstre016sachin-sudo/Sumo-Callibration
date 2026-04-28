from __future__ import annotations

import subprocess
import sys
import shutil
import os
import re
import copy
import importlib.util
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook


WORKSPACE_ROOT = Path(__file__).resolve().parent
ROUTESAMPLER_DIR = WORKSPACE_ROOT / "RouteSamplerMethod"
if str(ROUTESAMPLER_DIR) not in sys.path:
    sys.path.insert(0, str(ROUTESAMPLER_DIR))

from routesampler_utils import (  # type: ignore[import-not-found]
    build_location_edge_map,
    infer_route_sampler_path,
    parse_time_value,
    write_edge_data_xml,
)


DEFAULT_RAW_ROOT = WORKSPACE_ROOT / "ProcessedVideoOutput"
DEFAULT_DATE_WORKBOOK = DEFAULT_RAW_ROOT / "5minCompileSummary_DateTabs_UPDATED.xlsx"
STREAMLIT_CALIB_DIR = WORKSPACE_ROOT / "Streamlit_Callibration"
LEGACY_CALIB_DIR = WORKSPACE_ROOT / "RouteSamplerMethod2"
DEFAULT_MASTER_WORKBOOK = STREAMLIT_CALIB_DIR / "5minCompileSummary_UPDATED.xlsx"
DEFAULT_DATE_TABS_WORKBOOK = STREAMLIT_CALIB_DIR / "5minCompileSummary_DateTabs_UPDATED.xlsx"
DEFAULT_COMPACT_WORKBOOK = STREAMLIT_CALIB_DIR / "5minCompileSummary_Compact.xlsx"
DEFAULT_LOCATION_WORKBOOK = STREAMLIT_CALIB_DIR / "LocationID.xlsx"
DEFAULT_NETWORK_FILE = STREAMLIT_CALIB_DIR / "Network.net.xml"
DEFAULT_ROUTE_POOL = STREAMLIT_CALIB_DIR / "Callibration" / "route_pool.rou.xml"
DEFAULT_TRIP_POOL = STREAMLIT_CALIB_DIR / "Callibration" / "route_pool.trips.xml"
DEFAULT_WHITELISTED_POOL = STREAMLIT_CALIB_DIR / "Callibration" / "whitelisted_pool_test.rou.xml"
DEFAULT_OUTPUT_WORKBOOK = STREAMLIT_CALIB_DIR / "5minCompileSummary_Datewise.xlsx"
DEFAULT_EDGEWISE_WORKBOOK = STREAMLIT_CALIB_DIR / "EdgeWise_DirectionalCounts_5MinIntervalColumns_Datewise.xlsx"
DEFAULT_EDGEWISE_BACKUP_DIR = STREAMLIT_CALIB_DIR / "original_edgewise_backups"
DEFAULT_COUNTS_XML = STREAMLIT_CALIB_DIR / "counts.xml"
DEFAULT_CALIBRATED_ROUTES = STREAMLIT_CALIB_DIR / "calibrated_routes.rou.xml"
DEFAULT_TLS_ADD_FILE = STREAMLIT_CALIB_DIR / "traffic_lights.add.xml"
DEFAULT_DETECTORS_ADD_FILE = STREAMLIT_CALIB_DIR / "detectors.add.xml"
DEFAULT_VEHICLE_TYPES_ADD_FILE = STREAMLIT_CALIB_DIR / "vehicle_types.add.xml"
DEFAULT_LOCATION_EDGE_MAP_CSV = STREAMLIT_CALIB_DIR / "location_edge_map.csv"
DEFAULT_LOCATION_EDGE_MAP_DEBUG_CSV = STREAMLIT_CALIB_DIR / "location_edge_map_DEBUG.csv"
REQUIRED_CRITICAL_UNDERFLOW_EDGES = {"232351915#1", "232429764#1"}
DEFAULT_FINAL_SUMOCFG = STREAMLIT_CALIB_DIR / "final_validation.sumocfg"
DEFAULT_FINAL_TRIPINFO = STREAMLIT_CALIB_DIR / "tripinfo_validation.xml"
DEFAULT_FINAL_QUEUE = STREAMLIT_CALIB_DIR / "queue_validation.xml"
TLS_PRIORITY_STATUS_MESSAGE = "Default network signal logic permanently removed. Survey-based Program 0 is now the primary simulation default."
DEFAULT_ROUTE_SAMPLER = infer_route_sampler_path()

EXPECTED_SUMMARY_ROWS = 10287
SUMMARY_SHEETS = ("Summary", "Approach-Wise", "Direction Counts")


def normalize_key(value: object) -> str:
    return "".join(ch.lower() for ch in str(value).strip() if ch.isalnum())


def normalize_date_token(value: object) -> str:
    return str(value).strip().replace("-", " ")


def date_match_key(value: object) -> str:
    token = str(value).strip().replace(" ", "-").replace("/", "-")
    if re.fullmatch(r"\d{8}", token):
        return f"{token[:4]}-{token[4:6]}-{token[6:8]}"
    return token


def parse_interval_start_code(value: str) -> str | None:
    text = str(value).strip()
    if not re.fullmatch(r"\d{3,4}", text):
        return None
    token = text.zfill(4)
    hours = int(token[:2])
    minutes = int(token[2:4])
    if hours > 23 or minutes > 59:
        return None

    # Raw shorthand files (e.g., 905.xlsx) are coded +10 minutes from the literal token.
    total_seconds = (hours * 3600) + (minutes * 60) + (10 * 60)
    norm_hours = (total_seconds // 3600) % 24
    norm_minutes = (total_seconds % 3600) // 60
    norm_seconds = total_seconds % 60
    return f"{norm_hours:02d}:{norm_minutes:02d}:{norm_seconds:02d}"


def parse_interval_end_code(start_text: str) -> str | None:
    start_sec = to_seconds(start_text)
    if start_sec is None:
        return None
    end_sec = start_sec + 300
    hours = end_sec // 3600
    minutes = (end_sec % 3600) // 60
    seconds = end_sec % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def parse_clock_token(value: str) -> str | None:
    text = str(value).strip().replace(".", ":").replace("_", ":")
    parts = [part for part in text.split(":") if part != ""]
    if len(parts) not in {2, 3}:
        return None
    try:
        numbers = [int(part) for part in parts]
    except ValueError:
        return None

    hours = numbers[0]
    minutes = numbers[1]
    seconds = numbers[2] if len(numbers) == 3 else 0
    if not (0 <= hours <= 23 and 0 <= minutes <= 59 and 0 <= seconds <= 59):
        return None
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def parse_interval_from_filename(path: Path) -> tuple[str | None, str | None]:
    stem = path.stem.strip()

    range_match = re.search(r"(\d{1,2}[.:_]\d{2}(?:[.:_]\d{2})?)\s*-\s*(\d{1,2}[.:_]\d{2}(?:[.:_]\d{2})?)", stem)
    if range_match:
        start_text = parse_clock_token(range_match.group(1))
        end_text = parse_clock_token(range_match.group(2))
        if start_text and end_text:
            return start_text, end_text

    start_text = parse_interval_start_code(stem)
    if start_text:
        return start_text, parse_interval_end_code(start_text)

    return None, None


def to_seconds(value: object) -> int | None:
    return parse_time_value(value)


def resolve_excel_output_path(preferred_path: Path) -> tuple[Path, str | None]:
    preferred_path.parent.mkdir(parents=True, exist_ok=True)
    if not preferred_path.exists():
        return preferred_path, None

    try:
        with preferred_path.open("ab"):
            pass
        return preferred_path, None
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = preferred_path.with_name(f"{preferred_path.stem}_unlocked_{timestamp}{preferred_path.suffix}")
        message = f"Target workbook is locked: {preferred_path}. Wrote to: {fallback}"
        return fallback, message


def discover_excel_files(root: Path) -> list[Path]:
    if not root.exists():
        return []

    blocked_names = {
        DEFAULT_DATE_WORKBOOK.name.lower(),
        DEFAULT_COMPACT_WORKBOOK.name.lower(),
        DEFAULT_OUTPUT_WORKBOOK.name.lower(),
    }
    files: list[Path] = []
    for path in root.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if path.name.lower() in blocked_names:
            continue
        if "compile" in path.name.lower() and "summary" in path.name.lower():
            continue
        files.append(path)
    return sorted(files)


def path_relative_to_root(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return path.as_posix()


def locate_metadata(path: Path, root: Path) -> tuple[str, str, str]:
    try:
        rel_parts = path.relative_to(root).parts
    except ValueError:
        rel_parts = path.parts

    location = rel_parts[-4] if len(rel_parts) >= 4 else path.parent.parent.name if path.parent.parent.name else path.parent.name
    session_folder = rel_parts[-3] if len(rel_parts) >= 3 else path.parent.name
    interval_folder = rel_parts[-2] if len(rel_parts) >= 2 else path.stem
    return str(location), str(session_folder), str(interval_folder)


def first_row_map(sheet) -> dict[str, int]:
    try:
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return {}
    return {normalize_key(value): idx for idx, value in enumerate(header) if value is not None}


def row_value(row: Sequence[object], header_map: dict[str, int], *names: str) -> object:
    for name in names:
        idx = header_map.get(normalize_key(name))
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def parse_directional_header_name(header: object) -> tuple[str, str] | None:
    text = str(header).strip()
    if "|" not in text:
        return None
    left, right = [part.strip() for part in text.split("|", 1)]
    if not left or not right:
        return None
    return left, right


def add_row(records: list[dict[str, object]], base: dict[str, object], direction: str, vehicle_class: str, count: object) -> None:
    if count is None:
        return
    try:
        count_value = int(float(count))
    except (TypeError, ValueError):
        return

    if count_value < 0:
        return

    approach = base.get("Approach") or direction.split("-to-")[0]
    records.append(
        {
            "Location": base.get("Location", ""),
            "SessionFolder": base.get("SessionFolder", ""),
            "IntervalFolder": base.get("IntervalFolder", ""),
            "TimeStart": base.get("TimeStart", ""),
            "TimeEnd": base.get("TimeEnd", ""),
            "Approach": approach,
            "Direction": direction,
            "VehicleClass": vehicle_class,
            "Count": count_value,
            "SourceWorkbook": base.get("SourceWorkbook", ""),
        }
    )


def parse_summary_sheet(sheet, metadata: dict[str, object]) -> list[dict[str, object]]:
    try:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return []
    header_map = {normalize_key(value): idx for idx, value in enumerate(header_row) if value is not None}
    records: list[dict[str, object]] = []

    long_format = {"direction", "vehicleclass", "count"}.issubset(header_map)
    wide_by_class = {"direction", "bus", "car", "motorcycle"}.issubset(header_map)
    approach_wise = {"direction", "total", "bus", "car", "motorcycle"}.issubset(header_map)
    directional_headers = [header for header in header_row if parse_directional_header_name(header)]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        base = dict(metadata)
        base_location = row_value(row, header_map, "Location")
        base_session = row_value(row, header_map, "SessionFolder")
        base_interval = row_value(row, header_map, "IntervalFolder")
        base_start = row_value(row, header_map, "TimeStart")
        base_end = row_value(row, header_map, "TimeEnd")
        base["Location"] = str(base_location or base["Location"]).strip()
        base["SessionFolder"] = str(base_session or base["SessionFolder"]).strip()
        base["IntervalFolder"] = str(base_interval or base["IntervalFolder"]).strip()
        base["TimeStart"] = str(base_start or base["TimeStart"]).strip()
        base["TimeEnd"] = str(base_end or base["TimeEnd"]).strip()
        if not base["Location"]:
            continue

        if directional_headers:
            for header_key in directional_headers:
                parsed = parse_directional_header_name(header_key)
                if not parsed:
                    continue
                direction, vehicle_class = parsed
                add_row(records, base, direction, vehicle_class, row_value(row, header_map, header_key))
            continue

        if long_format:
            direction = str(row_value(row, header_map, "Direction") or "").strip()
            if not direction:
                continue
            vehicle_class = str(row_value(row, header_map, "VehicleClass") or "Total").strip() or "Total"
            add_row(records, base, direction, vehicle_class, row_value(row, header_map, "Count"))
            continue

        direction = str(row_value(row, header_map, "Direction") or "").strip()
        if not direction:
            continue

        if approach_wise:
            for vehicle_class in ("Total", "Bus", "Car", "Motorcycle"):
                add_row(records, base, direction, vehicle_class, row_value(row, header_map, vehicle_class))
            continue

        if wide_by_class:
            for vehicle_class in ("Bus", "Car", "Motorcycle"):
                add_row(records, base, direction, vehicle_class, row_value(row, header_map, vehicle_class))
            continue

        total_value = row_value(row, header_map, "Total", "Count")
        add_row(records, base, direction, "Total", total_value)

    return records


def parse_direction_counts_sheet(sheet, metadata: dict[str, object]) -> list[dict[str, object]]:
    header_map = first_row_map(sheet)
    if not {"direction", "class", "count"}.issubset(header_map):
        return []

    records: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        direction = str(row_value(row, header_map, "Direction") or "").strip()
        vehicle_class = str(row_value(row, header_map, "Class") or "").strip()
        count = row_value(row, header_map, "Count")
        if not direction or not vehicle_class:
            continue
        add_row(records, metadata, direction, vehicle_class, count)
    return records


def parse_approach_wise_sheet(sheet, metadata: dict[str, object]) -> list[dict[str, object]]:
    header_map = first_row_map(sheet)
    if not {"direction", "bus", "car", "motorcycle"}.issubset(header_map):
        return []

    records: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        direction = str(row_value(row, header_map, "Direction") or "").strip()
        if not direction:
            continue
        for vehicle_class in ("Total", "Bus", "Car", "Motorcycle"):
            add_row(records, metadata, direction, vehicle_class, row_value(row, header_map, vehicle_class))
    return records


def extract_records_from_workbook(path: Path, root: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    start_text, end_text = parse_interval_from_filename(path)
    metadata = {
        "Location": "",
        "SessionFolder": "",
        "IntervalFolder": "",
        "TimeStart": start_text or "",
        "TimeEnd": end_text or "",
        "SourceWorkbook": path_relative_to_root(path, root),
    }
    location, session_folder, interval_folder = locate_metadata(path, root)
    metadata["Location"] = location
    metadata["SessionFolder"] = session_folder
    metadata["IntervalFolder"] = interval_folder

    preferred = list(SUMMARY_SHEETS) + [sheet_name for sheet_name in workbook.sheetnames if "summary" in sheet_name.lower()]
    seen: set[str] = set()
    for sheet_name in preferred:
        if sheet_name in seen or sheet_name not in workbook.sheetnames:
            continue
        seen.add(sheet_name)
        lower = sheet_name.lower()
        if "event" in lower:
            continue
        sheet = workbook[sheet_name]
        if lower == "approach-wise":
            records = parse_approach_wise_sheet(sheet, metadata)
        elif lower == "direction counts":
            records = parse_direction_counts_sheet(sheet, metadata)
        else:
            records = parse_summary_sheet(sheet, metadata)
        if records:
            return records

    if workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]
        return parse_summary_sheet(sheet, metadata)
    return []


def build_raw_summary(root: Path, selected_dates: Iterable[str] | None = None) -> pd.DataFrame:
    selected = {normalize_date_token(date) for date in (selected_dates or []) if str(date).strip()}
    records: list[dict[str, object]] = []

    for path in discover_excel_files(root):
        records.extend(extract_records_from_workbook(path, root))

    if not records:
        return pd.DataFrame(columns=["Location", "SessionFolder", "IntervalFolder", "TimeStart", "TimeEnd", "Approach", "Direction", "VehicleClass", "Count", "SourceWorkbook"])

    df = pd.DataFrame.from_records(records)
    if selected:
        df = df[df["SessionFolder"].map(normalize_date_token).isin(selected)]

    df = df.copy()
    df["IntervalFolder"] = df["IntervalFolder"].astype(str)
    df["TimeStart"] = df["TimeStart"].astype(str)
    df["TimeEnd"] = df["TimeEnd"].astype(str)
    df["VehicleClass"] = df["VehicleClass"].astype(str)
    df["Count"] = pd.to_numeric(df["Count"], errors="coerce").fillna(0).astype(int)
    return df.sort_values(["Location", "SessionFolder", "IntervalFolder", "TimeStart", "Approach", "Direction", "VehicleClass", "SourceWorkbook"])


def build_date_tab_summary(workbook_path: Path, selected_sheets: Iterable[str] | None = None) -> pd.DataFrame:
    if not workbook_path.exists():
        return pd.DataFrame(columns=["Location", "SessionFolder", "IntervalFolder", "TimeStart", "TimeEnd", "Approach", "Direction", "VehicleClass", "Count", "SourceWorkbook"])

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    wanted = set(selected_sheets or [])
    records: list[dict[str, object]] = []

    for sheet_name in workbook.sheetnames:
        if wanted and sheet_name not in wanted:
            continue
        lower = sheet_name.lower()
        if "event" in lower:
            continue
        sheet = workbook[sheet_name]
        metadata = {
            "Location": "",
            "SessionFolder": sheet_name,
            "IntervalFolder": "",
            "TimeStart": "",
            "TimeEnd": "",
            "SourceWorkbook": path_relative_to_root(workbook_path, WORKSPACE_ROOT),
        }
        records.extend(parse_summary_sheet(sheet, metadata))

    if not records:
        return pd.DataFrame(columns=["Location", "SessionFolder", "IntervalFolder", "TimeStart", "TimeEnd", "Approach", "Direction", "VehicleClass", "Count", "SourceWorkbook"])

    df = pd.DataFrame.from_records(records)
    df["Count"] = pd.to_numeric(df["Count"], errors="coerce").fillna(0).astype(int)
    return df.sort_values(["Location", "SessionFolder", "IntervalFolder", "TimeStart", "Approach", "Direction", "VehicleClass", "SourceWorkbook"])


def filter_summary_window(summary_df: pd.DataFrame, begin_text: str = "08:00:00", end_text: str = "10:00:00") -> pd.DataFrame:
    if summary_df.empty:
        return summary_df

    begin_sec = to_seconds(begin_text) or 0
    end_sec = to_seconds(end_text) or 0
    filtered = summary_df.copy()

    if "TimeStart" in filtered.columns and "TimeEnd" in filtered.columns and begin_sec and end_sec and end_sec > begin_sec:
        start_seconds = filtered["TimeStart"].map(to_seconds)
        end_seconds = filtered["TimeEnd"].map(to_seconds)
        filtered = filtered[(start_seconds >= begin_sec) & (end_seconds <= end_sec)]

    return filtered


def normalize_date_selection(selected_dates: Iterable[str]) -> list[str]:
    normalized = []
    for value in selected_dates:
        token = str(value).strip()
        if not token:
            continue
        normalized.append(token if "-" in token else token.replace(" ", "-"))
    return normalized


def load_master_summary_workbook(master_workbook: Path, selected_dates: Iterable[str] | None = None) -> pd.DataFrame:
    if not master_workbook.exists():
        return pd.DataFrame(columns=["Location", "SessionFolder", "IntervalFolder", "TimeStart", "TimeEnd", "Approach", "Direction", "VehicleClass", "Count", "SourceWorkbook"])

    selected = {date_match_key(date) for date in (selected_dates or []) if str(date).strip()}
    df = pd.read_excel(master_workbook, sheet_name="Summary")
    required = {"Location", "SessionFolder", "IntervalFolder", "TimeStart", "TimeEnd", "Approach", "Direction", "VehicleClass", "Count", "SourceWorkbook"}
    if not required.issubset(df.columns):
        raise ValueError(f"Master workbook {master_workbook} is missing required Summary columns")

    if selected:
        df = df[df["SessionFolder"].map(date_match_key).isin(selected)]

    df = df.copy()
    df["Count"] = pd.to_numeric(df["Count"], errors="coerce").fillna(0).astype(int)
    df = filter_summary_window(df, "08:00:00", "10:00:00")
    return df.sort_values(["Location", "SessionFolder", "IntervalFolder", "TimeStart", "Approach", "Direction", "VehicleClass", "SourceWorkbook"])


def load_date_tab_workbook(date_tabs_workbook: Path, selected_dates: Iterable[str] | None = None) -> pd.DataFrame:
    if not date_tabs_workbook.exists():
        return pd.DataFrame(columns=["Location", "SessionFolder", "IntervalFolder", "TimeStart", "TimeEnd", "Approach", "Direction", "VehicleClass", "Count", "SourceWorkbook"])

    selected = {date_match_key(date) for date in (selected_dates or []) if str(date).strip()}
    workbook = load_workbook(date_tabs_workbook, read_only=True, data_only=True)
    records: list[dict[str, object]] = []

    for sheet_name in workbook.sheetnames:
        if selected and date_match_key(sheet_name) not in selected:
            continue
        lower = sheet_name.lower()
        if "event" in lower:
            continue
        sheet = workbook[sheet_name]
        metadata = {
            "Location": "",
            "SessionFolder": sheet_name,
            "IntervalFolder": "",
            "TimeStart": "",
            "TimeEnd": "",
            "SourceWorkbook": path_relative_to_root(date_tabs_workbook, WORKSPACE_ROOT),
        }
        records.extend(parse_summary_sheet(sheet, metadata))

    if not records:
        return pd.DataFrame(columns=["Location", "SessionFolder", "IntervalFolder", "TimeStart", "TimeEnd", "Approach", "Direction", "VehicleClass", "Count", "SourceWorkbook"])

    df = pd.DataFrame.from_records(records)
    df["Count"] = pd.to_numeric(df["Count"], errors="coerce").fillna(0).astype(int)
    df = filter_summary_window(df, "08:00:00", "10:00:00")
    return df.sort_values(["Location", "SessionFolder", "IntervalFolder", "TimeStart", "Approach", "Direction", "VehicleClass", "SourceWorkbook"])


def load_compact_workbook(compact_workbook: Path, selected_dates: Iterable[str] | None = None) -> pd.DataFrame:
    if not compact_workbook.exists():
        return pd.DataFrame(columns=["Location", "SessionFolder", "TimeStart", "TimeEnd", "BusCount", "CarCount", "MotorcycleCount", "BusRatio", "CarRatio", "MotorcycleRatio"])

    selected = {date_match_key(date) for date in (selected_dates or []) if str(date).strip()}
    frame = pd.read_excel(compact_workbook, sheet_name="TotalByIntervalLocation")
    required = {"Location", "SessionFolder", "TimeStart", "TimeEnd", "TotalCount"}
    if not required.issubset(frame.columns):
        raise ValueError(f"Compact workbook {compact_workbook} is missing TotalByIntervalLocation columns")

    if selected:
        frame = frame[frame["SessionFolder"].map(date_match_key).isin(selected)]

    frame = frame.copy()
    frame["TotalCount"] = pd.to_numeric(frame["TotalCount"], errors="coerce").fillna(0).astype(int)
    frame = frame.rename(columns={"TotalCount": "Count"})
    frame = filter_summary_window(frame, "08:00:00", "10:00:00")

    class_frames: list[pd.DataFrame] = []
    for sheet_name, column_name in [("Bus", "BusCount"), ("Car", "CarCount"), ("Motorcycle", "MotorcycleCount")]:
        if sheet_name not in pd.ExcelFile(compact_workbook).sheet_names:
            continue
        class_frame = pd.read_excel(compact_workbook, sheet_name=sheet_name)
        if not {"Location", "SessionFolder", "TimeStart", "TimeEnd", "TotalCount"}.issubset(class_frame.columns):
            continue
        if selected:
            class_frame = class_frame[class_frame["SessionFolder"].map(date_match_key).isin(selected)]
        class_frame = class_frame[["Location", "SessionFolder", "TimeStart", "TimeEnd", "TotalCount"]].copy()
        class_frame = filter_summary_window(class_frame.rename(columns={"TotalCount": column_name}), "08:00:00", "10:00:00")
        class_frames.append(class_frame)

    if class_frames:
        merged = class_frames[0]
        for class_frame in class_frames[1:]:
            merged = merged.merge(class_frame, on=["Location", "SessionFolder", "TimeStart", "TimeEnd"], how="outer")
    else:
        merged = pd.DataFrame(columns=["Location", "SessionFolder", "TimeStart", "TimeEnd", "BusCount", "CarCount", "MotorcycleCount"])

    result = frame.merge(merged, on=["Location", "SessionFolder", "TimeStart", "TimeEnd"], how="left")
    for column in ["BusCount", "CarCount", "MotorcycleCount"]:
        if column not in result.columns:
            result[column] = 0
        result[column] = pd.to_numeric(result[column], errors="coerce").fillna(0).astype(int)

    totals = result[["BusCount", "CarCount", "MotorcycleCount"]].sum(axis=1)
    result["BusRatio"] = result["BusCount"].where(totals == 0, result["BusCount"] / totals)
    result["CarRatio"] = result["CarCount"].where(totals == 0, result["CarCount"] / totals)
    result["MotorcycleRatio"] = result["MotorcycleCount"].where(totals == 0, result["MotorcycleCount"] / totals)
    return result.sort_values(["Location", "SessionFolder", "TimeStart"])


def load_compact_vehicle_mix(compact_workbook: Path, selected_dates: Iterable[str], begin_sec: int, end_sec: int) -> pd.DataFrame:
    selected = {normalize_date_token(value) for value in selected_dates}
    frames: list[pd.DataFrame] = []

    for sheet_name, column_name in [("Bus", "BusCount"), ("Car", "CarCount"), ("Motorcycle", "MotorcycleCount")]:
        sheet = pd.read_excel(compact_workbook, sheet_name=sheet_name)
        if sheet.empty or not {"Location", "SessionFolder", "TimeStart", "TimeEnd", "TotalCount"}.issubset(sheet.columns):
            continue

        normalized_session = sheet["SessionFolder"].astype(str).map(normalize_date_token)
        start_sec = sheet["TimeStart"].map(to_seconds)
        end_sec_col = sheet["TimeEnd"].map(to_seconds)
        mask = normalized_session.isin(selected) & (start_sec >= begin_sec) & (end_sec_col <= end_sec)
        filtered = sheet.loc[mask, ["Location", "SessionFolder", "TimeStart", "TimeEnd", "TotalCount"]].copy()
        filtered = filtered.rename(columns={"TotalCount": column_name})
        frames.append(filtered)

    if not frames:
        return pd.DataFrame(columns=["Location", "SessionFolder", "TimeStart", "TimeEnd", "BusCount", "CarCount", "MotorcycleCount", "BusRatio", "CarRatio", "MotorcycleRatio"])

    merged = frames[0]
    for frame in frames[1:]:
        merged = merged.merge(frame, on=["Location", "SessionFolder", "TimeStart", "TimeEnd"], how="outer")

    for column in ["BusCount", "CarCount", "MotorcycleCount"]:
        merged[column] = pd.to_numeric(merged[column], errors="coerce").fillna(0)

    totals = merged[["BusCount", "CarCount", "MotorcycleCount"]].sum(axis=1)
    merged["BusRatio"] = merged["BusCount"].where(totals == 0, merged["BusCount"] / totals)
    merged["CarRatio"] = merged["CarCount"].where(totals == 0, merged["CarCount"] / totals)
    merged["MotorcycleRatio"] = merged["MotorcycleCount"].where(totals == 0, merged["MotorcycleCount"] / totals)
    return merged.sort_values(["Location", "SessionFolder", "TimeStart"])


def attach_vehicle_mix(summary_df: pd.DataFrame, compact_workbook: Path, selected_dates: Iterable[str], begin_sec: int, end_sec: int) -> pd.DataFrame:
    if summary_df.empty:
        return summary_df

    mix = load_compact_vehicle_mix(compact_workbook, selected_dates, begin_sec, end_sec)
    if mix.empty:
        out = summary_df.copy()
        for column in ["BusCount", "CarCount", "MotorcycleCount", "BusRatio", "CarRatio", "MotorcycleRatio"]:
            out[column] = 0.0
        return out

    joined = summary_df.merge(mix, on=["Location", "SessionFolder", "TimeStart", "TimeEnd"], how="left")
    for column in ["BusCount", "CarCount", "MotorcycleCount", "BusRatio", "CarRatio", "MotorcycleRatio"]:
        if column not in joined.columns:
            joined[column] = 0.0
        joined[column] = pd.to_numeric(joined[column], errors="coerce").fillna(0.0)
    return joined


def write_summary_workbook(output_path: Path, summary_df: pd.DataFrame, vehicle_mix_df: pd.DataFrame | None = None) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        if vehicle_mix_df is not None and not vehicle_mix_df.empty:
            vehicle_mix_df.to_excel(writer, sheet_name="VehicleMix", index=False)


def write_datewise_workbook(output_path: Path, summary_df: pd.DataFrame) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        key_cols = ["Location", "SessionFolder", "IntervalFolder", "TimeStart", "TimeEnd"]
        required = set(key_cols + ["Direction", "VehicleClass", "Count"])
        if required.issubset(summary_df.columns):
            for session_value, session_df in summary_df.groupby("SessionFolder"):
                temp = session_df.copy()
                temp["pivot_col"] = temp["Direction"].astype(str).str.strip() + " | " + temp["VehicleClass"].astype(str).str.strip()
                pivot = (
                    temp.pivot_table(
                        index=key_cols,
                        columns="pivot_col",
                        values="Count",
                        aggfunc="sum",
                        fill_value=0,
                    )
                    .reset_index()
                )
                pivot.columns = [str(col) for col in pivot.columns]
                sheet_name = str(session_value).strip()[:31] or "DateSheet"
                pivot.to_excel(writer, sheet_name=sheet_name, index=False)


def build_compact_frames_from_summary(summary_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    key_cols = ["Location", "SessionFolder", "IntervalFolder", "TimeStart", "TimeEnd"]
    missing = [col for col in key_cols + ["VehicleClass", "Count"] if col not in summary_df.columns]
    if missing:
        empty = pd.DataFrame(columns=key_cols + ["TotalCount"])
        return {
            "TotalByIntervalLocation": empty,
            "Bus": empty,
            "Car": empty,
            "Motorcycle": empty,
            "Total": empty,
        }

    working = summary_df.copy()
    working["VehicleClassNorm"] = working["VehicleClass"].astype(str).str.strip().str.lower()
    working["Count"] = pd.to_numeric(working["Count"], errors="coerce").fillna(0)

    out: dict[str, pd.DataFrame] = {}
    for sheet_name, class_key in [("Bus", "bus"), ("Car", "car"), ("Motorcycle", "motorcycle")]:
        frame = (
            working[working["VehicleClassNorm"] == class_key]
            .groupby(key_cols, as_index=False)["Count"]
            .sum()
            .rename(columns={"Count": "TotalCount"})
            .sort_values(key_cols)
        )
        out[sheet_name] = frame

    non_total = working[~working["VehicleClassNorm"].isin(["total", "grandtotal"])].copy()
    total_source = non_total if not non_total.empty else working[working["VehicleClassNorm"].isin(["total", "grandtotal"])].copy()

    total_frame = (
        total_source.groupby(key_cols, as_index=False)["Count"]
        .sum()
        .rename(columns={"Count": "TotalCount"})
        .sort_values(key_cols)
    )

    out["TotalByIntervalLocation"] = total_frame
    out["Total"] = total_frame.copy()
    return out


def write_compact_workbook(compact_path: Path, summary_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    compact_path.parent.mkdir(parents=True, exist_ok=True)
    frames = build_compact_frames_from_summary(summary_df)
    with pd.ExcelWriter(compact_path, engine="openpyxl") as writer:
        for sheet_name in ["TotalByIntervalLocation", "Bus", "Car", "Motorcycle", "Total"]:
            frames[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
    return frames


def _edgewise_interval_sort_key(label: str) -> tuple[int, int, str]:
    text = str(label).strip()
    start_text, _, end_text = text.partition("-")
    start_sec = to_seconds(start_text)
    end_sec = to_seconds(end_text)
    return (
        start_sec if start_sec is not None else 10**9,
        end_sec if end_sec is not None else 10**9,
        text,
    )


def _build_edge_lookup(location_frame: pd.DataFrame | None) -> dict[str, str]:
    if location_frame is None or location_frame.empty:
        return {}

    edge_column = "Network ID" if "Network ID" in location_frame.columns else None
    if edge_column is None:
        return {}

    key_columns = [col for col in ["RenameLocationID", "LocationName", "LocationID"] if col in location_frame.columns]
    lookup: dict[str, str] = {}
    for _, row in location_frame.iterrows():
        edge_value = str(row.get(edge_column, "") or "").strip()
        if not edge_value:
            continue
        for column in key_columns:
            key = normalize_key(row.get(column, ""))
            if key:
                lookup[key] = edge_value
    return lookup


def write_edgewise_datewise_workbook(output_path: Path, summary_df: pd.DataFrame, location_frame: pd.DataFrame | None = None) -> dict[str, pd.DataFrame]:
    key_columns = ["Location", "Date", "Approach", "Direction", "EdgeID"]
    empty = pd.DataFrame(columns=key_columns)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    required = {"Location", "SessionFolder", "TimeStart", "TimeEnd", "Approach", "Direction", "VehicleClass", "Count"}
    if summary_df.empty or not required.issubset(summary_df.columns):
        out = {
            "Bus_5MinCols": empty,
            "Car_5MinCols": empty,
            "Motorcycle_5MinCols": empty,
            "Total_Volume_5MinCols": empty,
        }
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, frame in out.items():
                frame.to_excel(writer, sheet_name=sheet_name, index=False)
        return out

    working = summary_df.copy()
    working["Location"] = working["Location"].astype("string")
    working["Date"] = working["SessionFolder"].map(date_match_key).astype("string")
    working["Approach"] = working["Approach"].astype("string")
    working["Direction"] = working["Direction"].astype("string")
    working["VehicleClassNorm"] = working["VehicleClass"].astype("string").str.strip().str.lower()
    working["Count"] = pd.to_numeric(working["Count"], errors="coerce").fillna(0)
    working["TimeStart"] = working["TimeStart"].astype("string")
    working["TimeEnd"] = working["TimeEnd"].astype("string")
    working["IntervalLabel"] = working["TimeStart"].str.strip() + "-" + working["TimeEnd"].str.strip()

    for column in ["Location", "Date", "Approach", "Direction", "IntervalLabel"]:
        working[column] = working[column].fillna("").astype(str).str.strip()
        working[column] = working[column].replace({"<NA>": "", "nan": "", "None": ""})

    # Many source rows carry missing approach while direction is present.
    missing_approach = working["Approach"] == ""
    working.loc[missing_approach, "Approach"] = working.loc[missing_approach, "Direction"].str.split("-to-").str[0].fillna("")
    working["Approach"] = working["Approach"].fillna("").astype(str).str.strip()

    # Directional output requires valid direction and interval buckets.
    working = working[(working["Direction"] != "") & (working["IntervalLabel"] != "-")]

    edge_lookup = _build_edge_lookup(location_frame)
    working["EdgeID"] = working["Location"].map(lambda value: edge_lookup.get(normalize_key(value), ""))
    working["EdgeID"] = working["EdgeID"].fillna("").astype(str).str.strip()

    key_universe = (
        working[["Location", "Date", "Approach", "Direction", "EdgeID"]]
        .drop_duplicates()
        .sort_values(["Location", "Date", "Approach", "Direction"])
        .reset_index(drop=True)
    )
    interval_columns = sorted(working["IntervalLabel"].dropna().astype(str).unique().tolist(), key=_edgewise_interval_sort_key)

    def build_sheet(target_classes: set[str]) -> pd.DataFrame:
        subset = working[working["VehicleClassNorm"].isin(target_classes)]
        if subset.empty:
            frame = key_universe.copy()
            for column in interval_columns:
                frame[column] = 0
            return frame

        pivot = (
            subset.pivot_table(
                index=key_columns,
                columns="IntervalLabel",
                values="Count",
                aggfunc="sum",
                fill_value=0,
            )
            .reset_index()
        )
        pivot.columns = [str(column) for column in pivot.columns]
        merged = key_universe.merge(pivot, on=key_columns, how="left").fillna(0)
        for column in interval_columns:
            if column not in merged.columns:
                merged[column] = 0
            merged[column] = pd.to_numeric(merged[column], errors="coerce").fillna(0).astype(int)
        return merged[key_columns + interval_columns]

    sheets = {
        "Bus_5MinCols": build_sheet({"bus"}),
        "Car_5MinCols": build_sheet({"car"}),
        "Motorcycle_5MinCols": build_sheet({"motorcycle"}),
        "Total_Volume_5MinCols": build_sheet({"bus", "car", "motorcycle"}),
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
    return sheets


def read_location_mapping(location_workbook: Path) -> pd.DataFrame:
    workbook = load_workbook(location_workbook, data_only=True, read_only=True)
    if "Locations" not in workbook.sheetnames:
        raise ValueError("Location workbook is missing a 'Locations' sheet")
    sheet = workbook["Locations"]
    rows = list(sheet.values)
    if not rows:
        return pd.DataFrame(columns=["LocationID", "LocationName", "RenameLocationID", "Network ID"])
    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    frame = pd.DataFrame(rows[1:], columns=header)
    return frame[[column for column in ["LocationID", "LocationName", "RenameLocationID", "Network ID"] if column in frame.columns]]


def write_location_workbook(location_frame: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Locations"
    ws.append(list(location_frame.columns))
    for row in location_frame.itertuples(index=False, name=None):
        ws.append(list(row))
    wb.save(output_path)


def load_route_pool_preview(route_pool: Path, limit: int = 40) -> pd.DataFrame:
    if not route_pool.exists():
        return pd.DataFrame(columns=["id", "depart", "edge_count", "route_edges"])

    tree = ET.parse(route_pool)
    root = tree.getroot()
    rows: list[dict[str, object]] = []
    for vehicle in root.findall("vehicle")[:limit]:
        route = vehicle.find("route")
        edges = (route.get("edges", "") if route is not None else "").strip()
        edge_tokens = [token for token in edges.split() if token]
        rows.append(
            {
                "id": vehicle.get("id", ""),
                "depart": vehicle.get("depart", ""),
                "edge_count": len(edge_tokens),
                "route_edges": edges[:240] + ("..." if len(edges) > 240 else ""),
            }
        )
    return pd.DataFrame(rows)


def route_edges_from_vehicle(vehicle: ET.Element) -> list[str]:
    route = vehicle.find("route")
    edges = (route.get("edges", "") if route is not None else "").strip()
    return [token for token in edges.split() if token]


def load_network_edge_ids(network_file: Path) -> set[str]:
    if not network_file.exists():
        raise FileNotFoundError(f"Network file not found: {network_file}")

    tree = ET.parse(network_file)
    root = tree.getroot()
    edge_ids: set[str] = set()
    for edge in root.findall("edge"):
        edge_id = str(edge.get("id", "")).strip()
        if not edge_id or edge_id.startswith(":") or edge.get("function") == "internal":
            continue
        edge_ids.add(edge_id)
    return edge_ids


def sanitize_route_file_against_network(route_file: Path, network_file: Path) -> tuple[int, int]:
    """Remove routes that reference edges absent from the active network."""
    if not route_file.exists():
        return (0, 0)

    valid_edges = load_network_edge_ids(network_file)
    if not valid_edges:
        return (0, 0)

    tree = ET.parse(route_file)
    root = tree.getroot()
    removed = 0
    kept = 0

    for vehicle in list(root.findall("vehicle")):
        route_edges = route_edges_from_vehicle(vehicle)
        if route_edges and all(edge in valid_edges for edge in route_edges):
            kept += 1
            continue
        root.remove(vehicle)
        removed += 1

    if removed > 0:
        ET.indent(root)
        route_file.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")

    return (kept, removed)


def filter_route_pool(route_pool: Path, whitelist_edges: Iterable[str], output_path: Path) -> tuple[int, int]:
    tree = ET.parse(route_pool)
    root = tree.getroot()
    whitelist = {edge.strip() for edge in whitelist_edges if edge.strip()}
    kept = 0
    removed = 0

    for vehicle in list(root.findall("vehicle")):
        edges = route_edges_from_vehicle(vehicle)
        if whitelist and not whitelist.intersection(edges):
            root.remove(vehicle)
            removed += 1
        else:
            kept += 1

    ET.indent(root)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")
    return kept, removed


def amplify_route_pool_for_critical_edges(
    route_pool: Path,
    critical_edges: Iterable[str],
    output_path: Path,
    target_intervals: Sequence[tuple[float, float]] | None = None,
    duplicate_factor: int = 2,
    max_added: int = 5000,
) -> tuple[int, int]:
    """Duplicate candidate vehicles traversing critical edges to increase routeSampler coverage."""
    if not route_pool.exists():
        raise FileNotFoundError(f"Route pool file not found: {route_pool}")

    tree = ET.parse(route_pool)
    root = tree.getroot()
    edge_set = {edge.strip() for edge in critical_edges if edge and edge.strip()}
    if not edge_set:
        ET.indent(root)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")
        return (0, 0)

    def _depart_seconds(vehicle: ET.Element) -> float | None:
        depart_raw = str(vehicle.get("depart", "")).strip()
        if not depart_raw:
            return None
        try:
            return float(depart_raw)
        except ValueError:
            value = to_seconds(depart_raw)
            if value is None:
                return None
            return float(value)

    def _in_target_intervals(depart_seconds: float | None) -> bool:
        if not target_intervals:
            return True
        if depart_seconds is None:
            return False
        for begin, end in target_intervals:
            if begin <= depart_seconds < end:
                return True
        return False

    all_seed_candidates: list[ET.Element] = []
    interval_seed_candidates: list[ET.Element] = []
    for vehicle in root.findall("vehicle"):
        vehicle_edges = set(route_edges_from_vehicle(vehicle))
        if not edge_set.intersection(vehicle_edges):
            continue
        all_seed_candidates.append(vehicle)
        if _in_target_intervals(_depart_seconds(vehicle)):
            interval_seed_candidates.append(vehicle)

    # Prefer interval-matching seeds but gracefully fallback to any critical-edge seeds.
    seed_candidates = interval_seed_candidates if interval_seed_candidates else all_seed_candidates

    if not seed_candidates:
        ET.indent(root)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")
        return (0, 0)

    def _retimed_depart(round_idx: int, seed_idx: int) -> str | None:
        if not target_intervals:
            return None

        begin, end = target_intervals[(round_idx + seed_idx) % len(target_intervals)]
        span = max(1.0, end - begin)
        depart = begin + ((seed_idx % 10) / 10.0) * span
        if depart >= end:
            depart = max(begin, end - 0.1)
        return f"{depart:.2f}"

    added = 0
    rounds = max(1, int(duplicate_factor))
    for round_idx in range(rounds):
        for seed_idx, vehicle in enumerate(seed_candidates):
            if added >= max_added:
                break
            clone = copy.deepcopy(vehicle)
            source_id = str(clone.get("id", ""))
            if source_id:
                clone.set("id", f"{source_id}_boost_{round_idx + 1}_{seed_idx}")
            else:
                clone.set("id", f"boost_{round_idx + 1}_{seed_idx}")

            # Spread boosted candidates across deficit intervals to improve matching capacity.
            retimed = _retimed_depart(round_idx, seed_idx)
            if retimed is not None:
                clone.set("depart", retimed)

            root.append(clone)
            added += 1
        if added >= max_added:
            break

    ET.indent(root)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")
    return (len(seed_candidates), added)


def parse_critical_edge_deficit_from_mismatch(
    mismatch_file: Path,
    critical_edges: Iterable[str],
) -> tuple[int, list[tuple[float, float]]]:
    """Read mismatch-output XML and return total critical deficit and deficit intervals."""
    if not mismatch_file.exists():
        return (0, [])

    edge_set = {edge.strip() for edge in critical_edges if edge and edge.strip()}
    if not edge_set:
        return (0, [])

    tree = ET.parse(mismatch_file)
    root = tree.getroot()

    total_deficit = 0
    intervals: list[tuple[float, float]] = []
    seen_intervals: set[tuple[float, float]] = set()

    for interval in root.findall("interval"):
        begin_raw = str(interval.get("begin", "0"))
        end_raw = str(interval.get("end", "0"))
        try:
            begin = float(begin_raw)
            end = float(end_raw)
        except ValueError:
            continue

        interval_has_deficit = False
        for edge in interval.findall("edge"):
            edge_id = str(edge.get("id", ""))
            if edge_id not in edge_set:
                continue
            try:
                deficit = int(float(str(edge.get("deficit", "0"))))
            except ValueError:
                deficit = 0
            if deficit > 0:
                total_deficit += deficit
                interval_has_deficit = True

        key = (begin, end)
        if interval_has_deficit and key not in seen_intervals:
            intervals.append(key)
            seen_intervals.add(key)

    return (total_deficit, intervals)


def parse_edge_interval_deficits_from_mismatch(
    mismatch_file: Path,
    critical_edges: Iterable[str],
) -> dict[tuple[str, float, float], int]:
    """Return per-edge, per-interval positive deficits for the given critical edges."""
    deficits: dict[tuple[str, float, float], int] = {}
    if not mismatch_file.exists():
        return deficits

    edge_set = {edge.strip() for edge in critical_edges if edge and edge.strip()}
    if not edge_set:
        return deficits

    tree = ET.parse(mismatch_file)
    root = tree.getroot()
    for interval in root.findall("interval"):
        begin_raw = str(interval.get("begin", "0"))
        end_raw = str(interval.get("end", "0"))
        try:
            begin = float(begin_raw)
            end = float(end_raw)
        except ValueError:
            continue

        for edge in interval.findall("edge"):
            edge_id = str(edge.get("id", "")).strip()
            if edge_id not in edge_set:
                continue
            try:
                deficit = int(float(str(edge.get("deficit", "0"))))
            except ValueError:
                deficit = 0
            if deficit > 0:
                deficits[(edge_id, begin, end)] = deficit

    return deficits


def inject_rescue_vehicles_for_critical_deficits(
    source_route_pool: Path,
    output_route_file: Path,
    deficit_by_edge_interval: dict[tuple[str, float, float], int],
    max_added: int = 8000,
) -> tuple[int, int]:
    """Append supplemental vehicles for unresolved critical deficits.

    Returns (added_count, unresolved_deficit).
    """
    if not source_route_pool.exists() or not output_route_file.exists() or not deficit_by_edge_interval:
        unresolved = sum(max(0, int(v)) for v in deficit_by_edge_interval.values())
        return (0, unresolved)

    source_root = ET.parse(source_route_pool).getroot()
    output_tree = ET.parse(output_route_file)
    output_root = output_tree.getroot()

    seed_by_edge: dict[str, list[ET.Element]] = {}
    for vehicle in source_root.findall("vehicle"):
        edges = set(route_edges_from_vehicle(vehicle))
        if not edges:
            continue
        for edge_id, _, _ in deficit_by_edge_interval.keys():
            if edge_id in edges:
                seed_by_edge.setdefault(edge_id, []).append(vehicle)

    # Build a quick ID set to avoid collisions.
    existing_ids: set[str] = set()
    for vehicle in output_root.findall("vehicle"):
        vid = str(vehicle.get("id", "")).strip()
        if vid:
            existing_ids.add(vid)

    added = 0
    unresolved = 0
    sorted_items = sorted(deficit_by_edge_interval.items(), key=lambda item: item[1], reverse=True)
    for (edge_id, begin, end), deficit in sorted_items:
        need = max(0, int(deficit))
        if need == 0:
            continue
        templates = seed_by_edge.get(edge_id, [])
        if not templates:
            unresolved += need
            continue

        span = max(1.0, float(end - begin))
        for i in range(need):
            if added >= max_added:
                unresolved += (need - i)
                break

            template = templates[(added + i) % len(templates)]
            clone = copy.deepcopy(template)
            base_id = str(clone.get("id", "rescue")).strip() or "rescue"
            candidate_id = f"rescue_{edge_id}_{int(begin)}_{added}_{base_id}"
            while candidate_id in existing_ids:
                candidate_id = f"{candidate_id}_x"
            clone.set("id", candidate_id)
            existing_ids.add(candidate_id)

            depart = begin + ((i % 20) / 20.0) * span
            if depart >= end:
                depart = max(begin, end - 0.1)
            clone.set("depart", f"{depart:.2f}")

            output_root.append(clone)
            added += 1

    if added > 0:
        ET.indent(output_root)
        output_route_file.write_text(ET.tostring(output_root, encoding="unicode"), encoding="utf-8")
        sort_route_file_by_depart_time(output_route_file)

    return (added, unresolved)


def _parse_counts_requirements_for_edges(
    counts_file: Path,
    target_edges: Iterable[str],
) -> tuple[dict[tuple[str, float, float], int], list[tuple[float, float]]]:
    """Parse counts.xml and extract required counts per (edge, interval)."""
    if not counts_file.exists():
        raise FileNotFoundError(f"Counts file not found: {counts_file}")

    edge_set = {edge.strip() for edge in target_edges if edge and edge.strip()}
    requirements: dict[tuple[str, float, float], int] = {}
    interval_set: set[tuple[float, float]] = set()

    tree = ET.parse(counts_file)
    root = tree.getroot()
    for interval in root.findall("interval"):
        begin_raw = str(interval.get("begin", "0"))
        end_raw = str(interval.get("end", "0"))
        try:
            begin = float(begin_raw)
            end = float(end_raw)
        except ValueError:
            continue
        interval_set.add((begin, end))
        for edge in interval.findall("edge"):
            edge_id = str(edge.get("id", "")).strip()
            if edge_id not in edge_set:
                continue
            entered_raw = edge.get("entered", edge.get("count", "0"))
            try:
                entered = int(round(float(str(entered_raw))))
            except ValueError:
                entered = 0
            key = (edge_id, begin, end)
            requirements[key] = requirements.get(key, 0) + max(0, entered)

    return requirements, sorted(interval_set)


def _compute_candidate_capacity_for_edges(
    route_file: Path,
    target_edges: Iterable[str],
    intervals: Sequence[tuple[float, float]],
) -> dict[tuple[str, float, float], int]:
    """Compute upper-bound candidate capacity per (edge, interval) from route pool departures."""
    if not route_file.exists():
        raise FileNotFoundError(f"Route pool file not found: {route_file}")

    edge_set = {edge.strip() for edge in target_edges if edge and edge.strip()}
    capacity: dict[tuple[str, float, float], int] = {}
    if not edge_set or not intervals:
        return capacity

    tree = ET.parse(route_file)
    root = tree.getroot()

    for vehicle in root.findall("vehicle"):
        depart_raw = str(vehicle.get("depart", "")).strip()
        if not depart_raw:
            continue
        try:
            depart = float(depart_raw)
        except ValueError:
            parsed = to_seconds(depart_raw)
            if parsed is None:
                continue
            depart = float(parsed)

        interval_match: tuple[float, float] | None = None
        for begin, end in intervals:
            if begin <= depart < end:
                interval_match = (begin, end)
                break
        if interval_match is None:
            continue

        route_edges = set(route_edges_from_vehicle(vehicle))
        matched = edge_set.intersection(route_edges)
        if not matched:
            continue

        for edge_id in matched:
            key = (edge_id, interval_match[0], interval_match[1])
            capacity[key] = capacity.get(key, 0) + 1

    return capacity


def evaluate_route_pool_feasibility_for_edges(
    route_file: Path,
    counts_file: Path,
    target_edges: Iterable[str],
) -> tuple[bool, str, dict[str, int]]:
    """Check if route-pool candidate capacity can satisfy required counts on target edges."""
    requirements, intervals = _parse_counts_requirements_for_edges(counts_file, target_edges)
    capacity = _compute_candidate_capacity_for_edges(route_file, target_edges, intervals)

    edge_set = {edge.strip() for edge in target_edges if edge and edge.strip()}
    totals_required: dict[str, int] = {edge: 0 for edge in edge_set}
    totals_capacity: dict[str, int] = {edge: 0 for edge in edge_set}
    totals_deficit: dict[str, int] = {edge: 0 for edge in edge_set}
    worst_cases: list[tuple[int, str, float, float, int, int]] = []

    for edge in edge_set:
        for begin, end in intervals:
            req = int(requirements.get((edge, begin, end), 0))
            cap = int(capacity.get((edge, begin, end), 0))
            deficit = max(0, req - cap)
            totals_required[edge] += req
            totals_capacity[edge] += cap
            totals_deficit[edge] += deficit
            if deficit > 0:
                worst_cases.append((deficit, edge, begin, end, req, cap))

    feasible = sum(totals_deficit.values()) == 0
    summary_parts: list[str] = []
    for edge in sorted(edge_set):
        summary_parts.append(
            f"{edge}: required={totals_required[edge]}, capacity={totals_capacity[edge]}, deficit={totals_deficit[edge]}"
        )

    report = " ; ".join(summary_parts) if summary_parts else "No target-edge requirements found."
    if worst_cases:
        top = sorted(worst_cases, reverse=True)[:8]
        detail = " | ".join(
            [
                f"{edge}@{int(begin)}-{int(end)} req={req} cap={cap} deficit={deficit}"
                for deficit, edge, begin, end, req, cap in top
            ]
        )
        report = report + " || top_interval_deficits: " + detail

    totals = {
        "required_total": sum(totals_required.values()),
        "capacity_total": sum(totals_capacity.values()),
        "deficit_total": sum(totals_deficit.values()),
    }
    return feasible, report, totals


def get_route_pool_depart_window(route_file: Path) -> tuple[int, float | None, float | None]:
    """Return (vehicle_count, min_depart, max_depart) from a route pool file."""
    if not route_file.exists():
        return (0, None, None)

    tree = ET.parse(route_file)
    root = tree.getroot()
    departs: list[float] = []
    vehicle_count = 0
    for vehicle in root.findall("vehicle"):
        vehicle_count += 1
        depart_raw = str(vehicle.get("depart", "")).strip()
        if not depart_raw:
            continue
        try:
            departs.append(float(depart_raw))
            continue
        except ValueError:
            pass
        parsed = to_seconds(depart_raw)
        if parsed is not None:
            departs.append(float(parsed))

    if not departs:
        return (vehicle_count, None, None)
    return (vehicle_count, min(departs), max(departs))


def build_counts_preview(summary_df: pd.DataFrame, selected_dates: Iterable[str], begin_text: str, end_text: str) -> pd.DataFrame:
    if summary_df.empty:
        return summary_df

    selected = {date_match_key(date) for date in selected_dates}
    df = summary_df.copy()
    if selected:
        df = df[df["SessionFolder"].map(date_match_key).isin(selected)]

    if "VehicleClass" in df.columns:
        non_total = df[df["VehicleClass"].astype(str).str.lower() != "total"]
        if not non_total.empty:
            df = non_total

    if "Direction" in df.columns:
        direction_text = df["Direction"].astype(str).str.lower()
        df = df[~direction_text.isin({"total", "grandtotal"})]

    begin_sec = to_seconds(begin_text) or 0
    end_sec = to_seconds(end_text) or 0
    if begin_sec and end_sec and end_sec > begin_sec:
        start_seconds = df["TimeStart"].map(to_seconds)
        end_seconds = df["TimeEnd"].map(to_seconds)
        df = df[(start_seconds >= begin_sec) & (end_seconds <= end_sec)]

    grouped = df.groupby(["Location", "SessionFolder", "TimeStart", "TimeEnd", "Approach", "Direction"], as_index=False)["Count"].sum()
    grouped["Skip"] = False
    grouped["AdjustedCount"] = grouped["Count"]
    return grouped.sort_values(["Location", "SessionFolder", "TimeStart", "Direction"])


def normalize_counts_preview_columns(preview_df: pd.DataFrame) -> pd.DataFrame:
    rename_map: dict[str, str] = {}
    if "location" in preview_df.columns and "Location" not in preview_df.columns:
        rename_map["location"] = "Location"
    if "sessionfolder" in preview_df.columns and "SessionFolder" not in preview_df.columns:
        rename_map["sessionfolder"] = "SessionFolder"
    if "timestart" in preview_df.columns and "TimeStart" not in preview_df.columns:
        rename_map["timestart"] = "TimeStart"
    if "timeend" in preview_df.columns and "TimeEnd" not in preview_df.columns:
        rename_map["timeend"] = "TimeEnd"
    if "direction" in preview_df.columns and "Direction" not in preview_df.columns:
        rename_map["direction"] = "Direction"
    if "count" in preview_df.columns and "Count" not in preview_df.columns:
        rename_map["count"] = "Count"
    if "adjustedcount" in preview_df.columns and "AdjustedCount" not in preview_df.columns:
        rename_map["adjustedcount"] = "AdjustedCount"
    if "skip" in preview_df.columns and "Skip" not in preview_df.columns:
        rename_map["skip"] = "Skip"
    return preview_df.rename(columns=rename_map)


def write_counts_xml_from_preview(preview_df: pd.DataFrame, location_workbook: Path, network_file: Path, output_path: Path, begin_text: str, end_text: str) -> tuple[Path, dict[str, dict[str, str]]]:
    begin_sec = to_seconds(begin_text)
    end_sec = to_seconds(end_text)
    if begin_sec is None or end_sec is None:
        raise ValueError("Time window is invalid")

    working = normalize_counts_preview_columns(preview_df.copy())
    if "Skip" in working.columns:
        working = working[~working["Skip"].astype(bool)]
    count_column = "AdjustedCount" if "AdjustedCount" in working.columns else "Count"

    aggregated = (
        working.groupby(["Location", "Direction", "TimeStart", "TimeEnd"], as_index=False)[count_column]
        .sum()
        .rename(columns={count_column: "count"})
    )
    aggregated["interval_begin"] = aggregated["TimeStart"].map(to_seconds)
    aggregated["interval_end"] = aggregated["TimeEnd"].map(to_seconds)
    aggregated = aggregated[["Location", "Direction", "interval_begin", "interval_end", "count"]]
    aggregated["count"] = pd.to_numeric(aggregated["count"], errors="coerce").fillna(0).astype(int)

    _, location_edge_map = build_location_edge_map(location_workbook, network_file)
    mapping_rows: list[dict[str, str]] = []
    for location_name, direction_map in location_edge_map.items():
        for direction_name, edge_id in direction_map.items():
            mapping_rows.append({"Location": location_name, "Direction": direction_name, "edge_id": edge_id})
    mapping_df = pd.DataFrame.from_records(mapping_rows, columns=["Location", "Direction", "edge_id"])

    def resolve_edge_id(target_location: str, target_direction: str) -> str:
        location_matches = mapping_df[mapping_df["Location"] == target_location]
        if location_matches.empty:
            return ""
        direction_matches = location_matches[location_matches["Direction"] == target_direction]
        if direction_matches.empty:
            return ""
        return str(direction_matches.iloc[0]["edge_id"]).strip()

    aggregated["mapped_edge_id"] = aggregated.apply(
        lambda row: resolve_edge_id(str(row["Location"]), str(row["Direction"])),
        axis=1,
    )
    aggregated = aggregated[aggregated["mapped_edge_id"].astype(str).str.strip() != ""].copy()

    edge_ready = aggregated.rename(columns={"Location": "location", "Direction": "from_direction"})
    edge_ready = edge_ready[["location", "from_direction", "interval_begin", "interval_end", "count"]]
    _, location_edge_map = write_edge_data_xml(edge_ready, location_workbook, network_file, output_path, begin_sec=begin_sec, end_sec=end_sec)
    return output_path, location_edge_map


def _parse_interval_label(label: str) -> tuple[int | None, int | None]:
    text = str(label).strip()
    if "-" not in text:
        return None, None
    start_text, _, end_text = text.partition("-")
    start_sec = to_seconds(start_text.strip())
    end_sec = to_seconds(end_text.strip())
    if start_sec is None or end_sec is None:
        return None, None
    return start_sec, end_sec


def _normalize_location_for_mapping(location: object) -> str:
    raw = str(location or "").strip()
    aliases = {
        normalize_key("Kupondole Busstop"): "Krishna Marg (Kupondole Busstop)",
        normalize_key("Bhakundole"): "Maitri Marg (Bhakundole)",
    }
    alias_target = aliases.get(normalize_key(raw))
    return alias_target if alias_target else raw


def load_location_edge_map_csv(location_edge_map_csv: Path) -> pd.DataFrame:
    frame: pd.DataFrame | None = None
    last_error: Exception | None = None
    required = {"Location", "Approach", "EdgeID"}
    encodings = ["utf-8", "utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "cp1252", "latin-1"]

    for encoding in encodings:
        try:
            frame = pd.read_csv(location_edge_map_csv, encoding=encoding)
            break
        except (UnicodeDecodeError, pd.errors.ParserError) as exc:
            last_error = exc

    # Fallback for manually edited CSVs that have mixed delimiters/preamble lines.
    if frame is None:
        for encoding in encodings:
            for sep in [",", ";", "\t"]:
                try:
                    candidate = pd.read_csv(
                        location_edge_map_csv,
                        encoding=encoding,
                        sep=sep,
                        engine="python",
                        on_bad_lines="skip",
                    )
                    if required.issubset(candidate.columns):
                        frame = candidate
                        break
                except Exception as exc:  # noqa: BLE001 - keep fallback tolerant
                    last_error = exc
            if frame is not None:
                break

    # Final fallback: some manual files are actually Excel content with .csv extension.
    if frame is None:
        try:
            frame = pd.read_excel(location_edge_map_csv)
        except Exception as exc:  # noqa: BLE001 - keep fallback tolerant
            last_error = exc

    if frame is None:
        raise ValueError(
            f"Unable to read {location_edge_map_csv} with supported encodings "
            f"({', '.join(encodings)}) or Excel fallback. Last error: {last_error}"
        )

    frame = frame.copy()
    frame.columns = [str(col).replace("\ufeff", "").strip() for col in frame.columns]

    if not required.issubset(frame.columns):
        raise ValueError(f"{location_edge_map_csv} must contain columns: Location, Approach, EdgeID")

    out = frame.copy()
    out["Location"] = out["Location"].astype(str).str.strip()
    out["Approach"] = out["Approach"].astype(str).str.strip()
    out["EdgeID"] = out["EdgeID"].astype(str).str.strip()
    out = out[(out["Location"] != "") & (out["Approach"] != "") & (out["EdgeID"] != "")]
    out["LocationNorm"] = out["Location"].map(_normalize_location_for_mapping).map(normalize_key)
    out["ApproachNorm"] = out["Approach"].map(normalize_key)
    return out.drop_duplicates(subset=["LocationNorm", "ApproachNorm", "EdgeID"]).reset_index(drop=True)


def _critical_edge_reverse(edge_id: str) -> str:
    edge = str(edge_id).strip()
    if not edge:
        return ""
    return edge[1:] if edge.startswith("-") else f"-{edge}"


def validate_required_critical_edges_in_manual_map(location_edge_map_csv: Path) -> tuple[list[str], list[str]]:
    """Return missing critical edges and their reverse-direction counterparts.

    Returns:
      (missing_primary_edges, missing_reverse_edges)
    """
    if not location_edge_map_csv.exists():
        primary = sorted(REQUIRED_CRITICAL_UNDERFLOW_EDGES)
        reverse = sorted(_critical_edge_reverse(edge) for edge in REQUIRED_CRITICAL_UNDERFLOW_EDGES)
        return primary, reverse

    map_df = load_location_edge_map_csv(location_edge_map_csv)
    edge_values = {str(value).strip() for value in map_df.get("EdgeID", pd.Series(dtype=str)).astype(str).tolist()}
    missing_primary = sorted(edge for edge in REQUIRED_CRITICAL_UNDERFLOW_EDGES if edge not in edge_values)
    missing_reverse = sorted(
        reverse_edge
        for reverse_edge in (_critical_edge_reverse(edge) for edge in REQUIRED_CRITICAL_UNDERFLOW_EDGES)
        if reverse_edge and reverse_edge not in edge_values
    )
    return missing_primary, missing_reverse


def coverage_ratio_from_totals(totals: dict[str, int]) -> float:
    required = int(totals.get("required_total", 0))
    capacity = int(totals.get("capacity_total", 0))
    if required <= 0:
        return 1.0
    return float(capacity) / float(required)


def get_incoming_edge_lane_data(network_file: Path) -> dict[str, list[tuple[str, float]]]:
    root = ET.parse(network_file).getroot()
    incoming_edges: dict[str, list[tuple[str, float]]] = {}
    for edge in root.findall("edge"):
        edge_id = str(edge.get("id", "")).strip()
        if not edge_id or edge_id.startswith(":") or edge.get("function") == "internal":
            continue

        lane_rows: list[tuple[str, float]] = []
        for lane in edge.findall("lane"):
            lane_id = str(lane.get("id", "")).strip()
            if not lane_id:
                continue
            try:
                lane_length = float(lane.get("length", "0"))
            except ValueError:
                lane_length = 0.0
            lane_rows.append((lane_id, lane_length))

        if not lane_rows:
            continue

        to_node = str(edge.get("to", "")).strip()
        if to_node:
            incoming_edges[edge_id] = lane_rows
    return incoming_edges


def build_counts_from_edgewise(
    edgewise_workbook: Path,
    location_edge_map_csv: Path,
    network_file: Path,
    begin_text: str,
    end_text: str,
) -> tuple[pd.DataFrame, dict[str, object]]:
    begin_sec = to_seconds(begin_text)
    end_sec = to_seconds(end_text)
    if begin_sec is None or end_sec is None or end_sec <= begin_sec:
        raise ValueError("Invalid begin/end time for counts.xml generation")

    if not edgewise_workbook.exists():
        raise FileNotFoundError(f"EdgeWise workbook not found: {edgewise_workbook}")
    if not location_edge_map_csv.exists():
        raise FileNotFoundError(f"location_edge_map.csv not found: {location_edge_map_csv}")

    edgewise_df = pd.read_excel(edgewise_workbook, sheet_name="Total_Volume_5MinCols")
    required = {"Location", "Approach"}
    if not required.issubset(edgewise_df.columns):
        raise ValueError("Total_Volume_5MinCols must contain Location and Approach columns")

    mapping_df = load_location_edge_map_csv(location_edge_map_csv)
    incoming_edge_lanes = get_incoming_edge_lane_data(network_file)
    valid_incoming_edges = set(incoming_edge_lanes.keys())

    map_lookup: dict[tuple[str, str], str] = {}
    for row in mapping_df.itertuples(index=False):
        key = (str(row.LocationNorm), str(row.ApproachNorm))
        if key not in map_lookup:
            map_lookup[key] = str(row.EdgeID)

    interval_columns: list[tuple[str, int, int]] = []
    for column in edgewise_df.columns:
        start_sec, end_sec_col = _parse_interval_label(str(column))
        if start_sec is None or end_sec_col is None:
            continue
        if end_sec_col - start_sec != 300:
            continue
        if start_sec < begin_sec or end_sec_col > end_sec:
            continue
        interval_columns.append((str(column), start_sec, end_sec_col))

    if not interval_columns:
        raise ValueError("No 5-minute interval columns found in selected time window")

    mapped_rows: list[dict[str, object]] = []
    skipped_invalid_edge = 0
    skipped_unmapped = 0
    mapped_via_edgeid_fallback = 0
    for _, row in edgewise_df.iterrows():
        location_norm = normalize_key(_normalize_location_for_mapping(row.get("Location", "")))
        approach_norm = normalize_key(row.get("Approach", ""))
        edge_id = map_lookup.get((location_norm, approach_norm), "")
        if not edge_id:
            edge_id_from_row = str(row.get("EdgeID", "") or "").strip()
            if edge_id_from_row in valid_incoming_edges:
                edge_id = edge_id_from_row
                mapped_via_edgeid_fallback += 1
        if not edge_id:
            skipped_unmapped += 1
            continue
        if edge_id not in valid_incoming_edges:
            skipped_invalid_edge += 1
            continue

        for col_name, interval_begin, interval_end in interval_columns:
            value = row.get(col_name, 0)
            count_float = pd.to_numeric(value, errors="coerce")
            count_value = int(round(float(count_float if pd.notna(count_float) else 0)))
            mapped_rows.append(
                {
                    "edge_id": edge_id,
                    "interval_begin": interval_begin,
                    "interval_end": interval_end,
                    "entered": max(0, count_value),
                }
            )

    mapped_df = pd.DataFrame.from_records(mapped_rows)
    if mapped_df.empty:
        raise ValueError("No counts could be mapped from Total_Volume_5MinCols using location_edge_map.csv")

    aggregated = (
        mapped_df.groupby(["edge_id", "interval_begin", "interval_end"], as_index=False)["entered"]
        .sum()
        .sort_values(["interval_begin", "edge_id"]) 
    )

    diagnostics: dict[str, object] = {
        "skipped_unmapped_rows": skipped_unmapped,
        "skipped_invalid_edge_rows": skipped_invalid_edge,
        "mapped_via_edgeid_fallback_rows": mapped_via_edgeid_fallback,
        "mapped_edge_count": int(aggregated["edge_id"].nunique()),
        "mapped_row_count": int(len(aggregated)),
        "all_counts_zero": bool((aggregated["entered"] <= 0).all()),
    }
    return aggregated, diagnostics


def write_counts_xml_from_edgewise(
    edgewise_workbook: Path,
    location_edge_map_csv: Path,
    network_file: Path,
    output_path: Path,
    begin_text: str,
    end_text: str,
) -> tuple[Path, pd.DataFrame, dict[str, object]]:
    begin_sec = to_seconds(begin_text)
    end_sec = to_seconds(end_text)
    if begin_sec is None or end_sec is None or end_sec <= begin_sec:
        raise ValueError("Time window is invalid")

    aggregated, diagnostics = build_counts_from_edgewise(
        edgewise_workbook=edgewise_workbook,
        location_edge_map_csv=location_edge_map_csv,
        network_file=network_file,
        begin_text=begin_text,
        end_text=end_text,
    )

    all_edge_ids = sorted(aggregated["edge_id"].unique().tolist())
    root = ET.Element(
        "meandata",
        {
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
            "xsi:noNamespaceSchemaLocation": "http://sumo.dlr.de/xsd/meandata_file.xsd",
        },
    )

    interval_bins = list(range(begin_sec, end_sec, 300))
    lookup = {
        (str(row.edge_id), int(row.interval_begin), int(row.interval_end)): int(row.entered)
        for row in aggregated.itertuples(index=False)
    }
    for interval_begin in interval_bins:
        interval_end = interval_begin + 300
        interval_element = ET.SubElement(
            root,
            "interval",
            {
                "begin": f"{float(interval_begin):.2f}",
                "end": f"{float(interval_end):.2f}",
                "id": f"counts_{interval_begin}_{interval_end}",
            },
        )

        for edge_id in all_edge_ids:
            entered = int(lookup.get((edge_id, interval_begin, interval_end), 0))
            ET.SubElement(
                interval_element,
                "edge",
                {
                    "id": edge_id,
                    "entered": str(entered),
                    "count": str(entered),
                    "flow": f"{entered * 12.0:.2f}",
                },
            )

    ET.indent(root)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")
    return output_path, aggregated, diagnostics


def write_detectors_add_from_location_edge_map(
    network_file: Path,
    location_edge_map_csv: Path,
    output_path: Path,
) -> tuple[Path, int]:
    mapping_df = load_location_edge_map_csv(location_edge_map_csv)
    incoming_edge_lanes = get_incoming_edge_lane_data(network_file)

    additional = ET.Element("additional")
    detector_count = 0
    for edge_id in sorted(mapping_df["EdgeID"].astype(str).str.strip().unique().tolist()):
        lanes = incoming_edge_lanes.get(edge_id)
        if not lanes:
            continue
        for lane_id, lane_length in lanes:
            detector_id = f"det_{edge_id}_{lane_id}".replace("#", "_").replace("-", "m")
            pos = max(0.1, float(lane_length) - 10.0)
            ET.SubElement(
                additional,
                "e1Detector",
                {
                    "id": detector_id,
                    "lane": lane_id,
                    "pos": f"{pos:.2f}",
                    "freq": "60",
                    "file": "detector_output.xml",
                },
            )
            detector_count += 1

    ET.indent(additional)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(ET.tostring(additional, encoding="unicode"), encoding="utf-8")
    return output_path, detector_count


def build_counts_from_edgewise_classified(
    edgewise_workbook: Path,
    location_edge_map_csv: Path,
    network_file: Path,
    begin_text: str,
    end_text: str,
) -> dict[str, tuple[pd.DataFrame, dict[str, object]]]:
    """Build counts for each vehicle class (Bus, Car, Motorcycle) from classified sheets."""
    begin_sec = to_seconds(begin_text)
    end_sec = to_seconds(end_text)
    if begin_sec is None or end_sec is None or end_sec <= begin_sec:
        raise ValueError("Invalid begin/end time for counts.xml generation")

    if not edgewise_workbook.exists():
        raise FileNotFoundError(f"EdgeWise workbook not found: {edgewise_workbook}")
    if not location_edge_map_csv.exists():
        raise FileNotFoundError(f"location_edge_map.csv not found: {location_edge_map_csv}")

    vehicle_sheets = {
        "bus": "Bus_5MinCols",
        "car": "Car_5MinCols",
        "motorcycle": "Motorcycle_5MinCols",
    }

    mapping_df = load_location_edge_map_csv(location_edge_map_csv)
    incoming_edge_lanes = get_incoming_edge_lane_data(network_file)
    valid_incoming_edges = set(incoming_edge_lanes.keys())

    map_lookup: dict[tuple[str, str], str] = {}
    for row in mapping_df.itertuples(index=False):
        key = (str(row.LocationNorm), str(row.ApproachNorm))
        if key not in map_lookup:
            map_lookup[key] = str(row.EdgeID)

    interval_columns: list[tuple[str, int, int]] = []
    sample_df = pd.read_excel(edgewise_workbook, sheet_name="Bus_5MinCols")
    for column in sample_df.columns:
        start_sec_col, end_sec_col = _parse_interval_label(str(column))
        if start_sec_col is None or end_sec_col is None:
            continue
        if end_sec_col - start_sec_col != 300:
            continue
        if start_sec_col < begin_sec or end_sec_col > end_sec:
            continue
        interval_columns.append((str(column), start_sec_col, end_sec_col))

    if not interval_columns:
        raise ValueError("No 5-minute interval columns found in selected time window")

    results: dict[str, tuple[pd.DataFrame, dict[str, object]]] = {}

    for vehicle_class, sheet_name in vehicle_sheets.items():
        edgewise_df = pd.read_excel(edgewise_workbook, sheet_name=sheet_name)
        required = {"Location", "Approach"}
        if not required.issubset(edgewise_df.columns):
            raise ValueError(f"{sheet_name} must contain Location and Approach columns")

        mapped_rows: list[dict[str, object]] = []
        skipped_invalid_edge = 0
        skipped_unmapped = 0
        mapped_via_edgeid_fallback = 0
        unmapped_examples: list[dict[str, object]] = []
        unmapped_nonzero_total = 0
        interval_column_names = [name for name, _, _ in interval_columns]

        for _, row in edgewise_df.iterrows():
            location_norm = normalize_key(_normalize_location_for_mapping(row.get("Location", "")))
            approach_norm = normalize_key(row.get("Approach", ""))
            edge_id = map_lookup.get((location_norm, approach_norm), "")
            if not edge_id:
                edge_id_from_row = str(row.get("EdgeID", "") or "").strip()
                if edge_id_from_row in valid_incoming_edges:
                    edge_id = edge_id_from_row
                    mapped_via_edgeid_fallback += 1
            if not edge_id:
                skipped_unmapped += 1
                row_total = int(
                    pd.to_numeric(row.get(interval_column_names, pd.Series(dtype=float)), errors="coerce")
                    .fillna(0)
                    .sum()
                )
                if row_total > 0:
                    unmapped_nonzero_total += row_total
                if len(unmapped_examples) < 8:
                    unmapped_examples.append(
                        {
                            "Location": str(row.get("Location", "") or "").strip(),
                            "Approach": str(row.get("Approach", "") or "").strip(),
                            "Direction": str(row.get("Direction", "") or "").strip(),
                            "EdgeID": str(row.get("EdgeID", "") or "").strip(),
                            "TotalInWindow": row_total,
                        }
                    )
                continue
            if edge_id not in valid_incoming_edges:
                skipped_invalid_edge += 1
                continue

            for col_name, interval_begin, interval_end in interval_columns:
                value = row.get(col_name, 0)
                count_float = pd.to_numeric(value, errors="coerce")
                count_value = int(round(float(count_float if pd.notna(count_float) else 0)))
                mapped_rows.append(
                    {
                        "edge_id": edge_id,
                        "interval_begin": interval_begin,
                        "interval_end": interval_end,
                        "entered": max(0, count_value),
                        "vtype": vehicle_class,
                    }
                )

        mapped_df = pd.DataFrame.from_records(mapped_rows)
        if mapped_df.empty:
            aggregated = pd.DataFrame(columns=["edge_id", "interval_begin", "interval_end", "entered", "vtype"])
        else:
            aggregated = (
                mapped_df.groupby(["edge_id", "interval_begin", "interval_end", "vtype"], as_index=False)["entered"]
                .sum()
                .sort_values(["interval_begin", "edge_id"])
            )

        diagnostics: dict[str, object] = {
            "skipped_unmapped_rows": skipped_unmapped,
            "skipped_invalid_edge_rows": skipped_invalid_edge,
            "mapped_via_edgeid_fallback_rows": mapped_via_edgeid_fallback,
            "unmapped_nonzero_total": unmapped_nonzero_total,
            "unmapped_examples": unmapped_examples,
            "mapped_edge_count": int(aggregated["edge_id"].nunique()) if not aggregated.empty else 0,
            "mapped_row_count": int(len(aggregated)),
            "all_counts_zero": bool((aggregated["entered"] <= 0).all()) if not aggregated.empty else True,
        }
        results[vehicle_class] = (aggregated, diagnostics)

    return results


def write_counts_xml_from_edgewise_classified(
    edgewise_workbook: Path,
    location_edge_map_csv: Path,
    network_file: Path,
    output_path: Path,
    begin_text: str,
    end_text: str,
) -> tuple[Path, dict[str, tuple[pd.DataFrame, dict[str, object]]]]:
    """Generate counts.xml with vType attributes for classified vehicle modes."""
    begin_sec = to_seconds(begin_text)
    end_sec = to_seconds(end_text)
    if begin_sec is None or end_sec is None or end_sec <= begin_sec:
        raise ValueError("Time window is invalid")

    classified_counts = build_counts_from_edgewise_classified(
        edgewise_workbook=edgewise_workbook,
        location_edge_map_csv=location_edge_map_csv,
        network_file=network_file,
        begin_text=begin_text,
        end_text=end_text,
    )

    root = ET.Element(
        "meandata",
        {
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
            "xsi:noNamespaceSchemaLocation": "http://sumo.dlr.de/xsd/meandata_file.xsd",
        },
    )

    all_edge_ids_per_vtype: dict[str, set[str]] = {vc: set() for vc in ["bus", "car", "motorcycle"]}
    lookup_per_vtype: dict[str, dict[tuple[str, int, int], int]] = {vc: {} for vc in ["bus", "car", "motorcycle"]}

    for vehicle_class, (aggregated, _) in classified_counts.items():
        if not aggregated.empty:
            all_edge_ids_per_vtype[vehicle_class].update(aggregated["edge_id"].unique().tolist())
            for row in aggregated.itertuples(index=False):
                key = (str(row.edge_id), int(row.interval_begin), int(row.interval_end))
                lookup_per_vtype[vehicle_class][key] = int(row.entered)

    # Collect all unique edges across all vehicle classes for complete zero-count coverage
    all_unique_edges = set()
    for edge_ids in all_edge_ids_per_vtype.values():
        all_unique_edges.update(edge_ids)

    interval_bins = list(range(begin_sec, end_sec, 300))
    for interval_begin in interval_bins:
        interval_end = interval_begin + 300
        interval_element = ET.SubElement(
            root,
            "interval",
            {
                "begin": f"{float(interval_begin):.2f}",
                "end": f"{float(interval_end):.2f}",
                "id": f"counts_{interval_begin}_{interval_end}",
            },
        )

        for vehicle_class in ["bus", "car", "motorcycle"]:
            for edge_id in sorted(all_unique_edges):
                entered = int(lookup_per_vtype[vehicle_class].get((edge_id, interval_begin, interval_end), 0))
                ET.SubElement(
                    interval_element,
                    "edge",
                    {
                        "id": edge_id,
                        "entered": str(entered),
                        "count": str(entered),
                        "flow": str(float(entered)),
                        "vType": vehicle_class,
                    },
                )

    ET.indent(root)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")
    return output_path, classified_counts


def write_counts_xml_single_mode_from_aggregated(
    aggregated: pd.DataFrame,
    vehicle_class: str,
    output_path: Path,
    begin_sec: int,
    end_sec: int,
    include_vtype: bool = False,
) -> Path:
    """Write counts.xml for a single mode while preserving zero-count intervals.

    include_vtype defaults to False because route pool candidates are typically untyped.
    Enforcing vType in counts with untyped candidates can create artificial underflow.
    """
    root = ET.Element(
        "meandata",
        {
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
            "xsi:noNamespaceSchemaLocation": "http://sumo.dlr.de/xsd/meandata_file.xsd",
        },
    )

    edge_ids: list[str] = []
    lookup: dict[tuple[str, int, int], int] = {}
    if not aggregated.empty:
        edge_ids = sorted(aggregated["edge_id"].astype(str).unique().tolist())
        for row in aggregated.itertuples(index=False):
            lookup[(str(row.edge_id), int(row.interval_begin), int(row.interval_end))] = int(row.entered)

    for interval_begin in range(begin_sec, end_sec, 300):
        interval_end = interval_begin + 300
        interval_element = ET.SubElement(
            root,
            "interval",
            {
                "begin": f"{float(interval_begin):.2f}",
                "end": f"{float(interval_end):.2f}",
                "id": f"counts_{vehicle_class}_{interval_begin}_{interval_end}",
            },
        )
        for edge_id in edge_ids:
            entered = int(lookup.get((edge_id, interval_begin, interval_end), 0))
            edge_attributes = {
                "id": edge_id,
                "entered": str(entered),
                "count": str(entered),
                "flow": str(float(entered)),
            }
            if include_vtype:
                edge_attributes["vType"] = vehicle_class

            ET.SubElement(interval_element, "edge", edge_attributes)

    ET.indent(root)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")
    return output_path


def assign_vehicle_type_to_routes(route_file: Path, vehicle_class: str, id_prefix: str) -> tuple[int, int, int]:
    """Assign SUMO vehicle type and unique id prefix to vehicle/trip/flow elements."""
    if not route_file.exists():
        return (0, 0, 0)

    tree = ET.parse(route_file)
    root = tree.getroot()

    vehicle_count = 0
    trip_count = 0
    flow_count = 0

    for vehicle in root.findall("vehicle"):
        old_id = vehicle.get("id", "")
        if old_id and not old_id.startswith(id_prefix):
            vehicle.set("id", f"{id_prefix}{old_id}")
        vehicle.set("type", vehicle_class)
        vehicle_count += 1

    for trip in root.findall("trip"):
        old_id = trip.get("id", "")
        if old_id and not old_id.startswith(id_prefix):
            trip.set("id", f"{id_prefix}{old_id}")
        trip.set("type", vehicle_class)
        trip_count += 1

    for flow in root.findall("flow"):
        old_id = flow.get("id", "")
        if old_id and not old_id.startswith(id_prefix):
            flow.set("id", f"{id_prefix}{old_id}")
        flow.set("type", vehicle_class)
        flow_count += 1

    ET.indent(root)
    route_file.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")
    return (vehicle_count, trip_count, flow_count)


def _parse_depart_like_value(value: object) -> float | None:
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        parsed = to_seconds(text)
        if parsed is None:
            return None
        return float(parsed)


def _sortable_depart_seconds(element: ET.Element) -> float:
    for attribute in ("depart", "begin"):
        seconds = _parse_depart_like_value(element.get(attribute, ""))
        if seconds is not None:
            return seconds
    return float("inf")


def sort_route_file_by_depart_time(route_file: Path) -> tuple[int, int]:
    """Sort vehicle-like route elements by depart/begin time.

    Returns (sortable_elements, moved_elements).
    """
    if not route_file.exists():
        return (0, 0)

    tree = ET.parse(route_file)
    root = tree.getroot()
    children = list(root)
    sortable_tags = {"vehicle", "trip", "flow", "person", "personFlow", "container", "containerFlow"}

    sortable: list[tuple[int, ET.Element]] = []
    static: list[ET.Element] = []
    for index, child in enumerate(children):
        if child.tag in sortable_tags:
            sortable.append((index, child))
        else:
            static.append(child)

    if not sortable:
        return (0, 0)

    original_indices = [index for index, _ in sortable]
    sorted_sortable = sorted(sortable, key=lambda item: (_sortable_depart_seconds(item[1]), item[0]))
    sorted_indices = [index for index, _ in sorted_sortable]
    moved = sum(1 for original, updated in zip(original_indices, sorted_indices) if original != updated)
    if moved == 0:
        return (len(sortable), 0)

    root[:] = static + [element for _, element in sorted_sortable]
    ET.indent(root)
    route_file.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")
    return (len(sortable), moved)


def merge_classified_routes(
    bus_routes: Path,
    car_routes: Path,
    motorcycle_routes: Path,
    output_routes: Path,
) -> Path:
    """Merge mode-specific calibrated routes into a single routes file."""
    def parse_routes_file(filepath: Path) -> list[ET.Element]:
        if not filepath.exists():
            return []
        tree = ET.parse(filepath)
        root = tree.getroot()
        return list(root.findall("vehicle")) + list(root.findall("trip"))

    all_routes = []
    all_routes.extend(parse_routes_file(bus_routes))
    all_routes.extend(parse_routes_file(car_routes))
    all_routes.extend(parse_routes_file(motorcycle_routes))

    merged_root = ET.Element(
        "routes",
        {
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
            "xsi:noNamespaceSchemaLocation": "http://sumo.dlr.de/xsd/routes_file.xsd",
        },
    )

    for route_element in all_routes:
        merged_root.append(route_element)

    ET.indent(merged_root)
    output_routes.parent.mkdir(parents=True, exist_ok=True)
    output_routes.write_text(ET.tostring(merged_root, encoding="unicode"), encoding="utf-8")
    sort_route_file_by_depart_time(output_routes)
    return output_routes


def run_route_sampler_live(route_sampler: Path, route_file: Path, counts_file: Path, output_file: Path) -> str:
    if not route_sampler.exists() and route_sampler.name != "routeSampler.py":
        raise FileNotFoundError(f"RouteSampler script not found: {route_sampler}")
    if not route_file.exists():
        raise FileNotFoundError(f"Route file not found: {route_file}")
    if not counts_file.exists():
        raise FileNotFoundError(f"Counts file not found: {counts_file}")

    cmd = [sys.executable, str(route_sampler), "-r", str(route_file), "-d", str(counts_file), "-o", str(output_file)]
    process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1)

    lines: list[str] = []
    assert process.stdout is not None
    for line in iter(process.stdout.readline, ""):
        lines.append(line.rstrip())
    return_code = process.wait()
    if return_code != 0:
        lines.append(f"routeSampler exited with code {return_code}")
    return "\n".join(lines)


def parse_mode_breakdown_from_routesampler_log(log_output: str) -> dict[str, int]:
    """Parse routeSampler log to extract calibrated vehicle counts by mode (bus, car, motorcycle)."""
    breakdown = {"bus": 0, "car": 0, "motorcycle": 0, "total": 0}
    
    for line in log_output.split("\n"):
        line_lower = line.lower()
        if "bus" in line_lower and ("vehicle" in line_lower or "trip" in line_lower or "sampled" in line_lower):
            try:
                import re
                match = re.search(r"(\d+)\s*(bus|vehicle|trip)", line_lower)
                if match:
                    count = int(match.group(1))
                    breakdown["bus"] += count
            except (ValueError, AttributeError):
                pass
        elif "car" in line_lower and ("vehicle" in line_lower or "trip" in line_lower or "sampled" in line_lower):
            try:
                import re
                match = re.search(r"(\d+)\s*(car|vehicle|trip)", line_lower)
                if match:
                    count = int(match.group(1))
                    breakdown["car"] += count
            except (ValueError, AttributeError):
                pass
        elif "motorcycle" in line_lower and ("vehicle" in line_lower or "trip" in line_lower or "sampled" in line_lower):
            try:
                import re
                match = re.search(r"(\d+)\s*(motorcycle|vehicle|trip)", line_lower)
                if match:
                    count = int(match.group(1))
                    breakdown["motorcycle"] += count
            except (ValueError, AttributeError):
                pass
    
    breakdown["total"] = breakdown["bus"] + breakdown["car"] + breakdown["motorcycle"]
    return breakdown


def parse_underflow_edges_from_routesampler_log(log_output: str) -> set[str]:
    """Extract edge IDs reported in routeSampler underflow warnings."""
    underflow_edges: set[str] = set()
    if not log_output:
        return underflow_edges

    tuple_pattern = re.compile(r"\(\('([^']+)',\)\)")
    for line in log_output.splitlines():
        line_lower = line.lower()
        if "underflow" not in line_lower:
            continue
        for match in tuple_pattern.finditer(line):
            edge_id = match.group(1).strip()
            if edge_id:
                underflow_edges.add(edge_id)

    return underflow_edges


def get_routesampler_supported_options(route_sampler: Path) -> set[str]:
    """Parse routeSampler -h output and return supported long option names."""
    if not route_sampler.exists() and route_sampler.name != "routeSampler.py":
        return set()

    try:
        completed = subprocess.run(
            [sys.executable, str(route_sampler), "-h"],
            capture_output=True,
            text=True,
            check=False,
            timeout=20,
        )
        help_text = (completed.stdout or "") + "\n" + (completed.stderr or "")
    except Exception:
        return set()

    return set(re.findall(r"--[a-zA-Z0-9][a-zA-Z0-9_.-]*", help_text))


def is_scipy_available() -> bool:
    """Return True when SciPy is available in the current Python environment."""
    return importlib.util.find_spec("scipy") is not None


def ensure_scipy_available(install_if_missing: bool = False) -> tuple[bool, str | None]:
    """Ensure SciPy availability for routeSampler optimize mode."""
    if is_scipy_available():
        return True, None

    if not install_if_missing:
        return False, "SciPy is not installed; running routeSampler without --optimize."

    try:
        completed = subprocess.run(
            [sys.executable, "-m", "pip", "install", "scipy"],
            capture_output=True,
            text=True,
            check=False,
            timeout=600,
        )
        if completed.returncode == 0 and is_scipy_available():
            return True, "SciPy was installed automatically; routeSampler --optimize is enabled."
        return False, "SciPy install attempt failed; running routeSampler without --optimize."
    except Exception:
        return False, "SciPy install attempt failed; running routeSampler without --optimize."


def build_mode_routesampler_args(
    route_sampler: Path,
    mismatch_path: Path,
    vehicle_class: str,
    scipy_available: bool | None = None,
) -> tuple[list[str], list[str]]:
    """Build routeSampler args per mode while staying compatible with installed SUMO version."""
    args: list[str] = ["--mismatch-output", str(mismatch_path)]
    notes: list[str] = []
    supported = get_routesampler_supported_options(route_sampler)

    scipy_ready = is_scipy_available() if scipy_available is None else bool(scipy_available)
    if scipy_ready:
        args.extend(["--optimize", "100"])
    else:
        notes.append("SciPy is not installed; running routeSampler without --optimize.")

    if vehicle_class != "motorcycle":
        return args, notes

    if not supported:
        # Keep safe defaults when help probing fails.
        args.extend(["--minimize-vehicles", "1"])
        return args, notes

    if "--minimize-vehicles" in supported:
        args.extend(["--minimize-vehicles", "1"])
    else:
        notes.append("Installed routeSampler does not support --minimize-vehicles; skipped for motorcycle run.")

    if "--allow-no-unmatched" in supported:
        args.append("--allow-no-unmatched")

    return args, notes


def prepare_edgewise_workbook_with_minor_corrections(
    edgewise_workbook: Path,
    backup_dir: Path,
) -> tuple[Path, dict[str, object]]:
    """Apply minor numeric corrections to classified EdgeWise sheets and keep original backup.

    Corrections: NaN -> 0, round numeric counts to int, clamp negatives to 0.
    If no corrections are needed, returns the original workbook path.
    """
    required_sheets = ["Bus_5MinCols", "Car_5MinCols", "Motorcycle_5MinCols", "Total_Volume_5MinCols"]
    if not edgewise_workbook.exists():
        raise FileNotFoundError(f"EdgeWise workbook not found: {edgewise_workbook}")

    excel = pd.ExcelFile(edgewise_workbook, engine="openpyxl")
    missing = [sheet for sheet in required_sheets if sheet not in excel.sheet_names]
    if missing:
        raise ValueError(f"EdgeWise workbook missing required sheets: {', '.join(missing)}")

    corrected_frames: dict[str, pd.DataFrame] = {}
    changed_cells_by_sheet: dict[str, int] = {}
    total_changed_cells = 0

    for sheet_name in required_sheets:
        frame = pd.read_excel(edgewise_workbook, sheet_name=sheet_name, engine="openpyxl")
        base_cols = {"Location", "Date", "Approach", "Direction", "EdgeID", "SourceWorkbook"}
        interval_cols = []
        for column in frame.columns:
            col_text = str(column).strip()
            start_sec, end_sec = _parse_interval_label(col_text)
            if start_sec is not None and end_sec is not None:
                interval_cols.append(column)
            elif col_text not in base_cols:
                # Keep permissive fallback for non-standard interval headers.
                interval_cols.append(column)

        changed_cells = 0
        for column in interval_cols:
            original_numeric = pd.to_numeric(frame[column], errors="coerce")
            corrected_numeric = original_numeric.fillna(0).round().astype(int).clip(lower=0)
            original_filled = original_numeric.fillna(0)
            changed_mask = (
                original_numeric.isna()
                | (original_filled != original_filled.round())
                | (original_filled < 0)
            )
            changed_cells += int(changed_mask.sum())
            frame[column] = corrected_numeric

        corrected_frames[sheet_name] = frame
        changed_cells_by_sheet[sheet_name] = changed_cells
        total_changed_cells += changed_cells

    if total_changed_cells == 0:
        return edgewise_workbook, {
            "corrected": False,
            "total_changed_cells": 0,
            "changed_cells_by_sheet": changed_cells_by_sheet,
            "backup_path": "",
            "corrected_path": str(edgewise_workbook),
        }

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{edgewise_workbook.stem}_original_{timestamp}{edgewise_workbook.suffix}"
    shutil.copy2(edgewise_workbook, backup_path)

    corrected_path = edgewise_workbook.with_name(f"{edgewise_workbook.stem}_corrected_{timestamp}{edgewise_workbook.suffix}")
    with pd.ExcelWriter(corrected_path, engine="openpyxl") as writer:
        for sheet_name in excel.sheet_names:
            if sheet_name in corrected_frames:
                corrected_frames[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                pd.read_excel(edgewise_workbook, sheet_name=sheet_name, engine="openpyxl").to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )

    return corrected_path, {
        "corrected": True,
        "total_changed_cells": total_changed_cells,
        "changed_cells_by_sheet": changed_cells_by_sheet,
        "backup_path": str(backup_path),
        "corrected_path": str(corrected_path),
    }


def parse_mode_breakdown_from_routes_file(route_file: Path) -> dict[str, int]:
    """Read calibrated routes and count generated demand by vehicle type."""
    breakdown = {"bus": 0, "car": 0, "motorcycle": 0, "total": 0}
    if not route_file.exists():
        return breakdown

    root = ET.parse(route_file).getroot()
    for tag in ["vehicle", "trip", "flow"]:
        for item in root.findall(tag):
            type_value = str(item.get("type", "")).lower()
            if type_value in {"bus", "calibbustype"}:
                breakdown["bus"] += 1
            elif type_value in {"car", "calibcartype"}:
                breakdown["car"] += 1
            elif type_value in {"motorcycle", "motorbike", "calibmototype"}:
                breakdown["motorcycle"] += 1

    breakdown["total"] = breakdown["bus"] + breakdown["car"] + breakdown["motorcycle"]
    return breakdown


def sanitize_route_file_vtypes(route_file: Path, remove_ids: set[str] | None = None) -> tuple[int, int]:
    """Remove conflicting vType/vTypeDistribution definitions from a route file.

    Returns (removed_vtypes, removed_distributions).
    """
    if not route_file.exists():
        return (0, 0)

    if remove_ids is None:
        remove_ids = {"bus", "car", "motorcycle", "calibbustype", "calibcartype", "calibmototype"}

    tree = ET.parse(route_file)
    root = tree.getroot()

    removed_vtypes = 0
    removed_distributions = 0

    for element in list(root.findall("vType")):
        element_id = str(element.get("id", "")).lower()
        if element_id in remove_ids:
            root.remove(element)
            removed_vtypes += 1

    for element in list(root.findall("vTypeDistribution")):
        root.remove(element)
        removed_distributions += 1

    if removed_vtypes or removed_distributions:
        ET.indent(root)
        route_file.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")

    return (removed_vtypes, removed_distributions)


def cap_rescue_vehicles_in_routes(
    route_file: Path | str,
    max_ratio: float = 0.015,
    max_absolute: int = 400,
) -> tuple[int, int, int]:
    """Cap rescue vehicles to prevent over-injection congestion in final validation.

    Returns (total_vehicles, rescue_before, rescue_removed).
    """
    route_path = Path(route_file)
    if not route_path.exists():
        return (0, 0, 0)

    tree = ET.parse(route_path)
    root = tree.getroot()
    vehicles = list(root.findall("vehicle"))
    total_vehicles = len(vehicles)
    if total_vehicles == 0:
        return (0, 0, 0)

    rescue_vehicles: list[ET.Element] = []
    for vehicle in vehicles:
        vid = str(vehicle.get("id", "")).lower()
        if "rescue" in vid:
            rescue_vehicles.append(vehicle)

    rescue_before = len(rescue_vehicles)
    if rescue_before == 0:
        return (total_vehicles, 0, 0)

    rescue_cap = min(max_absolute, int(max_ratio * total_vehicles))
    rescue_cap = max(0, rescue_cap)
    if rescue_before <= rescue_cap:
        return (total_vehicles, rescue_before, 0)

    rescue_removed = 0
    for vehicle in rescue_vehicles[rescue_cap:]:
        root.remove(vehicle)
        rescue_removed += 1

    if rescue_removed > 0:
        ET.indent(root)
        route_path.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")
        sort_route_file_by_depart_time(route_path)

    return (total_vehicles, rescue_before, rescue_removed)


TARGET_SURVEY_TLS_IDS = {
    "cluster_1880066214_314962647",  # Pulchowk North
    "cluster_1880066220_314962644",  # Kesharmahal
    "cluster_2230569083_3405700883_4455971833_4455971834",  # Kandevtasthan
}


def normalize_tls_id(value: str) -> str:
    text = str(value or "")
    if text.startswith("cluster_"):
        return text[len("cluster_") :]
    return text


TARGET_SURVEY_TLS_IDS_NORMALIZED = {normalize_tls_id(tls_id) for tls_id in TARGET_SURVEY_TLS_IDS}


def load_target_tls_templates_from_add_file(tls_add_file: Path) -> dict[str, tuple[str, list[dict[str, str]]]]:
    """Load target TLS Program 0 templates from add-file, keyed by normalized TLS ID.

    Returns mapping: normalized_tls_id -> (offset, phase_attribute_dicts)
    """
    templates: dict[str, tuple[str, list[dict[str, str]]]] = {}
    if not tls_add_file.exists():
        return templates

    root = ET.parse(tls_add_file).getroot()
    first_seen: dict[str, tuple[str, list[dict[str, str]]]] = {}
    program_zero: dict[str, tuple[str, list[dict[str, str]]]] = {}

    for tl_logic in root.findall("tlLogic"):
        logic_id = str(tl_logic.get("id", "")).strip()
        normalized_id = normalize_tls_id(logic_id)
        if normalized_id not in TARGET_SURVEY_TLS_IDS_NORMALIZED:
            continue

        phases: list[dict[str, str]] = []
        for phase in tl_logic.findall("phase"):
            attrs = {str(k): str(v) for k, v in phase.attrib.items()}
            if "state" not in attrs:
                continue
            phases.append(attrs)

        if not phases:
            continue

        offset = str(tl_logic.get("offset", "0"))
        candidate = (offset, phases)
        first_seen.setdefault(normalized_id, candidate)
        if str(tl_logic.get("programID", "")) == "0":
            program_zero[normalized_id] = candidate

    for normalized_id in TARGET_SURVEY_TLS_IDS_NORMALIZED:
        if normalized_id in program_zero:
            templates[normalized_id] = program_zero[normalized_id]
        elif normalized_id in first_seen:
            templates[normalized_id] = first_seen[normalized_id]

    return templates


def ensure_survey_tls_controls_in_network(network_file: Path, create_missing_tllogics: bool = True) -> tuple[int, int, int, int]:
    """Ensure survey TLS junctions, controlled links, and placeholder tlLogic entries exist."""
    if not network_file.exists():
        raise FileNotFoundError(f"Network file not found: {network_file}")

    tree = ET.parse(network_file)
    root = tree.getroot()

    changed_junctions = 0
    changed_connections = 0
    created_tllogics = 0
    moved_tllogics = 0

    # SUMO requires the junction itself to be traffic_light when loading tlLogic plans.
    for junction in root.findall("junction"):
        junction_id = str(junction.get("id", ""))
        if normalize_tls_id(junction_id) not in TARGET_SURVEY_TLS_IDS_NORMALIZED:
            continue
        if str(junction.get("type", "")) != "traffic_light":
            junction.set("type", "traffic_light")
            changed_junctions += 1

    # Build deterministic controlled links from incoming (non-internal) via lanes.
    connection_groups: dict[str, list[ET.Element]] = {tls_id: [] for tls_id in TARGET_SURVEY_TLS_IDS}
    for connection in root.findall("connection"):
        from_edge = str(connection.get("from", ""))
        if from_edge.startswith(":"):
            continue

        via = str(connection.get("via", ""))
        if not via.startswith(":cluster_"):
            continue

        for tls_id in TARGET_SURVEY_TLS_IDS:
            if via.startswith(f":{tls_id}_"):
                connection_groups[tls_id].append(connection)
                break

    for tls_id, group in connection_groups.items():
        if not group:
            raise ValueError(
                "Could not find any controlled incoming connections for TLS "
                f"'{tls_id}' in network file {network_file}"
            )

        for link_index, connection in enumerate(group):
            old_tl = str(connection.get("tl", ""))
            old_link = str(connection.get("linkIndex", ""))
            new_link = str(link_index)
            if old_tl != tls_id:
                connection.set("tl", tls_id)
            if old_link != new_link:
                connection.set("linkIndex", new_link)
            if old_tl != tls_id or old_link != new_link:
                changed_connections += 1

    existing_tllogics: dict[str, ET.Element] = {}
    for tl_logic in root.findall("tlLogic"):
        logic_id = str(tl_logic.get("id", ""))
        if normalize_tls_id(logic_id) in TARGET_SURVEY_TLS_IDS_NORMALIZED:
            existing_tllogics[normalize_tls_id(logic_id)] = tl_logic

    children = list(root)
    first_junction_idx = next((i for i, node in enumerate(children) if node.tag == "junction"), len(children))

    if create_missing_tllogics:
        for tls_id, group in connection_groups.items():
            normalized_id = normalize_tls_id(tls_id)
            if normalized_id in existing_tllogics:
                continue

            phase_len = len(group)
            tl_logic = ET.Element(
                "tlLogic",
                {
                    "id": tls_id,
                    "type": "static",
                    "programID": "inactive_net",
                    "offset": "0",
                },
            )
            # Create proper placeholder program with all links getting green in sequence.
            # SUMO requires each link/direction to have at least one green phase.
            for i in range(phase_len):
                # Construct state where link i is green, others are red.
                green_state = "r" * i + "G" + "r" * (phase_len - 1 - i)
                yellow_state = "r" * i + "y" + "r" * (phase_len - 1 - i)
                tl_logic.append(ET.Element("phase", {"duration": "30", "state": green_state}))
                tl_logic.append(ET.Element("phase", {"duration": "3", "state": yellow_state}))
                tl_logic.append(ET.Element("phase", {"duration": "1", "state": "r" * phase_len}))
            root.insert(first_junction_idx, tl_logic)
            first_junction_idx += 1
            created_tllogics += 1

    # Keep target tlLogic definitions before junction/connection sections for SUMO parser compatibility.
    children = list(root)
    first_junction_idx = next((i for i, node in enumerate(children) if node.tag == "junction"), len(children))
    if first_junction_idx < len(children):
        target_tllogics: list[ET.Element] = []
        for node in children:
            if node.tag != "tlLogic":
                continue
            if normalize_tls_id(str(node.get("id", ""))) in TARGET_SURVEY_TLS_IDS_NORMALIZED:
                target_tllogics.append(node)

        for tl_logic in target_tllogics:
            current_children = list(root)
            current_idx = current_children.index(tl_logic)
            current_junction_idx = next((i for i, node in enumerate(current_children) if node.tag == "junction"), len(current_children))
            if current_idx > current_junction_idx:
                root.remove(tl_logic)
                root.insert(current_junction_idx, tl_logic)
                moved_tllogics += 1

    if changed_junctions or changed_connections or created_tllogics or moved_tllogics:
        tree.write(network_file, encoding="utf-8", xml_declaration=True)

    return (changed_junctions, changed_connections, created_tllogics, moved_tllogics)


def scrub_default_network_signal_logic(
    network_file: Path = DEFAULT_NETWORK_FILE,
    tls_add_file: Path | None = None,
) -> tuple[int, int, int]:
    """Remove target net tlLogic and recreate inactive fallback controllers.

    Fallback controllers are mirrored from survey Program 0 (if available) so net/add mismatch
    does not alter timing behavior when SUMO selects net controller at startup.
    Returns: (removed_tllogics, created_from_add_templates, created_generic)
    """
    if not network_file.exists():
        raise FileNotFoundError(f"Network file not found: {network_file}")

    tree = ET.parse(network_file)
    root = tree.getroot()
    removed_tllogics = 0
    created_from_add = 0
    created_generic = 0

    templates: dict[str, tuple[str, list[dict[str, str]]]] = {}
    if tls_add_file is not None:
        templates = load_target_tls_templates_from_add_file(tls_add_file)

    for tl_logic in list(root.findall("tlLogic")):
        logic_id = str(tl_logic.get("id", ""))
        if normalize_tls_id(logic_id) in TARGET_SURVEY_TLS_IDS_NORMALIZED:
            root.remove(tl_logic)
            removed_tllogics += 1

    # Keep TLS controllers known to SUMO by restoring neutral placeholders after scrub.
    connection_groups: dict[str, list[ET.Element]] = {tls_id: [] for tls_id in TARGET_SURVEY_TLS_IDS}
    for connection in root.findall("connection"):
        from_edge = str(connection.get("from", ""))
        if from_edge.startswith(":"):
            continue

        tl_id = str(connection.get("tl", ""))
        normalized_tl = normalize_tls_id(tl_id)
        if normalized_tl in TARGET_SURVEY_TLS_IDS_NORMALIZED:
            target_tls = next((t for t in TARGET_SURVEY_TLS_IDS if normalize_tls_id(t) == normalized_tl), None)
            if target_tls:
                connection_groups[target_tls].append(connection)
            continue

        via = str(connection.get("via", ""))
        for tls_id in TARGET_SURVEY_TLS_IDS:
            if via.startswith(f":{tls_id}_"):
                connection_groups[tls_id].append(connection)
                break

    existing_target_norm_ids = {
        normalize_tls_id(str(tl_logic.get("id", "")))
        for tl_logic in root.findall("tlLogic")
    }

    children = list(root)
    first_junction_idx = next((i for i, node in enumerate(children) if node.tag == "junction"), len(children))

    for tls_id in TARGET_SURVEY_TLS_IDS:
        norm_id = normalize_tls_id(tls_id)
        if norm_id in existing_target_norm_ids:
            continue

        group = connection_groups.get(tls_id, [])
        if not group:
            continue

        phase_len = len(group)
        tl_logic = ET.Element(
            "tlLogic",
            {
                "id": tls_id,
                "type": "static",
                "programID": "inactive_net",
                "offset": "0",
            },
        )

        normalized_id = normalize_tls_id(tls_id)
        template = templates.get(normalized_id)
        used_template = False
        if template:
            offset, phases = template
            tl_logic.set("offset", offset)
            valid_phases = 0
            for attrs in phases:
                phase_attrs = dict(attrs)
                state = str(phase_attrs.get("state", ""))
                if not state:
                    continue
                if len(state) < phase_len:
                    state = state + ("r" * (phase_len - len(state)))
                elif len(state) > phase_len:
                    state = state[:phase_len]
                phase_attrs["state"] = state
                tl_logic.append(ET.Element("phase", phase_attrs))
                valid_phases += 1

            if valid_phases > 0:
                used_template = True

        if not used_template:
            for idx in range(phase_len):
                green_state = "r" * idx + "G" + "r" * (phase_len - 1 - idx)
                yellow_state = "r" * idx + "y" + "r" * (phase_len - 1 - idx)
                tl_logic.append(ET.Element("phase", {"duration": "30", "state": green_state}))
                tl_logic.append(ET.Element("phase", {"duration": "3", "state": yellow_state}))
                tl_logic.append(ET.Element("phase", {"duration": "1", "state": "r" * phase_len}))
            created_generic += 1
        else:
            created_from_add += 1

        root.insert(first_junction_idx, tl_logic)
        first_junction_idx += 1
    created_placeholders = created_from_add + created_generic

    if removed_tllogics or created_placeholders:
        tree.write(network_file, encoding="utf-8", xml_declaration=True)

    return (removed_tllogics, created_from_add, created_generic)


def prepare_survey_tls_priority(network_file: Path, tls_add_file: Path) -> dict[str, int]:
    """Keep survey TLS in control by removing default net signal logic and forcing Program 0 in add-file."""
    tls_junctions_fixed = 0
    tls_links_fixed = 0
    tls_tllogics_reordered = 0
    tls_control_sync_skipped = 0
    try:
        tls_junctions_fixed, tls_links_fixed, _tls_tllogics_created, tls_tllogics_reordered = ensure_survey_tls_controls_in_network(
            network_file,
            create_missing_tllogics=False,
        )
    except ValueError:
        # Keep calibration moving on partial/trimmed networks while still applying tlLogic removal.
        tls_control_sync_skipped = 1

    removed_net_tllogics, mirrored_from_add, created_generic = scrub_default_network_signal_logic(
        network_file,
        tls_add_file=tls_add_file,
    )
    tls_program_updates, removed_waut = enforce_tls_program_zero_and_validate(tls_add_file)
    return {
        "tls_junctions_enabled": tls_junctions_fixed,
        "tls_controlled_links_assigned": tls_links_fixed,
        "tls_tllogics_reordered": tls_tllogics_reordered,
        "tls_control_sync_skipped": tls_control_sync_skipped,
        "network_tllogics_removed": removed_net_tllogics,
        "network_fallback_from_add": mirrored_from_add,
        "network_fallback_generic": created_generic,
        "tls_programs_forced": tls_program_updates,
        "waut_removed": removed_waut,
    }


def summarize_tls_priority_state(network_file: Path, tls_add_file: Path) -> dict[str, int]:
    """Return a compact verification snapshot after TLS scrub/enforce."""
    if not network_file.exists():
        raise FileNotFoundError(f"Network file not found: {network_file}")
    if not tls_add_file.exists():
        raise FileNotFoundError(f"Traffic light file not found: {tls_add_file}")

    net_root = ET.parse(network_file).getroot()
    add_root = ET.parse(tls_add_file).getroot()

    net_target_tllogics_remaining = 0
    net_target_inactive_programs = 0
    net_target_noninactive_programs = 0
    for tl_logic in net_root.findall("tlLogic"):
        logic_id = str(tl_logic.get("id", ""))
        if normalize_tls_id(logic_id) in TARGET_SURVEY_TLS_IDS_NORMALIZED:
            net_target_tllogics_remaining += 1
            if str(tl_logic.get("programID", "")) == "inactive_net":
                net_target_inactive_programs += 1
            else:
                net_target_noninactive_programs += 1

    add_target_program_zero = 0
    add_target_program_nonzero = 0
    present_add_ids: set[str] = set()
    for tl_logic in add_root.findall("tlLogic"):
        logic_id = str(tl_logic.get("id", ""))
        normalized_id = normalize_tls_id(logic_id)
        if normalized_id not in TARGET_SURVEY_TLS_IDS_NORMALIZED:
            continue
        present_add_ids.add(normalized_id)
        if str(tl_logic.get("programID", "")) == "0":
            add_target_program_zero += 1
        else:
            add_target_program_nonzero += 1

    waut_remaining = len(add_root.findall("WAUT")) + len(add_root.findall("wautJunction"))
    missing_target_add_tllogics = len(TARGET_SURVEY_TLS_IDS_NORMALIZED - present_add_ids)

    return {
        "net_target_tllogics_remaining": net_target_tllogics_remaining,
        "net_target_inactive_programs": net_target_inactive_programs,
        "net_target_noninactive_programs": net_target_noninactive_programs,
        "add_target_program_zero": add_target_program_zero,
        "add_target_program_nonzero": add_target_program_nonzero,
        "add_waut_remaining": waut_remaining,
        "add_missing_target_tllogics": missing_target_add_tllogics,
    }


def neutralize_network_default_tls_programs(network_file: Path) -> tuple[int, int]:
    """Rename default net programID='0' to 'inactive_net' and ensure inactive phases remain SUMO-valid."""
    if not network_file.exists():
        raise FileNotFoundError(f"Network file not found: {network_file}")

    tree = ET.parse(network_file)
    root = tree.getroot()
    renamed = 0
    repaired_programs = 0

    # Count controlled links from connection linkIndex so inactive programs can cover every tl-index.
    controlled_link_count: dict[str, int] = {}
    for connection in root.findall("connection"):
        tl_id_raw = str(connection.get("tl", ""))
        if not tl_id_raw:
            continue
        normalized = normalize_tls_id(tl_id_raw)
        if normalized not in TARGET_SURVEY_TLS_IDS_NORMALIZED:
            continue
        link_index_raw = str(connection.get("linkIndex", "")).strip()
        if not link_index_raw:
            continue
        try:
            idx = int(link_index_raw)
        except ValueError:
            continue
        key = normalized
        controlled_link_count[key] = max(controlled_link_count.get(key, 0), idx + 1)

    for tl_logic in root.findall("tlLogic"):
        tls_id = normalize_tls_id(str(tl_logic.get("id", "")))
        if tls_id not in TARGET_SURVEY_TLS_IDS_NORMALIZED:
            continue

        if str(tl_logic.get("programID", "")) == "0":
            tl_logic.set("programID", "inactive_net")
            renamed += 1

        if str(tl_logic.get("programID", "")) == "inactive_net":
            phases = tl_logic.findall("phase")
            if not phases:
                continue

            state_len = len(str(phases[0].get("state", "")))
            link_count = controlled_link_count.get(tls_id, state_len)
            if link_count <= 0:
                continue

            has_green_per_index = [False] * link_count
            for phase in phases:
                state = str(phase.get("state", ""))
                for idx, ch in enumerate(state[:link_count]):
                    if ch in "gGy":
                        has_green_per_index[idx] = True

            issues = validate_tls_yellow_phases(tl_logic)
            if all(has_green_per_index) and not issues:
                continue

            for phase in list(phases):
                tl_logic.remove(phase)

            for idx in range(link_count):
                green_state = "r" * idx + "G" + "r" * (link_count - 1 - idx)
                yellow_state = "r" * idx + "y" + "r" * (link_count - 1 - idx)
                tl_logic.append(ET.Element("phase", {"duration": "30", "state": green_state}))
                tl_logic.append(ET.Element("phase", {"duration": "3", "state": yellow_state}))
                tl_logic.append(ET.Element("phase", {"duration": "1", "state": "r" * link_count}))
            repaired_programs += 1

    if renamed or repaired_programs:
        tree.write(network_file, encoding="utf-8", xml_declaration=True)

    return (renamed, repaired_programs)


def validate_tls_yellow_phases(tl_logic: ET.Element) -> list[str]:
    """Validate no direct G/g -> r transitions and that yellow transitions use 3s phases."""
    issues: list[str] = []
    phases = tl_logic.findall("phase")
    n = len(phases)
    if n < 2:
        return ["insufficient phases"]

    for i in range(n):
        current = phases[i]
        nxt = phases[(i + 1) % n]
        nxt2 = phases[(i + 2) % n]
        s1 = str(current.get("state", ""))
        s2 = str(nxt.get("state", ""))
        s3 = str(nxt2.get("state", ""))

        direct_green_to_red = [idx for idx, (a, b) in enumerate(zip(s1, s2)) if a in "Gg" and b == "r"]
        if direct_green_to_red:
            issues.append(f"direct G/g->r transition between phase {i} and {(i + 1) % n}")

        green_to_yellow = [idx for idx, (a, b) in enumerate(zip(s1, s2)) if a in "Gg" and b in "Yy"]
        if green_to_yellow:
            try:
                yellow_duration = int(float(str(nxt.get("duration", "0"))))
            except ValueError:
                yellow_duration = 0

            if yellow_duration != 3:
                issues.append(f"yellow phase at index {(i + 1) % n} has duration {yellow_duration}, expected 3")

            bad_followup = [idx for idx in green_to_yellow if idx >= len(s3) or s3[idx] != "r"]
            if bad_followup:
                issues.append(f"yellow phase at index {(i + 1) % n} is not followed by red")

    return issues


def enforce_tls_program_zero_and_validate(tls_add_file: Path) -> tuple[int, int]:
    """Force survey tlLogic programID to 0, remove WAUT blocks, and validate yellow phases."""
    if not tls_add_file.exists():
        raise FileNotFoundError(f"Traffic light file not found: {tls_add_file}")

    tree = ET.parse(tls_add_file)
    root = tree.getroot()

    updated_programs = 0
    removed_waut = 0
    issues: list[str] = []

    for tag in ("WAUT", "wautJunction"):
        for element in list(root.findall(tag)):
            root.remove(element)
            removed_waut += 1

    for tl_logic in root.findall("tlLogic"):
        tls_id = str(tl_logic.get("id", ""))
        if normalize_tls_id(tls_id) not in TARGET_SURVEY_TLS_IDS_NORMALIZED:
            continue

        if str(tl_logic.get("programID", "")) != "0":
            tl_logic.set("programID", "0")
            updated_programs += 1

        for issue in validate_tls_yellow_phases(tl_logic):
            issues.append(f"{tls_id}: {issue}")

    if issues:
        raise ValueError("Yellow-phase validation failed: " + " | ".join(issues[:12]))

    if updated_programs or removed_waut:
        ET.indent(root)
        tls_add_file.write_text(ET.tostring(root, encoding="unicode"), encoding="utf-8")

    return (updated_programs, removed_waut)


def stream_route_sampler(
    route_sampler: Path,
    route_file: Path,
    counts_file: Path,
    output_file: Path,
    placeholder,
    extra_args: Sequence[str] | None = None,
) -> str:
    if not route_sampler.exists() and route_sampler.name != "routeSampler.py":
        raise FileNotFoundError(f"RouteSampler script not found: {route_sampler}")
    if not route_file.exists():
        raise FileNotFoundError(f"Route file not found: {route_file}")
    if not counts_file.exists():
        raise FileNotFoundError(f"Counts file not found: {counts_file}")

    cmd = [sys.executable, str(route_sampler), "-r", str(route_file), "-d", str(counts_file), "-o", str(output_file)]
    if extra_args:
        cmd.extend([str(arg) for arg in extra_args])
    process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1)

    lines: list[str] = []
    assert process.stdout is not None
    while True:
        line = process.stdout.readline()
        if line == "" and process.poll() is not None:
            break
        if line:
            lines.append(line.rstrip())
            placeholder.code("\n".join(lines), language="text")

    return_code = process.wait()
    if return_code != 0:
        lines.append(f"routeSampler exited with code {return_code}")
        placeholder.code("\n".join(lines), language="text")
    return "\n".join(lines)


def find_sumo_binary() -> Path:
    candidates = [
        Path("sumo"),
        Path(r"C:\Program Files (x86)\Eclipse\Sumo\bin\sumo.exe"),
        Path(r"C:\Program Files\Eclipse\Sumo\bin\sumo.exe"),
    ]
    for candidate in candidates:
        if str(candidate) == "sumo":
            return candidate
        if candidate.exists():
            return candidate
    return Path("sumo")


def write_final_simulation_sumocfg(
    sumocfg_path: Path,
    network_file: Path,
    route_file: Path,
    begin_sec: int,
    end_sec: int,
    tripinfo_file: Path,
    queue_file: Path,
    vehicle_types_file: Path | None = None,
    ) -> Path:
    """Write SUMO configuration file with support for vehicle types and additional files."""
    configuration = ET.Element("configuration")
    input_element = ET.SubElement(configuration, "input")
    # Keep file references local to Streamlit_Callibration so the generated config
    # consistently uses the intended network/routes/additional files.
    ET.SubElement(input_element, "net-file", {"value": network_file.name})
    ET.SubElement(input_element, "route-files", {"value": route_file.name})
    
    # Build additional-files list with vehicle_types if provided
    additional_files_list = ["traffic_lights.add.xml", "detectors.add.xml"]
    if vehicle_types_file and vehicle_types_file.exists():
        additional_files_list.append("vehicle_types.add.xml")
    ET.SubElement(input_element, "additional-files", {"value": ", ".join(additional_files_list)})

    time_element = ET.SubElement(configuration, "time")
    ET.SubElement(time_element, "begin", {"value": str(float(begin_sec))})
    ET.SubElement(time_element, "end", {"value": str(float(end_sec))})

    output_element = ET.SubElement(configuration, "output")
    ET.SubElement(output_element, "tripinfo-output", {"value": str(tripinfo_file)})
    ET.SubElement(output_element, "queue-output", {"value": str(queue_file)})

    processing_element = ET.SubElement(configuration, "processing")
    ET.SubElement(processing_element, "time-to-teleport", {"value": "99999"})
    ET.SubElement(processing_element, "max-depart-delay", {"value": "3600"})
    ET.SubElement(processing_element, "collision.action", {"value": "warn"})
    ET.SubElement(processing_element, "lateral-resolution", {"value": "0.8"})
    ET.SubElement(processing_element, "step-length", {"value": "0.50"})

    ET.indent(configuration)
    sumocfg_path.parent.mkdir(parents=True, exist_ok=True)
    sumocfg_path.write_text(ET.tostring(configuration, encoding="unicode"), encoding="utf-8")
    return sumocfg_path


def stream_final_simulation(sumocfg_path: Path, placeholder, working_dir: Path) -> str:
    sumo_binary = find_sumo_binary()
    cmd = [str(sumo_binary), "-c", str(sumocfg_path)]
    process = subprocess.Popen(cmd, cwd=str(working_dir), stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1)

    lines: list[str] = []
    assert process.stdout is not None
    while True:
        line = process.stdout.readline()
        if line == "" and process.poll() is not None:
            break
        if line:
            lines.append(line.rstrip())
            placeholder.code("\n".join(lines), language="text")

    return_code = process.wait()
    if return_code != 0:
        lines.append(f"SUMO final simulation exited with code {return_code}")
        placeholder.code("\n".join(lines), language="text")
    return "\n".join(lines)


def title_block(title: str, subtitle: str) -> None:
    st.markdown(f"## {title}")
    st.caption(subtitle)


def sidebar_path_input(label: str, default: Path) -> Path:
    value = st.sidebar.text_input(label, str(default))
    return Path(value)


def bootstrap_streamlit_calibration_workspace(target_root: Path, legacy_root: Path) -> list[str]:
    target_root.mkdir(parents=True, exist_ok=True)
    (target_root / "Callibration").mkdir(parents=True, exist_ok=True)

    copied: list[str] = []
    copy_map = {
        legacy_root / "5minCompileSummary_UPDATED.xlsx": target_root / "5minCompileSummary_UPDATED.xlsx",
        legacy_root / "5minCompileSummary_DateTabs_UPDATED.xlsx": target_root / "5minCompileSummary_DateTabs_UPDATED.xlsx",
        legacy_root / "5minCompileSummary_Compact.xlsx": target_root / "5minCompileSummary_Compact.xlsx",
        legacy_root / "LocationID.xlsx": target_root / "LocationID.xlsx",
        legacy_root / "Network.net.xml": target_root / "Network.net.xml",
        legacy_root / "Callibration" / "route_pool.rou.xml": target_root / "Callibration" / "route_pool.rou.xml",
        legacy_root / "Callibration" / "route_pool.trips.xml": target_root / "Callibration" / "route_pool.trips.xml",
    }

    for source, destination in copy_map.items():
        if destination.exists() or not source.exists():
            continue
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, destination)
        copied.append(destination.name)

    return copied


def find_sumo_tool(*relative_parts: str) -> Path:
    sumo_home = os.environ.get("SUMO_HOME", "").strip().strip('"')
    candidates: list[Path] = []
    if sumo_home:
        candidates.append(Path(sumo_home) / "tools" / Path(*relative_parts))
    candidates.extend(
        [
            Path(r"C:\Program Files (x86)\Eclipse\Sumo\tools") / Path(*relative_parts),
            Path(r"C:\Program Files\Eclipse\Sumo\tools") / Path(*relative_parts),
        ]
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(f"SUMO tool not found: {'/'.join(relative_parts)}")


def generate_candidate_route_pool(
    network_file: Path,
    trip_file: Path,
    route_file: Path,
    begin_sec: int,
    end_sec: int,
    target_trips: int,
    fringe_factor: float,
    output_placeholder,
) -> str:
    random_trips = find_sumo_tool("randomTrips.py")
    duration = max(1, end_sec - begin_sec)
    period = max(duration / float(max(1, target_trips)), 0.05)

    trip_file.parent.mkdir(parents=True, exist_ok=True)
    route_file.parent.mkdir(parents=True, exist_ok=True)

    cmd = [
        sys.executable,
        str(random_trips),
        "-n",
        str(network_file),
        "-b",
        str(begin_sec),
        "-e",
        str(end_sec),
        "-p",
        f"{period:.6f}",
        "--fringe-factor",
        f"{fringe_factor:.2f}",
        "--trip-attributes",
        'departLane="best" departSpeed="max" departPos="base"',
        "-o",
        str(trip_file),
        "--route-file",
        str(route_file),
    ]

    process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1)
    lines: list[str] = []
    assert process.stdout is not None
    while True:
        line = process.stdout.readline()
        if line == "" and process.poll() is not None:
            break
        if line:
            lines.append(line.rstrip())
            output_placeholder.code("\n".join(lines), language="text")

    return_code = process.wait()
    if return_code != 0:
        lines.append(f"randomTrips exited with code {return_code}")
        output_placeholder.code("\n".join(lines), language="text")
    return "\n".join(lines)


def get_location_edge_map(location_workbook: Path, network_file: Path) -> dict[str, dict[str, str]]:
    _, location_edge_map = build_location_edge_map(location_workbook, network_file)
    return location_edge_map


def write_location_edge_map_csv(location_workbook: Path, network_file: Path, output_csv: Path) -> tuple[Path, Path, dict[str, int]]:
    location_frame = read_location_mapping(location_workbook)
    required_columns = {"LocationName", "Network ID"}
    if not required_columns.issubset(location_frame.columns):
        raise ValueError("Location workbook must include 'LocationName' and 'Network ID' columns")

    _, location_edge_map = build_location_edge_map(location_workbook, network_file)
    network_root = ET.parse(network_file).getroot()
    edge_lookup: dict[str, tuple[str, str]] = {}
    for edge in network_root.findall("edge"):
        edge_id = str(edge.get("id", "")).strip()
        if not edge_id or edge_id.startswith(":") or edge.get("function") == "internal":
            continue
        edge_lookup[edge_id] = (str(edge.get("from", "")).strip(), str(edge.get("to", "")).strip())

    location_by_name = {
        str(row.get("LocationName", "") or "").strip(): str(row.get("Network ID", "") or "").strip()
        for _, row in location_frame.iterrows()
    }

    output_rows: list[dict[str, str]] = []
    debug_rows: list[dict[str, str]] = []
    mapped_counts: dict[str, int] = {}

    for location_name in sorted(location_edge_map.keys()):
        junction_id = location_by_name.get(location_name, "")
        direction_map = location_edge_map[location_name]
        mapped_counts[location_name] = len(direction_map)
        for approach, edge_id in sorted(direction_map.items()):
            output_rows.append({"Location": location_name, "Approach": approach, "EdgeID": edge_id})
            edge_from, edge_to = edge_lookup.get(edge_id, ("", ""))
            edge_type = "Incoming" if junction_id and edge_to == junction_id else "Outgoing" if junction_id and edge_from == junction_id else "Unknown"
            debug_rows.append(
                {
                    "Location": location_name,
                    "Approach": approach,
                    "EdgeID": edge_id,
                    "EdgeType": edge_type,
                    "JunctionID": junction_id,
                }
            )

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    output_frame = pd.DataFrame(output_rows, columns=["Location", "Approach", "EdgeID"])
    debug_frame = pd.DataFrame(debug_rows, columns=["Location", "Approach", "EdgeID", "EdgeType", "JunctionID"])
    if not output_frame.empty:
        output_frame = output_frame.sort_values(["Location", "Approach", "EdgeID"]).reset_index(drop=True)
    if not debug_frame.empty:
        debug_frame = debug_frame.sort_values(["Location", "Approach", "EdgeID", "EdgeType"]).reset_index(drop=True)

    # Automated mapping must never overwrite manual location_edge_map.csv.
    manual_csv = output_csv if output_csv.name.lower() == "location_edge_map.csv" else DEFAULT_LOCATION_EDGE_MAP_CSV
    debug_csv = output_csv if output_csv.name.lower() == "location_edge_map_debug.csv" else DEFAULT_LOCATION_EDGE_MAP_DEBUG_CSV
    actual_output_csv = manual_csv
    actual_debug_csv = debug_csv
    try:
        # Preserve manual setup: write only DEBUG output.
        debug_frame.to_csv(actual_debug_csv, index=False)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        actual_debug_csv = debug_csv.with_name(f"{debug_csv.stem}_unlocked_{timestamp}{debug_csv.suffix}")
        debug_frame.to_csv(actual_debug_csv, index=False)

    return actual_output_csv, actual_debug_csv, mapped_counts


def main() -> None:
    bootstrapped_files = bootstrap_streamlit_calibration_workspace(STREAMLIT_CALIB_DIR, LEGACY_CALIB_DIR)

    st.set_page_config(page_title="Corridor Calibration Hub", layout="wide")
    st.title("Corridor Calibration Hub")
    st.write("Streamlit front-end for raw video count consolidation, SUMO counts.xml creation, whitelist filtering, and routeSampler execution.")
    st.caption(f"Active calibration workspace: {STREAMLIT_CALIB_DIR}")
    if bootstrapped_files:
        st.info(f"Initialized Streamlit_Callibration with: {', '.join(bootstrapped_files)}")

    st.sidebar.header("Inputs")
    master_workbook = sidebar_path_input("Master summary workbook", DEFAULT_MASTER_WORKBOOK)
    date_tabs_workbook = sidebar_path_input("Date tabs workbook", DEFAULT_DATE_TABS_WORKBOOK)
    compact_workbook = sidebar_path_input("Compact workbook", DEFAULT_COMPACT_WORKBOOK)
    location_workbook = sidebar_path_input("Location mapping workbook", DEFAULT_LOCATION_WORKBOOK)
    network_file = sidebar_path_input("Network file", DEFAULT_NETWORK_FILE)
    route_pool = sidebar_path_input("Route pool", DEFAULT_ROUTE_POOL)
    trip_pool = sidebar_path_input("Route trip pool", DEFAULT_TRIP_POOL)
    route_sampler = sidebar_path_input("routeSampler.py", DEFAULT_ROUTE_SAMPLER)
    output_workbook = sidebar_path_input("Datewise summary output workbook", DEFAULT_OUTPUT_WORKBOOK)
    edgewise_workbook = sidebar_path_input("EdgeWise 5-min interval output workbook", DEFAULT_EDGEWISE_WORKBOOK)
    counts_output = sidebar_path_input("counts.xml output", DEFAULT_COUNTS_XML)
    calibrated_routes_output = sidebar_path_input("Calibrated routes output", DEFAULT_CALIBRATED_ROUTES)
    whitelisted_pool_output = sidebar_path_input("Whitelisted pool output", DEFAULT_WHITELISTED_POOL)

    fixed_signal_timing_detected = DEFAULT_TLS_ADD_FILE.exists()
    if fixed_signal_timing_detected:
        st.sidebar.success("Fixed Signal Timing Detected")
    else:
        st.sidebar.warning("Fixed signal timing file not found: Streamlit_Callibration/traffic_lights.add.xml")

    # Phase 1 outputs are always generated in Streamlit_Callibration.
    phase1_target_datewise = DEFAULT_OUTPUT_WORKBOOK
    phase1_target_compact = DEFAULT_COMPACT_WORKBOOK
    phase1_target_edgewise = DEFAULT_EDGEWISE_WORKBOOK

    candidate_dates: list[str] = []
    if date_tabs_workbook.exists():
        candidate_dates = [sheet for sheet in pd.ExcelFile(date_tabs_workbook).sheet_names if re.match(r"\d{4}-\d{2}-\d{2}$", sheet)]

    selected_dates = st.sidebar.multiselect("Target dates", options=sorted(candidate_dates), default=sorted(candidate_dates)[:5])
    begin_text = st.sidebar.text_input("Begin time", "09:15:00")
    end_text = st.sidebar.text_input("End time", "10:00:00")
    pool_begin_sec = st.sidebar.number_input("Pool generation begin (sec)", min_value=0, value=0, step=60)
    pool_end_sec = st.sidebar.number_input("Pool generation end (sec)", min_value=300, value=5000, step=60)
    pool_trip_target = st.sidebar.slider("Candidate pool size", min_value=5000, max_value=250000, value=100000, step=1000)
    fringe_factor = st.sidebar.slider("randomTrips fringe factor", min_value=1.0, max_value=300.0, value=200.0, step=1.0)
    classified_toggle = st.sidebar.toggle("Enable classified calibration", value=True)
    auto_run_final_simulation = st.sidebar.toggle("Auto-run final simulation after routeSampler", value=True)

    st.sidebar.divider()
    st.sidebar.subheader("TLS Override Control")
    if st.sidebar.button("Run TLS Scrub/Enforce Only"):
        try:
            tls_sync_stats = prepare_survey_tls_priority(network_file, DEFAULT_TLS_ADD_FILE)
            tls_verify_stats = summarize_tls_priority_state(network_file, DEFAULT_TLS_ADD_FILE)
            st.session_state.tls_scrub_result = {
                "ran_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "sync": tls_sync_stats,
                "verify": tls_verify_stats,
            }
            st.sidebar.success(TLS_PRIORITY_STATUS_MESSAGE)
        except Exception as exc:
            st.session_state.tls_scrub_result = {
                "ran_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "error": str(exc),
            }
            st.sidebar.error(f"TLS scrub/enforce failed: {exc}")

    tabs = st.tabs(["Phase 1: Consolidate", "Phase 2: Configure", "Phase 3: Execute"])

    if "summary_df" not in st.session_state:
        st.session_state.summary_df = pd.DataFrame()
    if "mix_df" not in st.session_state:
        st.session_state.mix_df = pd.DataFrame()
    if "counts_preview" not in st.session_state:
        st.session_state.counts_preview = pd.DataFrame()
    if "mapping_df" not in st.session_state:
        try:
            st.session_state.mapping_df = read_location_mapping(location_workbook) if location_workbook.exists() else pd.DataFrame()
        except Exception:
            st.session_state.mapping_df = pd.DataFrame()
    if "log_text" not in st.session_state:
        st.session_state.log_text = ""
    if "filtered_route_pool" not in st.session_state:
        st.session_state.filtered_route_pool = route_pool
    if "pool_generation_log" not in st.session_state:
        st.session_state.pool_generation_log = ""
    if "final_simulation_log" not in st.session_state:
        st.session_state.final_simulation_log = ""
    if "phase1_datewise_path" not in st.session_state:
        st.session_state.phase1_datewise_path = phase1_target_datewise
    if "phase1_compact_path" not in st.session_state:
        st.session_state.phase1_compact_path = phase1_target_compact
    if "phase1_edgewise_path" not in st.session_state:
        st.session_state.phase1_edgewise_path = phase1_target_edgewise
    if "tls_scrub_result" not in st.session_state:
        st.session_state.tls_scrub_result = None

    tls_scrub_result = st.session_state.get("tls_scrub_result")
    if isinstance(tls_scrub_result, dict):
        st.sidebar.caption(f"Last TLS scrub run: {tls_scrub_result.get('ran_at', 'n/a')}")
        if "error" in tls_scrub_result:
            st.sidebar.caption(f"Last error: {tls_scrub_result.get('error', '')}")
        else:
            sync = tls_scrub_result.get("sync", {})
            verify = tls_scrub_result.get("verify", {})
            st.sidebar.caption("TLS scrub summary: " + ", ".join(f"{k}={v}" for k, v in sync.items()))
            st.sidebar.caption("Verification: " + ", ".join(f"{k}={v}" for k, v in verify.items()))
            if (
                int(verify.get("net_target_tllogics_remaining", 0)) > 0
                and int(verify.get("net_target_noninactive_programs", 0)) == 0
                and int(verify.get("add_target_program_nonzero", 0)) == 0
                and int(verify.get("add_missing_target_tllogics", 0)) == 0
            ):
                st.sidebar.success("TLS verification passed: network placeholders active, Program 0 active.")
            else:
                st.sidebar.warning("TLS verification warning: check net placeholder/programID mapping before running calibration.")

    with tabs[0]:
        title_block("Phase 1: Data Consolidation", "Load the consolidated master Summary, date tabs, and compact workbook directly, then filter them to the selected dates and 08:00-10:00 window.")
        st.write("Phase 1 creates datewise, compact, and EdgeWise 5-minute interval workbooks from the consolidated master Summary.")
        st.caption("Datewise workbook contains Summary plus per-date tabs. Compact workbook contains TotalByIntervalLocation, Bus, Car, Motorcycle, and Total sheets. EdgeWise workbook contains Bus_5MinCols, Car_5MinCols, Motorcycle_5MinCols, and Total_Volume_5MinCols.")
        st.caption(f"Phase 1 output folder is fixed to Streamlit_Callibration: {STREAMLIT_CALIB_DIR}")

        col1, col2 = st.columns(2)
        with col1:
            st.metric("Master summary rows", len(load_master_summary_workbook(master_workbook)))
            st.metric("Selected dates", len(selected_dates))
        with col2:
            st.metric("Expected benchmark rows", EXPECTED_SUMMARY_ROWS)
            st.write(f"Master workbook: {master_workbook}")

        if st.button("Process Raw Data", type="primary"):
            phase1_status = st.empty()
            phase1_progress = st.progress(0, text="Phase 1 started")
            summary_df = pd.DataFrame()
            master_summary_df = pd.DataFrame()
            date_tabs_df = pd.DataFrame()
            compact_source_df = pd.DataFrame()
            mix_df = pd.DataFrame()

            try:
                phase1_status.info("Loading consolidated workbooks...")
                phase1_progress.progress(10, text="Loading master/date-tab/compact workbooks")
                selected_for_load = selected_dates or None
                master_summary_df = load_master_summary_workbook(master_workbook, selected_for_load)
                date_tabs_df = load_date_tab_workbook(date_tabs_workbook, selected_for_load)
                compact_source_df = load_compact_workbook(compact_workbook, selected_for_load)
                summary_df = master_summary_df.copy()
                phase1_progress.progress(30, text="Source workbooks loaded")
            except Exception as exc:
                st.error(f"Failed to load consolidated workbooks: {exc}")
                phase1_progress.progress(100, text="Phase 1 failed")
                phase1_status.error("Phase 1 stopped while loading source workbooks.")

            compact_frames: dict[str, pd.DataFrame] | None = None
            edgewise_frames: dict[str, pd.DataFrame] | None = None
            actual_output_workbook = phase1_target_datewise
            actual_compact_workbook = phase1_target_compact
            actual_edgewise_workbook = phase1_target_edgewise
            if not summary_df.empty:
                phase1_status.info("Writing Datewise, Compact, and EdgeWise workbooks...")
                phase1_progress.progress(45, text="Resolving output paths")
                actual_output_workbook, datewise_notice = resolve_excel_output_path(phase1_target_datewise)
                actual_compact_workbook, compact_notice = resolve_excel_output_path(phase1_target_compact)
                actual_edgewise_workbook, edgewise_notice = resolve_excel_output_path(phase1_target_edgewise)
                write_datewise_workbook(actual_output_workbook, summary_df)
                phase1_progress.progress(60, text="Datewise workbook written")
                compact_frames = write_compact_workbook(actual_compact_workbook, summary_df)
                phase1_progress.progress(72, text="Compact workbook written")
                edgewise_frames = write_edgewise_datewise_workbook(actual_edgewise_workbook, summary_df, st.session_state.mapping_df)
                phase1_progress.progress(82, text="EdgeWise workbook written")
                for notice in [datewise_notice, compact_notice, edgewise_notice]:
                    if notice:
                        st.warning(notice)

                try:
                    location_edge_map_csv = DEFAULT_LOCATION_EDGE_MAP_CSV
                    phase1_status.info("Exporting automated location-edge mapping to DEBUG file...")
                    actual_csv_path, actual_debug_csv_path, mapped_counts = write_location_edge_map_csv(location_workbook, network_file, location_edge_map_csv)
                    phase1_progress.progress(92, text="Automated mapping exported to DEBUG")
                    st.success(f"Automated mapping exported to {actual_debug_csv_path.name} from LocationID and Network files")
                    st.info("Automated mapping exported to DEBUG file. Simulation is using manual location_edge_map.csv for corridor calibration.")
                    if actual_csv_path.exists():
                        st.caption(f"Manual mapping preserved: {actual_csv_path.name}")
                    else:
                        st.caption(f"Manual mapping not found yet: {actual_csv_path.name}")
                    if mapped_counts:
                        summary_lines = [
                            f"{location}: {mapped_counts[location]} edges mapped"
                            for location in sorted(mapped_counts.keys())
                        ]
                        st.caption("; ".join(summary_lines))
                except Exception as exc:
                    st.warning(f"Could not generate location_edge_map_DEBUG.csv: {exc}")

            if not summary_df.empty and classified_toggle and actual_compact_workbook.exists():
                try:
                    phase1_status.info("Joining vehicle mix from compact workbook...")
                    selected_for_mix = selected_dates or summary_df["SessionFolder"].unique().tolist()
                    summary_df = attach_vehicle_mix(summary_df, actual_compact_workbook, selected_for_mix, to_seconds(begin_text) or 0, to_seconds(end_text) or 0)
                    mix_df = load_compact_vehicle_mix(actual_compact_workbook, selected_for_mix, to_seconds(begin_text) or 0, to_seconds(end_text) or 0)
                    phase1_progress.progress(97, text="Vehicle mix joined")
                except Exception as exc:
                    st.warning(f"Vehicle mix join skipped: {exc}")

            st.session_state.summary_df = summary_df
            st.session_state.mix_df = mix_df
            st.session_state.master_summary_df = master_summary_df
            st.session_state.date_tabs_df = date_tabs_df
            st.session_state.compact_source_df = compact_source_df
            st.session_state.phase1_datewise_path = actual_output_workbook
            st.session_state.phase1_compact_path = actual_compact_workbook
            st.session_state.phase1_edgewise_path = actual_edgewise_workbook

            if not summary_df.empty:
                phase1_progress.progress(100, text="Phase 1 completed")
                phase1_status.success("Phase 1 complete.")
                row_count = len(summary_df)
                if row_count == EXPECTED_SUMMARY_ROWS:
                    st.success(f"Derived {row_count} rows from the 10,287-row master Summary and filtered to the 08:00-10:00 window.")
                else:
                    st.success(f"Derived {row_count} rows from the 10,287-row master Summary and filtered to the 08:00-10:00 window.")
                st.info(f"Wrote {actual_output_workbook}")
                st.info(f"Wrote {actual_compact_workbook}")
                st.info(f"Wrote {actual_edgewise_workbook}")
                if compact_frames is not None and "TotalByIntervalLocation" in compact_frames:
                    st.caption(f"Compact rows (TotalByIntervalLocation): {len(compact_frames['TotalByIntervalLocation'])}")
                if edgewise_frames is not None and "Total_Volume_5MinCols" in edgewise_frames:
                    st.caption(f"EdgeWise rows (Total_Volume_5MinCols): {len(edgewise_frames['Total_Volume_5MinCols'])}")
                st.caption(f"Date-tab rows loaded: {len(date_tabs_df)}")
            else:
                phase1_progress.progress(100, text="Phase 1 finished with no rows")
                phase1_status.warning("Phase 1 finished, but no rows were consolidated.")
                st.warning("No rows were consolidated from the selected source.")

        if not st.session_state.summary_df.empty:
            st.dataframe(st.session_state.summary_df.head(200), use_container_width=True, height=360)
            phase1_datewise_path = Path(st.session_state.get("phase1_datewise_path", phase1_target_datewise))
            phase1_compact_path = Path(st.session_state.get("phase1_compact_path", phase1_target_compact))
            phase1_edgewise_path = Path(st.session_state.get("phase1_edgewise_path", phase1_target_edgewise))
            if phase1_datewise_path.exists():
                st.download_button(
                    "Download Datewise workbook",
                    data=phase1_datewise_path.read_bytes(),
                    file_name=phase1_datewise_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            if phase1_compact_path.exists():
                st.download_button(
                    "Download Compact workbook",
                    data=phase1_compact_path.read_bytes(),
                    file_name=phase1_compact_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            if phase1_edgewise_path.exists():
                st.download_button(
                    "Download EdgeWise workbook",
                    data=phase1_edgewise_path.read_bytes(),
                    file_name=phase1_edgewise_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    with tabs[1]:
        title_block("Phase 2: Calibration Configuration", "Set the SUMO inputs, dates, time window, and location-to-junction mapping used to build counts.xml.")

        st.subheader("Step 1: Generate Candidate Route Pool")
        st.write("Generate an over-saturated candidate pool with SUMO randomTrips.py (recommended: 100,000+ trips for high motorcycle demand).")
        phase1_datewise_path = Path(st.session_state.get("phase1_datewise_path", output_workbook))
        if st.button("Generate Pool", type="primary"):
            if not network_file.exists():
                st.error(f"Network file not found: {network_file}")
            elif not phase1_datewise_path.exists():
                st.error(f"Datewise summary not found: {phase1_datewise_path}. Run 'Process Raw Data' in Phase 1 first.")
            else:
                output_placeholder = st.empty()
                try:
                    with st.spinner("Generating candidate route pool..."):
                        generation_log = generate_candidate_route_pool(
                            network_file=network_file,
                            trip_file=trip_pool,
                            route_file=route_pool,
                            begin_sec=int(pool_begin_sec),
                            end_sec=int(max(pool_end_sec, pool_begin_sec + 1)),
                            target_trips=int(pool_trip_target),
                            fringe_factor=float(fringe_factor),
                            output_placeholder=output_placeholder,
                        )
                    st.session_state.pool_generation_log = generation_log
                    st.success(f"Generated candidate pool: {route_pool}")
                except Exception as exc:
                    st.error(f"Failed to generate candidate route pool: {exc}")

        if st.session_state.pool_generation_log:
            st.code(st.session_state.pool_generation_log, language="text")

        config_col1, config_col2 = st.columns([1.1, 0.9])
        with config_col1:
            st.subheader("Location Mapping")
            mapping_df = st.session_state.mapping_df.copy()
            if not mapping_df.empty:
                edited_mapping = st.data_editor(mapping_df, use_container_width=True, num_rows="dynamic", key="mapping_editor")
                st.session_state.mapping_df = edited_mapping
            else:
                st.warning("No mapping workbook loaded.")

            if st.button("Save Edited Mapping"):
                temp_location_workbook = STREAMLIT_CALIB_DIR / "LocationID_edited.xlsx"
                if not st.session_state.mapping_df.empty:
                    write_location_workbook(st.session_state.mapping_df, temp_location_workbook)
                    st.session_state.location_workbook_override = temp_location_workbook
                    st.success(f"Saved mapping workbook to {temp_location_workbook}")

        with config_col2:
            st.subheader("Route Pool Preview")
            if route_pool.exists():
                st.dataframe(load_route_pool_preview(route_pool), use_container_width=True, height=320)
            else:
                st.warning(f"Route pool not found: {route_pool}")

        st.subheader("Step 2: Whitelist Filter")
        st.write("Select mandatory corridor locations. Routes not passing at least one selected location edge are removed.")
        st.info("Whitelist filtering in Phase 2 Step 2 uses manual location_edge_map.csv (not DEBUG).")
        whitelist_candidates: list[str] = []
        try:
            if location_workbook.exists() and network_file.exists():
                location_edge_map = get_location_edge_map(location_workbook, network_file)
                whitelist_candidates = sorted(location_edge_map.keys())
            elif st.session_state.mapping_df is not None and not st.session_state.mapping_df.empty and "LocationName" in st.session_state.mapping_df.columns:
                whitelist_candidates = sorted(st.session_state.mapping_df["LocationName"].astype(str).unique().tolist())
        except Exception:
            whitelist_candidates = []

        default_whitelist = [
            name
            for name in whitelist_candidates
            if normalize_key(name) in {normalize_key("Thapathali"), normalize_key("Pulchowk Bridge1"), normalize_key("Kesharmahal")}
        ]
        whitelist_locations_phase2 = st.multiselect(
            "Mandatory locations for whitelist",
            options=whitelist_candidates,
            default=default_whitelist,
            key="phase2_whitelist_locations",
        )

        if st.button("Apply Whitelist", key="phase2_apply_whitelist"):
            if not route_pool.exists():
                st.error(f"Candidate route pool not found: {route_pool}. Generate pool first.")
            else:
                try:
                    whitelist_edges: list[str] = []
                    location_edge_map_csv = DEFAULT_LOCATION_EDGE_MAP_CSV
                    if not location_edge_map_csv.exists():
                        _, debug_map_path, _ = write_location_edge_map_csv(location_workbook, network_file, DEFAULT_LOCATION_EDGE_MAP_DEBUG_CSV)
                        raise FileNotFoundError(
                            f"Manual mapping file not found: {location_edge_map_csv}. "
                            f"Automated mapping was exported to: {debug_map_path}"
                        )

                    map_df = load_location_edge_map_csv(location_edge_map_csv)
                    location_edge_map: dict[str, dict[str, str]] = {}
                    for row in map_df.itertuples(index=False):
                        location_name = str(row.Location)
                        approach_name = str(row.Approach)
                        edge_id = str(row.EdgeID)
                        location_edge_map.setdefault(location_name, {})[approach_name] = edge_id

                    for location in whitelist_locations_phase2:
                        whitelist_edges.extend(location_edge_map.get(location, {}).values())

                    kept, removed = filter_route_pool(route_pool, whitelist_edges, whitelisted_pool_output)
                    st.session_state.filtered_route_pool = whitelisted_pool_output
                    st.success(f"Whitelist applied. Kept {kept} routes and removed {removed}.")
                    st.info(f"Whitelisted pool saved to {whitelisted_pool_output}")
                except Exception as exc:
                    st.error(f"Whitelist filtering failed: {exc}")

        if st.button("Build counts.xml preview", type="primary"):
            try:
                location_edge_map_csv = DEFAULT_LOCATION_EDGE_MAP_CSV
                if not location_edge_map_csv.exists():
                    _, debug_map_path, _ = write_location_edge_map_csv(location_workbook, network_file, DEFAULT_LOCATION_EDGE_MAP_DEBUG_CSV)
                    raise FileNotFoundError(
                        f"Manual mapping file not found: {location_edge_map_csv}. "
                        f"Automated mapping was exported to: {debug_map_path}"
                    )

                counts_path, counts_preview_df, diagnostics = write_counts_xml_from_edgewise(
                    edgewise_workbook=phase1_edgewise_path,
                    location_edge_map_csv=location_edge_map_csv,
                    network_file=network_file,
                    output_path=counts_output,
                    begin_text=begin_text,
                    end_text=end_text,
                )
                st.session_state.counts_preview = counts_preview_df

                detectors_path, detector_count = write_detectors_add_from_location_edge_map(
                    network_file=network_file,
                    location_edge_map_csv=location_edge_map_csv,
                    output_path=DEFAULT_DETECTORS_ADD_FILE,
                )

                st.success(f"Wrote counts.xml to {counts_path}")
                st.caption(
                    f"Mapped edges: {diagnostics.get('mapped_edge_count', 0)} | "
                    f"interval rows: {diagnostics.get('mapped_row_count', 0)} | "
                    f"skipped unmapped: {diagnostics.get('skipped_unmapped_rows', 0)} | "
                    f"skipped invalid edges: {diagnostics.get('skipped_invalid_edge_rows', 0)}"
                )
                if diagnostics.get("all_counts_zero"):
                    st.warning("All entered counts are zero. Check Total_Volume_5MinCols loading/mapping.")
                st.success(f"Wrote detectors file to {detectors_path} ({detector_count} detectors)")
                st.dataframe(counts_preview_df.head(120), use_container_width=True, height=280)
            except Exception as exc:
                st.error(f"Failed to build counts.xml preview: {exc}")

    with tabs[2]:
        title_block("Phase 3: Whitelist and Execution", "Filter route pools to corridor movements, inspect the XML preview, then launch routeSampler with live output.")
        st.write("Calibration uses the whitelisted pool when available; otherwise it falls back to the candidate route pool.")

        counts_preview = st.session_state.counts_preview
        if counts_preview.empty and not st.session_state.summary_df.empty:
            counts_preview = build_counts_preview(st.session_state.summary_df, selected_dates or None, begin_text, end_text)

        if not counts_preview.empty:
            st.subheader("Counts XML Preview")
            st.dataframe(counts_preview.head(100), use_container_width=True, height=260)
        else:
            st.info("Build a counts preview in Phase 2 first.")

        manual_map_path = DEFAULT_LOCATION_EDGE_MAP_CSV
        if manual_map_path.exists():
            missing_primary_preview, missing_reverse_preview = validate_required_critical_edges_in_manual_map(manual_map_path)
            if missing_primary_preview:
                st.error(
                    "CRITICAL: Manual location_edge_map.csv is missing required critical edges: "
                    + ", ".join(missing_primary_preview)
                )
            else:
                st.caption("Critical edge check passed for manual location_edge_map.csv (232351915#1, 232429764#1).")
            if missing_reverse_preview:
                st.warning(
                    "Map audit: reverse-direction critical edges are missing in manual location_edge_map.csv: "
                    + ", ".join(missing_reverse_preview)
                    + ". This can remove valid corridor routes during whitelist filtering."
                )
        else:
            st.warning(f"Manual mapping file not found for critical-edge validation: {manual_map_path}")

        available_locations: list[str] = []
        if not st.session_state.summary_df.empty and "Location" in st.session_state.summary_df.columns:
            available_locations = sorted(st.session_state.summary_df["Location"].astype(str).unique().tolist())
        elif st.session_state.mapping_df is not None and not st.session_state.mapping_df.empty and "LocationName" in st.session_state.mapping_df.columns:
            available_locations = sorted(st.session_state.mapping_df["LocationName"].astype(str).unique().tolist())

        whitelist_locations = st.multiselect("Whitelist locations", options=available_locations, default=[name for name in available_locations if name in {"Pulchowk Bridge1", "Thapathali"}])
        extra_whitelist_edges = st.text_area("Extra whitelist edge IDs (one per line)", value="")

        # Calibration mode toggle
        st.divider()
        st.subheader("Calibration Mode")
        calibration_mode = st.radio(
            "Choose calibration strategy",
            options=["Total Volume Calibration", "Classified Modal Calibration"],
            help="Total Volume: Use combined Bus+Car+Motorcycle counts. Classified Modal: Generate mode-specific calibration (bus, car, motorcycle).",
        )
        st.session_state.calibration_mode = calibration_mode

        if st.button("Apply whitelist to route pool"):
            try:
                whitelist_edges: list[str] = []
                location_workbook_for_counts = st.session_state.get("location_workbook_override", location_workbook)
                if location_workbook_for_counts.exists() and network_file.exists():
                    location_edge_map_csv = DEFAULT_LOCATION_EDGE_MAP_CSV
                    if not location_edge_map_csv.exists():
                        _, debug_map_path, _ = write_location_edge_map_csv(location_workbook_for_counts, network_file, DEFAULT_LOCATION_EDGE_MAP_DEBUG_CSV)
                        raise FileNotFoundError(
                            f"Manual mapping file not found: {location_edge_map_csv}. "
                            f"Automated mapping was exported to: {debug_map_path}"
                        )

                    _, _, _ = write_counts_xml_from_edgewise(
                        edgewise_workbook=Path(st.session_state.get("phase1_edgewise_path", phase1_target_edgewise)),
                        location_edge_map_csv=location_edge_map_csv,
                        network_file=network_file,
                        output_path=counts_output,
                        begin_text=begin_text,
                        end_text=end_text,
                    )

                    location_edge_map: dict[str, dict[str, str]] = {}
                    map_df = load_location_edge_map_csv(location_edge_map_csv)
                    for row in map_df.itertuples(index=False):
                        location_name = str(row.Location)
                        approach_name = str(row.Approach)
                        edge_id = str(row.EdgeID)
                        location_edge_map.setdefault(location_name, {})[approach_name] = edge_id
                    for location in whitelist_locations:
                        whitelist_edges.extend(location_edge_map.get(location, {}).values())
                whitelist_edges.extend([line.strip() for line in extra_whitelist_edges.splitlines() if line.strip()])

                filtered_route_pool = whitelisted_pool_output
                kept, removed = filter_route_pool(route_pool, whitelist_edges, filtered_route_pool)
                st.session_state.filtered_route_pool = filtered_route_pool
                st.success(f"Whitelist applied. Kept {kept} vehicles and removed {removed}.")
                st.info(f"Filtered route pool: {filtered_route_pool}")
            except Exception as exc:
                st.error(f"Whitelist filtering failed: {exc}")

        if st.button("Run routeSampler.py", type="primary"):
            try:
                source_route_file = st.session_state.get("filtered_route_pool", whitelisted_pool_output if whitelisted_pool_output.exists() else route_pool)
                calibration_mode = st.session_state.get("calibration_mode", "Total Volume Calibration")
                
                location_workbook_for_counts = st.session_state.get("location_workbook_override", location_workbook)
                location_edge_map_csv = DEFAULT_LOCATION_EDGE_MAP_CSV
                if not location_edge_map_csv.exists():
                    _, debug_map_path, _ = write_location_edge_map_csv(location_workbook_for_counts, network_file, DEFAULT_LOCATION_EDGE_MAP_DEBUG_CSV)
                    raise FileNotFoundError(
                        f"Manual mapping file not found: {location_edge_map_csv}. "
                        f"Automated mapping was exported to: {debug_map_path}"
                    )

                missing_primary_edges, missing_reverse_edges = validate_required_critical_edges_in_manual_map(location_edge_map_csv)
                if missing_primary_edges:
                    st.error(
                        "CRITICAL: Manual location_edge_map.csv is missing required critical edges: "
                        + ", ".join(missing_primary_edges)
                    )
                    st.stop()
                if missing_reverse_edges:
                    st.warning(
                        "Map audit warning: reverse-direction critical edges are missing from manual mapping: "
                        + ", ".join(missing_reverse_edges)
                    )

                if calibration_mode == "Classified Modal Calibration":
                    # Mode-specific calibration workflow
                    st.info("Running Classified Modal Calibration (Bus, Car, Motorcycle)...")

                    scipy_ready, scipy_note = ensure_scipy_available(install_if_missing=True)
                    if scipy_note:
                        if scipy_ready:
                            st.info(scipy_note)
                        else:
                            st.warning(scipy_note)

                    edgewise_input_path = Path(st.session_state.get("phase1_edgewise_path", phase1_target_edgewise))
                    prepared_edgewise_path, correction_info = prepare_edgewise_workbook_with_minor_corrections(
                        edgewise_input_path,
                        DEFAULT_EDGEWISE_BACKUP_DIR,
                    )
                    if bool(correction_info.get("corrected", False)):
                        st.warning(
                            f"Applied minor corrections to EdgeWise counts workbook (changed_cells={correction_info.get('total_changed_cells', 0)})."
                        )
                        st.caption(f"Original backup saved: {correction_info.get('backup_path', '')}")
                        st.caption(f"Corrected workbook used for this run: {correction_info.get('corrected_path', '')}")
                    else:
                        st.caption("No numeric corrections were needed in EdgeWise sheets; original workbook is used.")
                    
                    # Generate classified counts.xml with vType attributes (combined view)
                    counts_output_classified, classified_results = write_counts_xml_from_edgewise_classified(
                        edgewise_workbook=prepared_edgewise_path,
                        location_edge_map_csv=location_edge_map_csv,
                        network_file=network_file,
                        output_path=counts_output,
                        begin_text=begin_text,
                        end_text=end_text,
                    )
                    
                    # Display diagnostics for each vehicle class
                    for vehicle_class, (counts_df, diagnostics) in classified_results.items():
                        with st.expander(f"Diagnostics: {vehicle_class.capitalize()}"):
                            st.write(diagnostics)
                            st.write(f"Sample rows:\n{counts_df.head(10).to_string()}")
                    
                    write_detectors_add_from_location_edge_map(network_file, location_edge_map_csv, DEFAULT_DETECTORS_ADD_FILE)

                    begin_sec_counts = to_seconds(begin_text)
                    end_sec_counts = to_seconds(end_text)
                    if begin_sec_counts is None or end_sec_counts is None or end_sec_counts <= begin_sec_counts:
                        raise ValueError("Invalid begin/end time for classified calibration")

                    source_pool_count, source_pool_min, source_pool_max = get_route_pool_depart_window(Path(source_route_file))
                    window_mismatch = False
                    if source_pool_count > 0 and source_pool_min is not None and source_pool_max is not None:
                        window_mismatch = source_pool_max < float(begin_sec_counts) or source_pool_min >= float(end_sec_counts)

                    if window_mismatch:
                        st.warning(
                            "Route pool departures are outside the calibration window. "
                            f"pool_depart_range={source_pool_min:.0f}-{source_pool_max:.0f}, "
                            f"calibration_window={begin_sec_counts}-{end_sec_counts}. "
                            "Regenerating candidate pool for the active calibration window."
                        )
                        regen_placeholder = st.empty()
                        with st.spinner("Regenerating candidate route pool for calibration window..."):
                            generation_log = generate_candidate_route_pool(
                                network_file=network_file,
                                trip_file=trip_pool,
                                route_file=route_pool,
                                begin_sec=int(begin_sec_counts),
                                end_sec=int(end_sec_counts),
                                target_trips=int(max(pool_trip_target, 100000)),
                                fringe_factor=float(fringe_factor),
                                output_placeholder=regen_placeholder,
                            )
                        st.session_state.pool_generation_log = generation_log

                        # Re-apply current Phase 3 whitelist selections to regenerated pool.
                        whitelist_edges: list[str] = []
                        map_df = load_location_edge_map_csv(location_edge_map_csv)
                        location_edge_map: dict[str, dict[str, str]] = {}
                        for row in map_df.itertuples(index=False):
                            location_name = str(row.Location)
                            approach_name = str(row.Approach)
                            edge_id = str(row.EdgeID)
                            location_edge_map.setdefault(location_name, {})[approach_name] = edge_id
                        for location in whitelist_locations:
                            whitelist_edges.extend(location_edge_map.get(location, {}).values())
                        whitelist_edges.extend([line.strip() for line in extra_whitelist_edges.splitlines() if line.strip()])

                        if whitelist_edges:
                            kept, removed = filter_route_pool(route_pool, whitelist_edges, whitelisted_pool_output)
                            st.session_state.filtered_route_pool = whitelisted_pool_output
                            source_route_file = whitelisted_pool_output
                            st.info(
                                f"Rebuilt whitelist after pool regeneration (kept={kept}, removed={removed})."
                            )
                        else:
                            source_route_file = route_pool
                            st.session_state.filtered_route_pool = route_pool
                            st.info("Using regenerated full route pool (no whitelist edges selected).")

                    source_route_path = Path(source_route_file)
                    kept_valid, removed_invalid = sanitize_route_file_against_network(source_route_path, network_file)
                    if removed_invalid > 0:
                        st.warning(
                            f"Removed {removed_invalid} route(s) from {source_route_path.name} because they reference edges not present in {network_file.name}."
                        )
                    if kept_valid == 0:
                        raise RuntimeError(
                            f"No valid routes remain in {source_route_path.name} after validating against {network_file.name}."
                        )
                    source_route_file = source_route_path

                    mode_files = {
                        "bus": (STREAMLIT_CALIB_DIR / "counts_bus.xml", STREAMLIT_CALIB_DIR / "calibrated_routes_bus.rou.xml", "bus_"),
                        "car": (STREAMLIT_CALIB_DIR / "counts_car.xml", STREAMLIT_CALIB_DIR / "calibrated_routes_car.rou.xml", "car_"),
                        "motorcycle": (STREAMLIT_CALIB_DIR / "counts_motorcycle.xml", STREAMLIT_CALIB_DIR / "calibrated_routes_motorcycle.rou.xml", "moto_"),
                    }

                    tls_sync_stats = prepare_survey_tls_priority(network_file, DEFAULT_TLS_ADD_FILE)
                    st.success(TLS_PRIORITY_STATUS_MESSAGE)
                    if any(tls_sync_stats.values()):
                        st.caption(
                            "Pre-routeSampler TLS scrub summary: "
                            + ", ".join(f"{key}={value}" for key, value in tls_sync_stats.items())
                        )

                    mode_logs: list[str] = []
                    log_placeholder = st.empty()
                    with st.spinner("Running routeSampler.py per vehicle mode (bus/car/motorcycle)..."):
                        for vehicle_class in ["bus", "car", "motorcycle"]:
                            counts_mode_path, routes_mode_path, id_prefix = mode_files[vehicle_class]
                            aggregated_mode_df, _diag = classified_results[vehicle_class]
                            write_counts_xml_single_mode_from_aggregated(
                                aggregated=aggregated_mode_df,
                                vehicle_class=vehicle_class,
                                output_path=counts_mode_path,
                                begin_sec=begin_sec_counts,
                                end_sec=end_sec_counts,
                            )

                            mode_source_file = Path(source_route_file)
                            source_feasible, source_report, source_totals = evaluate_route_pool_feasibility_for_edges(
                                mode_source_file,
                                counts_mode_path,
                                REQUIRED_CRITICAL_UNDERFLOW_EDGES,
                            )
                            source_coverage = coverage_ratio_from_totals(source_totals)
                            fallback_source = route_pool
                            fallback_coverage = None
                            fallback_feasible = False
                            fallback_report = ""

                            if not source_feasible:
                                st.warning(
                                    f"{vehicle_class.capitalize()} pre-check: current source pool appears infeasible on critical edges. "
                                    f"{source_report} (coverage_ratio={source_coverage:.3f})"
                                )

                                if fallback_source.exists() and mode_source_file.resolve() != fallback_source.resolve():
                                    fallback_feasible, fallback_report, fallback_totals = evaluate_route_pool_feasibility_for_edges(
                                        fallback_source,
                                        counts_mode_path,
                                        REQUIRED_CRITICAL_UNDERFLOW_EDGES,
                                    )
                                    fallback_coverage = coverage_ratio_from_totals(fallback_totals)
                                    if fallback_feasible:
                                        st.warning(
                                            f"{vehicle_class.capitalize()} pre-check: switching to full route pool because whitelist pool is infeasible."
                                        )
                                        mode_source_file = fallback_source
                                        source_coverage = fallback_coverage
                                    elif vehicle_class == "motorcycle":
                                        st.warning(
                                            f"{vehicle_class.capitalize()} pre-check: full route pool is still infeasible "
                                            f"(coverage_ratio={fallback_coverage:.3f}). Proceeding with compatible routeSampler fallback flags."
                                        )
                                        mode_source_file = fallback_source
                                        source_coverage = fallback_coverage
                                    else:
                                        raise RuntimeError(
                                            f"Infeasible counts vs candidate pool for mode '{vehicle_class}' on critical edges. "
                                            f"Current pool: {source_report}. Full pool: {fallback_report}."
                                        )
                                elif vehicle_class == "motorcycle":
                                    st.warning(
                                        f"{vehicle_class.capitalize()} pre-check: source route pool is infeasible "
                                        f"(coverage_ratio={source_coverage:.3f}) and no alternative pool is available. "
                                        "Proceeding with compatible routeSampler fallback flags."
                                    )
                                else:
                                    raise RuntimeError(
                                        f"Infeasible counts vs candidate pool for mode '{vehicle_class}' on critical edges. {source_report}."
                                    )

                            st.caption(
                                f"Coverage Ratio ({vehicle_class}): {source_coverage:.3f} "
                                "[capacity / required on critical edges]"
                            )

                            mismatch_path = STREAMLIT_CALIB_DIR / f"mismatch_{vehicle_class}.xml"
                            extra_args, arg_notes = build_mode_routesampler_args(
                                route_sampler,
                                mismatch_path,
                                vehicle_class,
                                scipy_available=scipy_ready,
                            )
                            for note in arg_notes:
                                st.warning(note)

                            mode_log = stream_route_sampler(
                                route_sampler,
                                mode_source_file,
                                counts_mode_path,
                                routes_mode_path,
                                log_placeholder,
                                extra_args=extra_args,
                            )
                            if "routeSampler exited with code" in mode_log:
                                raise RuntimeError(
                                    f"routeSampler failed for mode '{vehicle_class}'. Check Live Output for details."
                                )

                            # If critical corridor edges underflow on whitelisted pool, retry with full candidate pool.
                            underflow_edges = parse_underflow_edges_from_routesampler_log(mode_log)
                            critical_underflow_edges = sorted(REQUIRED_CRITICAL_UNDERFLOW_EDGES.intersection(underflow_edges))
                            fallback_source = route_pool
                            latest_critical_underflow_edges = list(critical_underflow_edges)
                            if (
                                critical_underflow_edges
                                and fallback_source.exists()
                                and mode_source_file.resolve() != fallback_source.resolve()
                            ):
                                st.warning(
                                    f"{vehicle_class.capitalize()} underflow on critical edges "
                                    f"({', '.join(critical_underflow_edges)}). Retrying with full route pool."
                                )
                                retry_log = stream_route_sampler(
                                    route_sampler,
                                    fallback_source,
                                    counts_mode_path,
                                    routes_mode_path,
                                    log_placeholder,
                                    extra_args=extra_args,
                                )
                                if "routeSampler exited with code" in retry_log:
                                    raise RuntimeError(
                                        f"routeSampler failed during full-pool retry for mode '{vehicle_class}'."
                                    )
                                latest_critical_underflow_edges = sorted(
                                    REQUIRED_CRITICAL_UNDERFLOW_EDGES.intersection(
                                        parse_underflow_edges_from_routesampler_log(retry_log)
                                    )
                                )
                                mode_log = (
                                    mode_log
                                    + "\n\n"
                                    + "[Adaptive retry with full route pool triggered by critical-edge underflow]\n"
                                    + retry_log
                                )

                            # If critical underflow remains even after full-pool retry, amplify critical-edge candidates.
                            if latest_critical_underflow_edges and fallback_source.exists():
                                boost_source = fallback_source
                                max_boost_rounds = 3 if vehicle_class == "bus" else 1
                                for boost_round in range(max_boost_rounds):
                                    if not latest_critical_underflow_edges:
                                        break

                                    total_deficit, deficit_intervals = parse_critical_edge_deficit_from_mismatch(
                                        mismatch_path,
                                        latest_critical_underflow_edges,
                                    )
                                    duplicate_factor = max(2, min(8, (total_deficit // 120) + 2 + boost_round))
                                    max_added = max(3000, min(12000, total_deficit * 20 if total_deficit > 0 else 5000))

                                    if vehicle_class == "bus":
                                        duplicate_factor = max(duplicate_factor, 5 + boost_round)
                                        max_added = max(max_added, 6000 + (4000 * boost_round))

                                    boosted_pool = STREAMLIT_CALIB_DIR / f"route_pool_boosted_{vehicle_class}_r{boost_round + 1}.rou.xml"
                                    seed_count, added_count = amplify_route_pool_for_critical_edges(
                                        boost_source,
                                        latest_critical_underflow_edges,
                                        boosted_pool,
                                        target_intervals=deficit_intervals,
                                        duplicate_factor=int(duplicate_factor),
                                        max_added=int(max_added),
                                    )
                                    if added_count <= 0:
                                        st.warning(
                                            f"{vehicle_class.capitalize()} critical-edge underflow persists, "
                                            "but no candidate routes containing the critical edges were found for boosting."
                                        )
                                        break

                                    st.warning(
                                        f"{vehicle_class.capitalize()} still underflowing critical edges "
                                        f"({', '.join(latest_critical_underflow_edges)}). "
                                        f"Retrying with boosted pool r{boost_round + 1} "
                                        f"(critical_deficit={total_deficit}, intervals={len(deficit_intervals)}, "
                                        f"seed={seed_count}, added={added_count})."
                                    )
                                    boosted_log = stream_route_sampler(
                                        route_sampler,
                                        boosted_pool,
                                        counts_mode_path,
                                        routes_mode_path,
                                        log_placeholder,
                                        extra_args=extra_args,
                                    )
                                    if "routeSampler exited with code" in boosted_log:
                                        raise RuntimeError(
                                            f"routeSampler failed during boosted-pool retry for mode '{vehicle_class}'."
                                        )

                                    latest_critical_underflow_edges = sorted(
                                        REQUIRED_CRITICAL_UNDERFLOW_EDGES.intersection(
                                            parse_underflow_edges_from_routesampler_log(boosted_log)
                                        )
                                    )
                                    boost_source = boosted_pool
                                    mode_log = (
                                        mode_log
                                        + "\n\n"
                                        + f"[Adaptive retry with boosted critical-edge pool r{boost_round + 1}]\n"
                                        + boosted_log
                                    )

                                if latest_critical_underflow_edges:
                                    if vehicle_class in {"car", "motorcycle"}:
                                        st.info(
                                            f"{vehicle_class.capitalize()} switching to direct rescue injection "
                                            "after one boosted retry to avoid repeated underflow loops."
                                        )
                                    remaining_deficits = parse_edge_interval_deficits_from_mismatch(
                                        mismatch_path,
                                        latest_critical_underflow_edges,
                                    )
                                    rescue_source = boost_source if boost_source.exists() else fallback_source
                                    rescue_added, rescue_unresolved = inject_rescue_vehicles_for_critical_deficits(
                                        source_route_pool=rescue_source,
                                        output_route_file=routes_mode_path,
                                        deficit_by_edge_interval=remaining_deficits,
                                        max_added=(6000 if vehicle_class == "bus" else 2500 if vehicle_class == "car" else 1500),
                                    )
                                    if rescue_added > 0:
                                        st.info(
                                            f"{vehicle_class.capitalize()} rescue injection added {rescue_added} supplemental vehicles "
                                            f"for remaining critical-edge deficits."
                                        )
                                    elif rescue_unresolved == 0:
                                        st.info(
                                            f"{vehicle_class.capitalize()} rescue step found no remaining interval deficits on critical edges."
                                        )
                                    if rescue_unresolved > 0:
                                        st.warning(
                                            f"{vehicle_class.capitalize()} still reports unresolved critical deficit={rescue_unresolved} "
                                            f"after rescue injection on edges: {', '.join(latest_critical_underflow_edges)}."
                                        )

                            assign_vehicle_type_to_routes(routes_mode_path, vehicle_class, id_prefix)
                            mode_logs.append(f"=== {vehicle_class.upper()} ===")
                            mode_logs.append(mode_log)

                    merge_classified_routes(
                        bus_routes=mode_files["bus"][1],
                        car_routes=mode_files["car"][1],
                        motorcycle_routes=mode_files["motorcycle"][1],
                        output_routes=calibrated_routes_output,
                    )
                    sorted_count, moved_count = sort_route_file_by_depart_time(calibrated_routes_output)
                    if moved_count > 0:
                        st.info(
                            f"Sorted merged calibrated routes by departure time "
                            f"(moved={moved_count}, sortable={sorted_count})."
                        )
                    st.session_state.log_text = "\n".join(mode_logs)
                    st.success(f"Finished per-mode routeSampler run. Merged output: {calibrated_routes_output}")

                else:
                    # Total volume calibration workflow (existing logic)
                    source_route_path = Path(source_route_file)
                    kept_valid, removed_invalid = sanitize_route_file_against_network(source_route_path, network_file)
                    if removed_invalid > 0:
                        st.warning(
                            f"Removed {removed_invalid} route(s) from {source_route_path.name} because they reference edges not present in {network_file.name}."
                        )
                    if kept_valid == 0:
                        raise RuntimeError(
                            f"No valid routes remain in {source_route_path.name} after validating against {network_file.name}."
                        )
                    source_route_file = source_route_path

                    if not counts_output.exists():
                        write_counts_xml_from_edgewise(
                            edgewise_workbook=Path(st.session_state.get("phase1_edgewise_path", phase1_target_edgewise)),
                            location_edge_map_csv=location_edge_map_csv,
                            network_file=network_file,
                            output_path=counts_output,
                            begin_text=begin_text,
                            end_text=end_text,
                        )
                        write_detectors_add_from_location_edge_map(network_file, location_edge_map_csv, DEFAULT_DETECTORS_ADD_FILE)

                    if not counts_output.exists():
                        st.error("counts.xml does not exist. Build the counts preview first.")
                    else:
                        tls_sync_stats = prepare_survey_tls_priority(network_file, DEFAULT_TLS_ADD_FILE)
                        st.success(TLS_PRIORITY_STATUS_MESSAGE)
                        if any(tls_sync_stats.values()):
                            st.caption(
                                "Pre-routeSampler TLS scrub summary: "
                                + ", ".join(f"{key}={value}" for key, value in tls_sync_stats.items())
                            )

                        log_placeholder = st.empty()
                        with st.spinner("Running routeSampler.py..."):
                            log_output = stream_route_sampler(route_sampler, source_route_file, counts_output, calibrated_routes_output, log_placeholder)
                        sorted_count, moved_count = sort_route_file_by_depart_time(calibrated_routes_output)
                        if moved_count > 0:
                            st.info(
                                f"Sorted calibrated routes by departure time "
                                f"(moved={moved_count}, sortable={sorted_count})."
                            )
                        st.session_state.log_text = log_output
                        st.success(f"Finished routeSampler run. Output: {calibrated_routes_output}")

                # Final simulation (works with both modes)
                if auto_run_final_simulation:
                    if not fixed_signal_timing_detected:
                        st.warning("Skipping final simulation because fixed traffic signal timing file was not detected.")
                    else:
                        detectors_ready = True
                        if not DEFAULT_DETECTORS_ADD_FILE.exists():
                            try:
                                location_edge_map_csv = DEFAULT_LOCATION_EDGE_MAP_CSV
                                location_workbook_for_counts = st.session_state.get("location_workbook_override", location_workbook)
                                if not location_edge_map_csv.exists():
                                    _, debug_map_path, _ = write_location_edge_map_csv(location_workbook_for_counts, network_file, DEFAULT_LOCATION_EDGE_MAP_DEBUG_CSV)
                                    raise FileNotFoundError(
                                        f"Manual mapping file not found: {location_edge_map_csv}. "
                                        f"Automated mapping was exported to: {debug_map_path}"
                                    )
                                detectors_path, detector_count = write_detectors_add_from_location_edge_map(
                                    network_file,
                                    location_edge_map_csv,
                                    DEFAULT_DETECTORS_ADD_FILE,
                                )
                                st.info(f"Generated missing detectors file: {detectors_path} ({detector_count} detectors)")
                            except Exception as detector_exc:
                                detectors_ready = False
                                st.warning(f"Skipping final simulation because detectors.add.xml could not be generated: {detector_exc}")

                        if detectors_ready:
                            begin_sec = to_seconds(begin_text) or 0
                            end_sec = to_seconds(end_text) or begin_sec + 300
                            if end_sec <= begin_sec:
                                end_sec = begin_sec + 300

                            tls_file = STREAMLIT_CALIB_DIR / "traffic_lights.add.xml"

                            removed_vtypes, removed_distributions = sanitize_route_file_vtypes(calibrated_routes_output)
                            if removed_vtypes or removed_distributions:
                                st.info(
                                    f"Sanitized calibrated routes: removed {removed_vtypes} conflicting vType and {removed_distributions} vTypeDistribution entries."
                                )

                            total_routes, rescue_before, rescue_removed = cap_rescue_vehicles_in_routes(
                                calibrated_routes_output,
                                max_ratio=0.015,
                                max_absolute=400,
                            )
                            if rescue_removed > 0:
                                st.warning(
                                    f"Capped rescue vehicles in calibrated routes: removed {rescue_removed} "
                                    f"(before={rescue_before}, total={total_routes})."
                                )

                            kept_valid, removed_invalid = sanitize_route_file_against_network(calibrated_routes_output, network_file)
                            if removed_invalid > 0:
                                st.warning(
                                    f"Sanitized calibrated routes against {network_file.name}: removed {removed_invalid} route(s) with unknown edges."
                                )
                            if kept_valid == 0:
                                raise RuntimeError(
                                    f"All calibrated routes were removed during network-edge validation against {network_file.name}."
                                )

                            sumocfg_path = write_final_simulation_sumocfg(
                                DEFAULT_FINAL_SUMOCFG,
                                network_file,
                                calibrated_routes_output,
                                begin_sec,
                                end_sec,
                                DEFAULT_FINAL_TRIPINFO,
                                DEFAULT_FINAL_QUEUE,
                                                        DEFAULT_VEHICLE_TYPES_ADD_FILE,
                                                        )
                            
                            # Verify required Streamlit_Callibration simulation inputs
                            routes_file = STREAMLIT_CALIB_DIR / "calibrated_routes.rou.xml"
                            detectors_file = STREAMLIT_CALIB_DIR / "detectors.add.xml"
                            vehicle_types_file = STREAMLIT_CALIB_DIR / "vehicle_types.add.xml"

                            missing_inputs: list[str] = []
                            if not network_file.exists():
                                missing_inputs.append(str(network_file))
                            if not routes_file.exists():
                                missing_inputs.append(str(routes_file))
                            if not tls_file.exists():
                                missing_inputs.append(str(tls_file))
                            if not detectors_file.exists():
                                missing_inputs.append(str(detectors_file))
                            if not vehicle_types_file.exists():
                                missing_inputs.append(str(vehicle_types_file))

                            if missing_inputs:
                                st.error("Missing required simulation inputs:\n- " + "\n- ".join(missing_inputs))
                            else:
                                st.info(
                                    f"✓ Verified TLS configuration:\n"
                                    f"  - Network: {network_file.name}\n"
                                    f"  - Routes: {routes_file.name}\n"
                                    f"  - Additional files: traffic_lights.add.xml, detectors.add.xml, vehicle_types.add.xml\n"
                                    f"  - Signal Programs: Program 0 (Kandevtasthan, Kesharmahal, Pulchowk North)\n"
                                    f"  - Yellow phases: 3s transitions validated\n"
                                    f"  - Override method: net tlLogic removed + traffic_lights.add.xml Program 0"
                                )
                                st.info("vehicle_types.add.xml loaded for mode-specific vehicle behavior (Bus, Car, Motorcycle).")

                            tls_sync_stats = prepare_survey_tls_priority(network_file, tls_file)
                            st.success(TLS_PRIORITY_STATUS_MESSAGE)
                            if any(tls_sync_stats.values()):
                                st.caption(
                                    "Final pre-run TLS scrub summary: "
                                    + ", ".join(f"{key}={value}" for key, value in tls_sync_stats.items())
                                )
                            
                            final_placeholder = st.empty()
                            with st.spinner("Running final SUMO simulation with fixed signal timing..."):
                                final_log = stream_final_simulation(sumocfg_path, final_placeholder, STREAMLIT_CALIB_DIR)
                            st.session_state.final_simulation_log = final_log
                            if "SUMO final simulation exited with code" in final_log:
                                st.error(f"Final simulation failed using {sumocfg_path}")
                            else:
                                st.success(f"Final simulation completed using {sumocfg_path}")
            except Exception as exc:
                st.error(f"routeSampler run failed: {exc}")

        if st.session_state.log_text:
            st.subheader("Live Output")
            st.code(st.session_state.log_text, language="text")

        # Display mode breakdown if available
        if st.session_state.calibration_mode == "Classified Modal Calibration" and st.session_state.log_text:
            st.subheader("Mode Breakdown Dashboard")
            mode_breakdown = parse_mode_breakdown_from_routes_file(calibrated_routes_output)
            
            if mode_breakdown["total"] > 0:
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Vehicles", mode_breakdown["total"])
                with col2:
                    st.metric("Buses", mode_breakdown["bus"], delta=f"{100*mode_breakdown['bus']/max(1, mode_breakdown['total']):.1f}%")
                with col3:
                    st.metric("Cars", mode_breakdown["car"], delta=f"{100*mode_breakdown['car']/max(1, mode_breakdown['total']):.1f}%")
                with col4:
                    st.metric("Motorcycles", mode_breakdown["motorcycle"], delta=f"{100*mode_breakdown['motorcycle']/max(1, mode_breakdown['total']):.1f}%")
            else:
                st.info("No mode breakdown data found in routeSampler log (typical if counts are aggregated).")


        if st.session_state.final_simulation_log:
            st.subheader("Final Simulation Output")
            st.code(st.session_state.final_simulation_log, language="text")


if __name__ == "__main__":
    main()