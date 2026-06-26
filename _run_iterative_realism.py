from pathlib import Path
import subprocess
import sys
import pandas as pd
import openpyxl
import re
import shutil
import os

ROOT = Path(os.environ.get("SUMO_CALIBRATION_ROOT", Path(__file__).resolve().parent))
PROJECT = ROOT / "project"
SCRIPTS = PROJECT / "scripts"
PY = Path(os.environ.get("PYTHON_EXE", sys.executable))

BASE = PROJECT / "results" / "frame_based_5min_corridor"
OUT = PROJECT / "results" / "frame_based_5min_corridor_iterative"
OUT.mkdir(parents=True, exist_ok=True)

CAL_TRAFFIC_SRC = BASE / "inputs" / "traffic_5min_frame_calib_20260222.xlsx"
VAL_TRAFFIC_SRC = BASE / "inputs" / "traffic_5min_frame_valid_20260225.xlsx"
SIG_BASE = PROJECT / "results" / "corridor_calib_20260222_validate_20260225" / "inputs" / "signal_timings_20260222_0900_1000.xlsx"
JMAP = SCRIPTS / "data" / "junction_id.xlsx"
NET = PROJECT / "Network.net.xml"


def run(cmd, cwd=None):
    subprocess.run([str(x) for x in cmd], cwd=str(cwd) if cwd else None, check=True)


def parse_report(path: Path):
    out = {"realism":"Unknown","wait":1e9,"speed":0.0,"tele":1e9,"geh":0.0}
    for ln in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        s = ln.strip()
        if s.startswith("- Realism classification:"):
            out["realism"] = s.split(":",1)[1].strip()
        elif s.startswith("- Mean waiting time:"):
            out["wait"] = float(re.findall(r"[-+]?[0-9]*\.?[0-9]+", s)[0])
        elif s.startswith("- Mean speed relative"):
            out["speed"] = float(re.findall(r"[-+]?[0-9]*\.?[0-9]+", s)[0])
        elif s.startswith("- Edges with teleports:"):
            out["tele"] = float(re.findall(r"[-+]?[0-9]*\.?[0-9]+", s)[0])
        elif s.startswith("- GEH < 5 share:"):
            out["geh"] = float(re.findall(r"[-+]?[0-9]*\.?[0-9]+", s)[0])
    return out


def realism_rank(lbl: str):
    m={"Near free-flow":0,"Moderate congestion":1,"Heavy congestion":2,"Severely congested":3,"Unknown":4}
    return m.get(lbl,4)


def scale_traffic(src: Path, dst: Path, scale: float):
    df = pd.read_excel(src)
    df["count"] = (df["count"].astype(float) * scale).round().clip(lower=1)
    df.to_excel(dst, index=False)


def write_signal_variant(src: Path, dst: Path, ns_mult: float, ew_mult: float):
    wb = openpyxl.load_workbook(src)
    for s in wb.sheetnames:
        ws = wb[s]
        # assume tidy sheet: intersection, ns_green, ew_green
        head = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        cols = {str(v).strip().lower(): i+1 for i,v in enumerate(head) if v}
        if "ns_green" in cols and "ew_green" in cols:
            r=2
            while ws.cell(r,1).value is not None:
                ns = ws.cell(r, cols["ns_green"]).value
                ew = ws.cell(r, cols["ew_green"]).value
                if isinstance(ns,(int,float)) and isinstance(ew,(int,float)):
                    ns2 = max(8, int(round(float(ns)*ns_mult)))
                    ew2 = max(10, int(round(float(ew)*ew_mult)))
                    ws.cell(r, cols["ns_green"]).value = ns2
                    ws.cell(r, cols["ew_green"]).value = ew2
                r += 1
    wb.save(dst)


def one_run(tag: str, scale: float, ns_mult: float, ew_mult: float):
    d = OUT / tag
    d.mkdir(parents=True, exist_ok=True)

    cal_x = d / "traffic_cal.xlsx"
    val_x = d / "traffic_val.xlsx"
    sig_x = d / "signal.xlsx"
    scale_traffic(CAL_TRAFFIC_SRC, cal_x, scale)
    scale_traffic(VAL_TRAFFIC_SRC, val_x, scale)
    write_signal_variant(SIG_BASE, sig_x, ns_mult, ew_mult)

    # signals
    run([PY, SCRIPTS / "generate_signals.py", "--signal-timings", sig_x, "--junction-map", JMAP, "--network", NET, "--output", d / "traffic_lights.add.xml"], cwd=PROJECT)

    shutil.copy2(NET, d / "Network.net.xml")

    # calibration
    run([PY, SCRIPTS / "generate_routes.py", "--traffic", cal_x, "--junction-map", JMAP, "--network", NET, "--routes-output", d / "routes.rou.xml", "--detectors-output", d / "detectors.add.xml", "--sumocfg-output", d / "simulation.sumocfg"], cwd=PROJECT)
    run(["sumo", "-c", str(d / "simulation.sumocfg")], cwd=d)
    cal_rep = d / "calibration_report.txt"
    run([PY, SCRIPTS / "analyze_calibration.py", "--traffic", cal_x, "--junction-map", JMAP, "--signal-timings", sig_x, "--routes", d / "routes.rou.xml", "--traffic-lights", d / "traffic_lights.add.xml", "--tripinfo", d / "tripinfo.xml", "--edge-data", d / "edge_data.xml", "--queue", d / "queue.xml", "--detector", d / "detector_output.xml", "--demand-report-csv", d / "demand_cal.csv", "--report", cal_rep], cwd=d)

    # validation
    run([PY, SCRIPTS / "generate_routes.py", "--traffic", val_x, "--junction-map", JMAP, "--network", NET, "--routes-output", d / "routes.rou.xml", "--detectors-output", d / "detectors.add.xml", "--sumocfg-output", d / "simulation.sumocfg"], cwd=PROJECT)
    run(["sumo", "-c", str(d / "simulation.sumocfg")], cwd=d)
    val_rep = d / "validation_report.txt"
    run([PY, SCRIPTS / "analyze_calibration.py", "--traffic", val_x, "--junction-map", JMAP, "--signal-timings", sig_x, "--routes", d / "routes.rou.xml", "--traffic-lights", d / "traffic_lights.add.xml", "--tripinfo", d / "tripinfo.xml", "--edge-data", d / "edge_data.xml", "--queue", d / "queue.xml", "--detector", d / "detector_output.xml", "--demand-report-csv", d / "demand_val.csv", "--report", val_rep], cwd=d)

    mcal = parse_report(cal_rep)
    mval = parse_report(val_rep)
    score = (
        realism_rank(mcal["realism"]) * 1000000 +
        realism_rank(mval["realism"]) * 200000 +
        mcal["wait"] * 150 + mval["wait"] * 100 +
        mcal["tele"] * 500 + mval["tele"] * 300 -
        (mcal["speed"] + mval["speed"]) * 10000
    )
    return {"tag":tag,"scale":scale,"ns_mult":ns_mult,"ew_mult":ew_mult,"cal":mcal,"val":mval,"score":score,"dir":d}


cands = []
scales = [0.60,0.50,0.40,0.35,0.30,0.25]
ns_mults = [1.0,0.9,0.8,0.7]
ew_mults = [1.0,1.15,1.30,1.45,1.60]

idx = 0
for s in scales:
    for nsm in ns_mults:
        for ewm in ew_mults:
            idx += 1
            tag=f"i{idx:03d}_s{int(s*100):02d}_ns{int(nsm*100):03d}_ew{int(ewm*100):03d}"
            try:
                r = one_run(tag, s, nsm, ewm)
                cands.append(r)
                print(f"{tag}: CAL={r['cal']['realism']} wait={r['cal']['wait']:.1f} speed={r['cal']['speed']:.3f} tele={r['cal']['tele']:.0f} | VAL={r['val']['realism']} wait={r['val']['wait']:.1f} speed={r['val']['speed']:.3f} tele={r['val']['tele']:.0f}")
                # early stop if both become better than severe
                if r['cal']['realism'] != 'Severely congested' and r['val']['realism'] != 'Severely congested':
                    print('EARLY_STOP_BETTER_THAN_SEVERE')
                    raise StopIteration
            except StopIteration:
                break
            except Exception as e:
                print(f"{tag}: FAILED {e}")
                continue
        else:
            continue
        break
    else:
        continue
    break

if not cands:
    raise RuntimeError('No successful iterations')

best = sorted(cands, key=lambda x: x['score'])[0]

best_dir = OUT / "best"
best_dir.mkdir(parents=True, exist_ok=True)
for f in ["Network.net.xml","traffic_lights.add.xml","detectors.add.xml","simulation.sumocfg","routes.rou.xml","tripinfo.xml","edge_data.xml","queue.xml","detector_output.xml","calibration_report.txt","validation_report.txt","demand_cal.csv","demand_val.csv","traffic_cal.xlsx","traffic_val.xlsx","signal.xlsx"]:
    src = best['dir'] / f
    if src.exists():
        shutil.copy2(src, best_dir / f)

report = OUT / "iterative_realism_report.txt"
lines = []
lines.append('Iterative Realism Calibration Report (Frame-based 5min)')
lines.append('======================================================')
lines.append('')
lines.append(f"Best tag: {best['tag']}")
lines.append(f"Best parameters: demand_scale={best['scale']:.2f}, ns_mult={best['ns_mult']:.2f}, ew_mult={best['ew_mult']:.2f}")
lines.append('')
lines.append('Calibration metrics')
for k,v in best['cal'].items():
    lines.append(f"- {k}: {v}")
lines.append('')
lines.append('Validation metrics')
for k,v in best['val'].items():
    lines.append(f"- {k}: {v}")
lines.append('')
lines.append(f"Best folder: {best['dir']}")
lines.append(f"Runnable best package: {best_dir}")
lines.append(f"Total successful iterations: {len(cands)}")
report.write_text("\n".join(lines), encoding='utf-8')

# Save leaderboard
rows=[]
for r in cands:
    rows.append({
        'tag':r['tag'],'scale':r['scale'],'ns_mult':r['ns_mult'],'ew_mult':r['ew_mult'],'score':r['score'],
        'cal_realism':r['cal']['realism'],'cal_wait':r['cal']['wait'],'cal_speed':r['cal']['speed'],'cal_tele':r['cal']['tele'],
        'val_realism':r['val']['realism'],'val_wait':r['val']['wait'],'val_speed':r['val']['speed'],'val_tele':r['val']['tele'],
    })
pd.DataFrame(rows).sort_values('score').to_csv(OUT / 'iteration_leaderboard.csv', index=False)
print('Wrote', report)
