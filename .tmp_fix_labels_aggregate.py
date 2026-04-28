from pathlib import Path
import re
from collections import OrderedDict
from openpyxl import load_workbook

base = Path(r"C:/Users/gupta/SumoCallibration")
target_root = base / "ProcessedVideoOutput" / "Krishna Marg (Kupondole Busstop)"

if not target_root.exists():
    print(f"Target folder not found: {target_root}")
    raise SystemExit(1)

TARGET_SHEETS = {"direction counts", "events"}

LONG_CANON = {
    "NORTH-TO-EAST": "North-to-East",
    "SOUTH-TO-EAST": "South-to-East",
    "EAST-TO-SOUTH": "East-to-South",
}
SHORT_CANON = {
    "NE": "NE",
    "SE": "SE",
    "ES": "ES",
}

LONG_ALIASES = {
    "NORTH-TO_WEST": "North-to-East",
    "NORTH-TO-EAST": "North-to-East",
    "NORTH-TO_EAST": "North-to-East",
    "NE": "North-to-East",

    "SOUTH-TO-EAST": "South-to-East",
    "SOUTH-TO_EAST": "South-to-East",
    "SE": "South-to-East",

    "EAST-TO-WEST": "East-to-South",
    "EAST-TO-SOUTH": "East-to-South",
    "EAST-TO_SOUTH": "East-to-South",
    "EW": "East-to-South",
    "ES": "East-to-South",

    "NW": "North-to-East",
}

SHORT_ALIASES = {
    "NORTH-TO-WEST": "NE",
    "NORTH-TO_WEST": "NE",
    "NORTH-TO-EAST": "NE",
    "NORTH-TO_EAST": "NE",
    "NE": "NE",
    "NW": "NE",

    "SOUTH-TO-EAST": "SE",
    "SOUTH-TO_EAST": "SE",
    "SE": "SE",

    "EAST-TO-WEST": "ES",
    "EAST-TO-SOUTH": "ES",
    "EAST-TO_SOUTH": "ES",
    "EW": "ES",
    "ES": "ES",
}

WORD_MAP_LONG = {
    "EAST-TO-WEST": "East-to-South",
    "NORTH-TO-WEST": "North-to-East",
    "NORTH-TO-EAST": "North-to-East",
    "SOUTH-TO-EAST": "South-to-East",
    "EAST-TO-SOUTH": "East-to-South",
    "EW": "East-to-South",
    "NW": "North-to-East",
    "NE": "North-to-East",
    "SE": "South-to-East",
    "ES": "East-to-South",
}

WORD_MAP_SHORT = {
    "EAST-TO-WEST": "ES",
    "NORTH-TO-WEST": "NE",
    "NORTH-TO-EAST": "NE",
    "SOUTH-TO-EAST": "SE",
    "EAST-TO-SOUTH": "ES",
    "EW": "ES",
    "NW": "NE",
    "NE": "NE",
    "SE": "SE",
    "ES": "ES",
}

TOKEN_RE = re.compile(r"(?<![A-Za-z0-9])(?:North-to[_-]West|North-to[_-]East|South-to[_-]East|East-to[_-]West|East-to[_-]South|EW|NW|NE|SE|ES)(?![A-Za-z0-9])", re.IGNORECASE)

def norm_key(v: str) -> str:
    return v.strip().replace("_", "-").upper()

def is_short_column(header: str) -> bool:
    h = header.strip().lower()
    return h in {"short", "short code", "shortcode", "code", "direction short", "direction_code", "direction code"}

def is_direction_column(header: str) -> bool:
    return "direction" in header.strip().lower()

def normalize_exact(value: str, mode: str):
    k = norm_key(value)
    if mode == "short":
        return SHORT_ALIASES.get(k)
    return LONG_ALIASES.get(k)

def normalize_text(value: str, mode: str):
    exact = normalize_exact(value, mode)
    if exact is not None:
        return exact
    mapping = WORD_MAP_SHORT if mode == "short" else WORD_MAP_LONG
    def repl(m):
        k = norm_key(m.group(0))
        return mapping.get(k, m.group(0))
    return TOKEN_RE.sub(repl, value)

def set_if_changed(cell, new_val):
    if cell.value != new_val:
        cell.value = new_val
        return 1
    return 0

files_scanned = 0
files_changed = 0
total_cell_edits = 0
total_merged_rows = 0
fallback_written = []
skipped_unreadable = []

for path in sorted(target_root.rglob("*.xlsx")):
    if path.name.startswith("~$"):
        continue

    files_scanned += 1
    try:
        wb = load_workbook(path)
    except Exception:
        skipped_unreadable.append(path)
        continue

    wb_changed = False
    wb_edits = 0

    for ws in wb.worksheets:
        sheet_name = ws.title.strip().lower()
        if sheet_name not in TARGET_SHEETS:
            continue

        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 1 or max_col < 1:
            continue

        headers = []
        header_lookup = {}
        for c in range(1, max_col + 1):
            hv = ws.cell(row=1, column=c).value
            hs = str(hv).strip() if hv is not None else ""
            headers.append(hs)
            if hs:
                header_lookup[hs.strip().lower()] = c

        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell.value, str):
                    continue
                header = headers[c - 1] if c - 1 < len(headers) else ""
                mode = "long"
                if is_short_column(header):
                    mode = "short"
                elif is_direction_column(header):
                    mode = "long"
                else:
                    exact_short = normalize_exact(cell.value, "short")
                    if exact_short is not None and cell.value.strip().upper().replace("_", "-") in {"NE","SE","ES","NW","EW"}:
                        mode = "short"
                new_val = normalize_text(cell.value, mode)
                edits = set_if_changed(cell, new_val)
                if edits:
                    wb_changed = True
                    wb_edits += edits

        if sheet_name == "direction counts":
            direction_col = header_lookup.get("direction")
            class_col = header_lookup.get("class")
            count_col = header_lookup.get("count")

            if direction_col and class_col and count_col:
                rows = []
                for r in range(2, ws.max_row + 1):
                    dir_cell = ws.cell(row=r, column=direction_col)
                    cls_cell = ws.cell(row=r, column=class_col)
                    cnt_cell = ws.cell(row=r, column=count_col)
                    direction = dir_cell.value
                    cls = cls_cell.value
                    cnt = cnt_cell.value
                    if direction is None and cls is None and cnt is None:
                        continue
                    direction_key = direction.strip() if isinstance(direction, str) else direction
                    class_key = cls.strip() if isinstance(cls, str) else cls
                    cnt_num = cnt
                    if isinstance(cnt_num, str):
                        try:
                            cnt_num = float(cnt_num.strip())
                        except Exception:
                            cnt_num = 0
                    if cnt_num is None:
                        cnt_num = 0
                    rows.append((direction_key, class_key, cnt_num))

                grouped = OrderedDict()
                for direction, cls, cnt in rows:
                    key = (direction, cls)
                    grouped[key] = grouped.get(key, 0) + cnt

                merged_rows = max(0, len(rows) - len(grouped))
                total_merged_rows += merged_rows

                out_rows = [(*k, v) for k, v in grouped.items()]
                start_row = 2
                end_needed = start_row + len(out_rows) - 1
                original_last = ws.max_row

                for idx, (direction, cls, cnt) in enumerate(out_rows, start=start_row):
                    wb_edits += set_if_changed(ws.cell(row=idx, column=direction_col), direction)
                    wb_edits += set_if_changed(ws.cell(row=idx, column=class_col), cls)
                    # preserve int if whole-number float
                    if isinstance(cnt, float) and cnt.is_integer():
                        cnt_out = int(cnt)
                    else:
                        cnt_out = cnt
                    wb_edits += set_if_changed(ws.cell(row=idx, column=count_col), cnt_out)

                clear_from = end_needed + 1
                if out_rows:
                    clear_from = end_needed + 1
                else:
                    clear_from = 2

                for r in range(clear_from, original_last + 1):
                    wb_edits += set_if_changed(ws.cell(row=r, column=direction_col), None)
                    wb_edits += set_if_changed(ws.cell(row=r, column=class_col), None)
                    wb_edits += set_if_changed(ws.cell(row=r, column=count_col), None)

                if merged_rows > 0 or wb_edits > 0:
                    wb_changed = wb_changed or (merged_rows > 0 or wb_edits > 0)

    if wb_changed and wb_edits > 0:
        try:
            wb.save(path)
            files_changed += 1
            total_cell_edits += wb_edits
        except PermissionError:
            fallback = path.with_name(f"{path.stem}_unlocked_fixed_labels{path.suffix}")
            wb.save(fallback)
            files_changed += 1
            total_cell_edits += wb_edits
            fallback_written.append(fallback)
        except Exception:
            try:
                fallback = path.with_name(f"{path.stem}_unlocked_fixed_labels{path.suffix}")
                wb.save(fallback)
                files_changed += 1
                total_cell_edits += wb_edits
                fallback_written.append(fallback)
            except Exception:
                pass

print(f"Files scanned: {files_scanned}")
print(f"Files changed: {files_changed}")
print(f"Total cell edits: {total_cell_edits}")
print(f"Aggregated rows merged: {total_merged_rows}")
print(f"Fallback files written: {len(fallback_written)}")
for p in fallback_written:
    print(f"- {p}")
