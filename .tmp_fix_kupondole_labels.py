from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

base = Path(r"C:/Users/gupta/SumoCallibration")
target = base / "ProcessedVideoOutput" / "Krishna Marg (Kupondole Busstop)"
SHEETS = {"direction counts", "events"}

if not target.exists():
    print(f"Target folder not found: {target}")
    raise SystemExit(1)


def normkey(s):
    return s.strip().upper().replace("_", "-")


def normalize_as_direction_text(v):
    if not isinstance(v, str):
        return v, False
    t = v.strip()
    if not t:
        return v, False
    k = normkey(t)
    if k in {"NORTH-TO-EAST", "NE", "NORTH-TO-WEST", "NW"}:
        n = "North-to-East"
    elif k in {"SOUTH-TO-EAST", "SE"}:
        n = "South-to-East"
    elif k in {"EAST-TO-SOUTH", "ES", "EAST-TO-WEST", "EW"}:
        n = "East-to-South"
    else:
        n = t
    return n, n != t


def normalize_as_code(v):
    if not isinstance(v, str):
        return v, False
    t = v.strip()
    if not t:
        return v, False
    k = normkey(t)
    if k in {"NORTH-TO-EAST", "NE", "NORTH-TO-WEST", "NW"}:
        n = "NE"
    elif k in {"SOUTH-TO-EAST", "SE"}:
        n = "SE"
    elif k in {"EAST-TO-SOUTH", "ES", "EAST-TO-WEST", "EW"}:
        n = "ES"
    else:
        n = t
    return n, n != t


def parse_count(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", ""))
    except Exception:
        return 0.0


def format_count(v):
    return int(round(v)) if abs(v - round(v)) < 1e-9 else float(v)


def infer_code_columns(ws):
    code_cols = set()
    max_col = ws.max_column
    max_row = ws.max_row

    headers = {}
    for c in range(1, max_col + 1):
        hv = ws.cell(1, c).value
        if hv is not None:
            h = str(hv).strip().lower()
            headers[c] = h
            if "code" in h or h in {"dir", "direction code", "movement code"}:
                code_cols.add(c)

    alias_keys = {
        "NE", "NW", "SE", "ES", "EW",
        "NORTH-TO-EAST", "NORTH-TO-WEST", "SOUTH-TO-EAST", "EAST-TO-SOUTH", "EAST-TO-WEST",
        "NORTH-TO_EAST", "SOUTH-TO_EAST", "EAST-TO_SOUTH"
    }

    for c in range(1, max_col + 1):
        if c in code_cols:
            continue
        total = 0
        alias_like = 0
        for r in range(2, max_row + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                total += 1
                if v.strip().upper() in alias_keys or v.strip().upper().replace("_", "-") in alias_keys:
                    alias_like += 1
        if total > 0 and alias_like / total >= 0.8:
            code_cols.add(c)

    return headers, code_cols


def normalize_sheet_labels(ws):
    changes = 0
    headers, code_cols = infer_code_columns(ws)

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if not isinstance(cell.value, str):
                continue
            if c in code_cols:
                nv, ch = normalize_as_code(cell.value)
            else:
                nv, ch = normalize_as_direction_text(cell.value)
            if ch:
                cell.value = nv
                changes += 1

    return changes


def aggregate_direction_counts(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        hv = ws.cell(1, c).value
        if hv is not None:
            headers[str(hv).strip().lower()] = c

    dcol = headers.get("direction")
    ccol = headers.get("class")
    ncol = headers.get("count")
    if not (dcol and ccol and ncol):
        return 0, 0

    grouped = defaultdict(float)
    class_display = {}
    order = []

    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, dcol).value
        cl = ws.cell(r, ccol).value
        ct = ws.cell(r, ncol).value
        dn, _ = normalize_as_direction_text(d if d is not None else "")
        dstr = str(dn).strip() if dn is not None else ""
        if dstr == "":
            continue
        cls = "" if cl is None else str(cl).strip()
        key = (dstr, cls.lower())
        if key not in grouped:
            order.append(key)
            class_display[key] = cls
        grouped[key] += parse_count(ct)

    original_rows = sum(1 for r in range(2, ws.max_row + 1) if str(ws.cell(r, dcol).value).strip() not in {"", "None"})
    merged = max(0, original_rows - len(grouped))

    old = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    for i, key in enumerate(order, start=2):
        ws.cell(i, dcol).value = key[0]
        ws.cell(i, ccol).value = class_display[key]
        ws.cell(i, ncol).value = format_count(grouped[key])

    new = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, 2 + len(old))]
    rewrite_changes = sum(1 for ro, rn in zip(old, new) for ov, nv in zip(ro, rn) if ov != nv)
    return rewrite_changes, merged


files = sorted([p for p in target.rglob("*.xlsx") if not p.name.startswith("~$")])

scanned = 0
changed_files = 0
total_cell_changes = 0
total_merged = 0
per_file = {}
fallbacks = []

for path in files:
    scanned += 1
    try:
        wb = load_workbook(path)
    except Exception:
        per_file[str(path)] = "skipped(load_error)"
        continue

    file_changes = 0
    file_merged = 0

    for ws in wb.worksheets:
        t = ws.title.strip().lower()
        if t not in SHEETS:
            continue
        file_changes += normalize_sheet_labels(ws)
        if t == "direction counts":
            rc, mg = aggregate_direction_counts(ws)
            file_changes += rc
            file_merged += mg

    if file_changes > 0 or file_merged > 0:
        changed_files += 1
        total_cell_changes += file_changes
        total_merged += file_merged
        per_file[str(path)] = f"cells={file_changes}, merged_rows={file_merged}"
        try:
            wb.save(path)
        except Exception:
            fb = path.with_name(f"{path.stem}_unlocked_fixed_labels.xlsx")
            wb.save(fb)
            fallbacks.append(str(fb))
    else:
        per_file[str(path)] = "cells=0, merged_rows=0"

print("=== Processing Report ===")
print(f"total files scanned: {scanned}")
print(f"files changed: {changed_files}")
print(f"total cell changes: {total_cell_changes}")
print(f"aggregated duplicate rows merged: {total_merged}")
print("per-file change counts:")
for k in sorted(per_file):
    print(f"- {k}: {per_file[k]}")
print("fallback paths:")
if fallbacks:
    for p in fallbacks:
        print(f"- {p}")
else:
    print("- (none)")

# verification
bad1 = 0
bad2 = 0
bad3 = 0
o1 = set(); o2 = set(); o3 = set()

for path in files:
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        continue
    for ws in wb.worksheets:
        t = ws.title.strip().lower()
        if t not in SHEETS:
            continue
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                s = v.strip()
                su = s.upper()
                if s == "East-to-West" or s == "North-to-West" or su == "EW" or su == "NW":
                    bad1 += 1
                    o1.add(str(path))
                if s in {"North-to_East", "South-to_East", "East-to_South"}:
                    bad2 += 1
                    o2.add(str(path))
        if t == "direction counts":
            hdr = {}
            for c in range(1, ws.max_column + 1):
                hv = ws.cell(1, c).value
                if hv is not None:
                    hdr[str(hv).strip().lower()] = c
            dcol = hdr.get("direction")
            ccol = hdr.get("class")
            if dcol and ccol:
                seen = set(); dups = 0
                for r in range(2, ws.max_row + 1):
                    dv = ws.cell(r, dcol).value
                    cv = ws.cell(r, ccol).value
                    if (dv is None or str(dv).strip()=="") and (cv is None or str(cv).strip()==""):
                        continue
                    k = (str(dv).strip(), str(cv).strip().lower())
                    if k in seen:
                        dups += 1
                    else:
                        seen.add(k)
                if dups:
                    bad3 += dups
                    o3.add(str(path))

print("=== Verification Report ===")
print(f"total East-to-West/North-to-West/EW/NW found: {bad1}")
print(f"total underscore variants found: {bad2}")
print(f"total duplicate (Direction, Class) rows in Direction Counts: {bad3}")
print("offending files for check 1:")
if o1:
    for p in sorted(o1):
        print(f"- {p}")
else:
    print("- (none)")
print("offending files for check 2:")
if o2:
    for p in sorted(o2):
        print(f"- {p}")
else:
    print("- (none)")
print("offending files for check 3:")
if o3:
    for p in sorted(o3):
        print(f"- {p}")
else:
    print("- (none)")
