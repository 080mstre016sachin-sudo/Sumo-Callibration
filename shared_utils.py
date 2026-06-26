"""Shared utility functions for SUMO Calibration scripts.

This module consolidates common patterns duplicated across
streamlit_app.py, generate_edgewise_from_raw.py,
process_20260222_reports.py, and summarize_zone_maps.py.
"""

from __future__ import annotations

import re
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Time parsing
# ---------------------------------------------------------------------------

def to_seconds(value: object) -> int | None:
    """Parse a HH:MM:SS time string into total seconds.

    Accepts common separators (colon, dot, underscore).
    Returns None when parsing fails.
    """
    text = str(value).strip().replace(".", ":").replace("_", ":")
    parts = text.split(":")
    if len(parts) != 3:
        return None
    try:
        h, m, s = (int(part) for part in parts)
    except ValueError:
        return None
    if not (0 <= h <= 23 and 0 <= m <= 59 and 0 <= s <= 59):
        return None
    return h * 3600 + m * 60 + s


def format_seconds(total_seconds: int) -> str:
    """Format total seconds as HH:MM:SS."""
    hh = total_seconds // 3600
    mm = (total_seconds % 3600) // 60
    ss = total_seconds % 60
    return f"{hh:02d}:{mm:02d}:{ss:02d}"


# ---------------------------------------------------------------------------
# Date / token normalization
# ---------------------------------------------------------------------------

def normalize_date_token(value: object) -> str | None:
    """Normalize a date string or compact token to ISO YYYY-MM-DD format.

    Handles formats like 20260222, 2026-02-22, 22-02-2026, 22/02/2026, etc.
    Returns None if parsing fails.
    """
    text = str(value).strip()
    if not text:
        return None

    compact = re.sub(r"\D", "", text)
    if len(compact) == 8:
        try:
            if compact.startswith("20"):
                return datetime.strptime(compact, "%Y%m%d").strftime("%Y-%m-%d")
            return datetime.strptime(compact, "%d%m%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass

    normalized = re.sub(r"[\s_/\\.]+", "-", text)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(normalized, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Interval parsing from filenames
# ---------------------------------------------------------------------------

_TIME_RANGE_RE = re.compile(
    r"(\d{1,2})[.:_](\d{2})(?:[.:_](\d{2}))?\s*-\s*(\d{1,2})[.:_](\d{2})(?:[.:_](\d{2}))?"
)
_NUMERIC_STEM_RE = re.compile(r"^\d{3,4}$")


def parse_interval_from_filename(path: Path) -> tuple[str | None, str | None]:
    """Extract (start_time, end_time) from a filename.

    Supports both explicit time-range patterns (e.g., '9.00-9.05') and
    numeric shorthand codes (e.g., '905' meaning 09:15:00-09:20:00 with +10m offset).

    Returns (start_hh:mm:ss, end_hh:mm:ss) or (None, None).
    """
    stem = path.stem.strip()

    match = _TIME_RANGE_RE.search(stem)
    if match:
        h1, m1, s1, h2, m2, s2 = match.groups()
        try:
            start = datetime(2000, 1, 1, int(h1), int(m1), int(s1 or 0))
            end = datetime(2000, 1, 1, int(h2), int(m2), int(s2 or 0))
            return start.strftime("%H:%M:%S"), end.strftime("%H:%M:%S")
        except ValueError:
            return None, None

    if _NUMERIC_STEM_RE.fullmatch(stem):
        token = stem.zfill(4)
        hh = int(token[:2])
        mm = int(token[2:])
        try:
            base = datetime(2000, 1, 1, hh, mm, 0)
        except ValueError:
            return None, None
        start = base + timedelta(minutes=10)
        end = start + timedelta(minutes=5)
        return start.strftime("%H:%M:%S"), end.strftime("%H:%M:%S")

    return None, None


# ---------------------------------------------------------------------------
# Text normalization helpers
# ---------------------------------------------------------------------------

def normalize_key(value: object) -> str:
    """Normalize a value to a lowercase alphanumeric key for fuzzy matching."""
    return "".join(ch.lower() for ch in str(value).strip() if ch.isalnum())


def normalize_direction(value: object) -> str:
    """Normalize direction text by collapsing whitespace."""
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def is_total_text(value: object) -> bool:
    """Check if a value represents a 'total' row to be excluded."""
    return str(value or "").strip().lower() in {"total", "grandtotal"}


# ---------------------------------------------------------------------------
# Excel utilities
# ---------------------------------------------------------------------------

def autosize_columns(sheet, max_width: int = 26) -> None:
    """Auto-size worksheet columns based on content length."""
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, max_width)


def write_dataframe_to_sheet(
    workbook: Workbook,
    sheet_name: str,
    dataframe: pd.DataFrame,
    freeze_panes: str = "A2",
) -> None:
    """Write a DataFrame to a new worksheet with bold headers and auto-sized columns.

    If a sheet with the same name exists, it is replaced.
    """
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    sheet = workbook.create_sheet(sheet_name)
    sheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        sheet.append(list(row))

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    autosize_columns(sheet)
    sheet.freeze_panes = freeze_panes


# ---------------------------------------------------------------------------
# Interval helpers
# ---------------------------------------------------------------------------

def interval_in_window(interval: str, begin_sec: int, end_sec: int) -> bool:
    """Check if a 'HH:MM:SS-HH:MM:SS' interval falls entirely within a time window."""
    start_text, _, end_text = interval.partition("-")
    start_sec = to_seconds(start_text)
    finish_sec = to_seconds(end_text)
    if start_sec is None or finish_sec is None:
        return False
    return start_sec >= begin_sec and finish_sec <= end_sec


def canonicalize_to_5min_interval(interval: str) -> str | None:
    """Snap an interval string to the nearest aligned 5-minute boundary."""
    start_text, _, _end_text = interval.partition("-")
    start_sec = to_seconds(start_text)
    if start_sec is None:
        return None
    canonical_start = (start_sec // 300) * 300
    canonical_end = canonical_start + 300
    if canonical_end > 24 * 3600:
        return None
    return f"{format_seconds(canonical_start)}-{format_seconds(canonical_end)}"
