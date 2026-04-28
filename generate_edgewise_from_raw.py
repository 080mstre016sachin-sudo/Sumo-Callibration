from __future__ import annotations

import re
import shutil
import argparse
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

WORKSPACE_ROOT = Path(__file__).resolve().parent
RAW_ROOT = WORKSPACE_ROOT / "ProcessedVideoOutput"
OUTPUT_PATH = RAW_ROOT / "EdgeWise_DirectionalCounts_5MinIntervalColumns_Datewise.xlsx"

ALLOWED_FOLDERS = [
    "Thapathali",
    "Jwagal",
    "Kandevtasthan",
    "Kesharmahal",
    "Krishna Marg (Kupondole Busstop)",
    "Krishnagalli",
    "Maitri Marg (Bhakundole)",
    "Patan Dhoka Road",
    "Pulchowk North",
    "Pulchowk South",
]

LOCATION_LABEL_MAP = {
    "Thapathali": "Thapathali",
    "Krishna Marg (Kupondole Busstop)": "Kupondole Busstop",
    "Kandevtasthan": "Kandevtasthan",
    "Jwagal": "Jwagal",
    "Patan Dhoka Road": "Patan Dhoka Road",
    "Maitri Marg (Bhakundole)": "Bhakundole",
    "Krishnagalli": "Krishnagalli",
    "Kesharmahal": "Kesharmahal",
    "Pulchowk North": "Pulchowk North",
    "Pulchowk South": "Pulchowk South",
}
LOCATION_FOLDER_KEY_MAP = {name.strip().lower(): name for name in LOCATION_LABEL_MAP}

LOCATION_ORDER = [
    "Thapathali",
    "Kupondole Busstop",
    "Kandevtasthan",
    "Jwagal",
    "Patan Dhoka Road",
    "Bhakundole",
    "Krishnagalli",
    "Kesharmahal",
    "Pulchowk North",
    "Pulchowk South",
]
LOCATION_ORDER_IDX = {name: idx for idx, name in enumerate(LOCATION_ORDER)}

TIME_RANGE_RE = re.compile(r"(\d{1,2})[.:_](\d{2})(?:[.:_](\d{2}))?\s*-\s*(\d{1,2})[.:_](\d{2})(?:[.:_](\d{2}))?")
NUMERIC_STEM_RE = re.compile(r"^\d{3,4}$")
SUMMARY_MARKERS = ("compile", "summary", "edgewise", "datewise", "compact")
TRAILING_STAMP_RE = re.compile(r"(\d{8})_(\d{6})$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate EdgeWise directional workbook from raw folders.")
    parser.add_argument(
        "--extra-root",
        action="append",
        default=[],
        help="Optional additional raw root path (can be passed multiple times), e.g. f:/Output",
    )
    return parser.parse_args()


def normalize_date_token(value: str) -> str | None:
    text = str(value).strip()
    if not text:
        return None

    compact = re.sub(r"\D", "", text)
    if len(compact) == 8:
        try:
            if compact.startswith("20"):
                return datetime.strptime(compact, "%Y%m%d").strftime("%Y-%m-%d")
            return datetime.strptime(compact, "%d%m%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass

    normalized = re.sub(r"[\s_/\\.]+", "-", text)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(normalized, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_interval_from_filename(path: Path) -> str | None:
    stem = path.stem.strip()
    match = TIME_RANGE_RE.search(stem)
    if match:
        h1, m1, s1, h2, m2, s2 = match.groups()
        try:
            start = datetime(2000, 1, 1, int(h1), int(m1), int(s1 or 0))
            end = datetime(2000, 1, 1, int(h2), int(m2), int(s2 or 0))
            return f"{start.strftime('%H:%M:%S')}-{end.strftime('%H:%M:%S')}"
        except ValueError:
            return None

    if NUMERIC_STEM_RE.fullmatch(stem):
        token = stem.zfill(4)
        hh = int(token[:2])
        mm = int(token[2:])
        try:
            base = datetime(2000, 1, 1, hh, mm, 0)
        except ValueError:
            return None
        start = base + timedelta(minutes=10)
        end = start + timedelta(minutes=5)
        return f"{start.strftime('%H:%M:%S')}-{end.strftime('%H:%M:%S')}"

    return None


def classify_filename_priority(path: Path) -> int:
    name = path.name.lower()
    if "approach_directional" in name:
        return 4
    if "zone_map" in name:
        return 3
    if "direction" in name:
        return 2
    if NUMERIC_STEM_RE.fullmatch(path.stem.strip()):
        return 1
    return 0


def parse_trailing_timestamp(path: Path) -> int:
    stem = path.stem.strip()
    match = TRAILING_STAMP_RE.search(stem)
    if not match:
        return -1
    date_part, time_part = match.groups()
    try:
        dt = datetime.strptime(f"{date_part}{time_part}", "%Y%m%d%H%M%S")
    except ValueError:
        return -1
    return int(dt.timestamp())


def choose_best_candidate(candidates: list[dict[str, object]]) -> dict[str, object]:
    def score(item: dict[str, object]) -> tuple[int, int, int, str]:
        # Prefer primary root, then richer filename type, then latest timestamp.
        root_idx = int(item["root_idx"])  # 0 = primary root
        file_priority = classify_filename_priority(Path(str(item["path"])))
        timestamp = parse_trailing_timestamp(Path(str(item["path"])))
        return (-root_idx, file_priority, timestamp, str(item["path"]))

    return sorted(candidates, key=score, reverse=True)[0]


def to_seconds(value: str) -> int | None:
    text = str(value).strip()
    parts = text.split(":")
    if len(parts) != 3:
        return None
    try:
        h, m, s = (int(part) for part in parts)
    except ValueError:
        return None
    if not (0 <= h <= 23 and 0 <= m <= 59 and 0 <= s <= 59):
        return None
    return h * 3600 + m * 60 + s


def interval_in_window(interval: str, begin_sec: int, end_sec: int) -> bool:
    start_text, _, end_text = interval.partition("-")
    start_sec = to_seconds(start_text)
    finish_sec = to_seconds(end_text)
    if start_sec is None or finish_sec is None:
        return False
    return start_sec >= begin_sec and finish_sec <= end_sec


def canonicalize_to_5min_interval(interval: str) -> str | None:
    start_text, _, _end_text = interval.partition("-")
    start_sec = to_seconds(start_text)
    if start_sec is None:
        return None
    canonical_start = (start_sec // 300) * 300
    canonical_end = canonical_start + 300
    if canonical_end > 24 * 3600:
        return None

    def fmt(total_seconds: int) -> str:
        hh = total_seconds // 3600
        mm = (total_seconds % 3600) // 60
        ss = total_seconds % 60
        return f"{hh:02d}:{mm:02d}:{ss:02d}"

    return f"{fmt(canonical_start)}-{fmt(canonical_end)}"


def normalize_direction(value: object) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def resolve_location_label(folder_name: str) -> str | None:
    key = normalize_direction(folder_name).lower()
    canonical_folder = LOCATION_FOLDER_KEY_MAP.get(key)
    if not canonical_folder:
        return None
    return LOCATION_LABEL_MAP.get(canonical_folder)


def is_total_text(value: object) -> bool:
    return str(value or "").strip().lower() == "total"


def get_sheet_case_insensitive(workbook, name: str):
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().lower() == name.strip().lower():
            return workbook[sheet_name]
    return None


def parse_direction_counts_sheet(sheet, fallback_interval: str | None) -> list[dict[str, object]]:
    header = [str(v).strip().lower() if v is not None else "" for v in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())]
    if not header:
        return []
    idx = {name: pos for pos, name in enumerate(header)}
    if not {"direction", "class", "count"}.issubset(idx):
        return []

    out: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: {"bus": 0.0, "car": 0.0, "motorcycle": 0.0})
    for row in sheet.iter_rows(min_row=2, values_only=True):
        direction = normalize_direction(row[idx["direction"]] if idx["direction"] < len(row) else "")
        vehicle_class = str(row[idx["class"]] if idx["class"] < len(row) else "").strip().lower()
        count_raw = row[idx["count"]] if idx["count"] < len(row) else 0

        if not direction or is_total_text(direction):
            continue

        if "bus" in vehicle_class:
            bucket = "bus"
        elif "car" in vehicle_class:
            bucket = "car"
        elif "motor" in vehicle_class or "bike" in vehicle_class:
            bucket = "motorcycle"
        else:
            continue

        try:
            count_val = float(count_raw)
        except (TypeError, ValueError):
            count_val = 0.0
        interval = fallback_interval or ""
        if not interval:
            continue
        out[(direction, interval)][bucket] += max(0.0, count_val)

    return [
        {
            "Direction": direction,
            "Interval": interval,
            "bus": values["bus"],
            "car": values["car"],
            "motorcycle": values["motorcycle"],
        }
        for (direction, interval), values in out.items()
    ]


def parse_summary_sheet(sheet, fallback_interval: str | None) -> list[dict[str, object]]:
    header = [str(v).strip().lower() if v is not None else "" for v in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())]
    if not header:
        return []
    idx = {name: pos for pos, name in enumerate(header)}
    has_wide = {"direction", "bus", "car", "motorcycle"}.issubset(idx)
    has_long = {"direction", "vehicleclass", "count"}.issubset(idx)
    if not (has_wide or has_long):
        return []

    out: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: {"bus": 0.0, "car": 0.0, "motorcycle": 0.0})
    for row in sheet.iter_rows(min_row=2, values_only=True):
        direction = normalize_direction(row[idx["direction"]] if idx["direction"] < len(row) else "")
        if not direction or is_total_text(direction):
            continue

        row_interval = fallback_interval or ""
        if "timestart" in idx and "timeend" in idx:
            start = normalize_direction(row[idx["timestart"]] if idx["timestart"] < len(row) else "")
            end = normalize_direction(row[idx["timeend"]] if idx["timeend"] < len(row) else "")
            if start and end and to_seconds(start) is not None and to_seconds(end) is not None:
                row_interval = f"{start}-{end}"
        if not row_interval:
            continue

        if has_wide:
            try:
                bus = max(0.0, float(row[idx["bus"]] if idx["bus"] < len(row) else 0))
            except (TypeError, ValueError):
                bus = 0.0
            try:
                car = max(0.0, float(row[idx["car"]] if idx["car"] < len(row) else 0))
            except (TypeError, ValueError):
                car = 0.0
            try:
                motorcycle = max(0.0, float(row[idx["motorcycle"]] if idx["motorcycle"] < len(row) else 0))
            except (TypeError, ValueError):
                motorcycle = 0.0

            out[(direction, row_interval)]["bus"] += bus
            out[(direction, row_interval)]["car"] += car
            out[(direction, row_interval)]["motorcycle"] += motorcycle
            continue

        vehicle_class = normalize_direction(row[idx["vehicleclass"]] if idx["vehicleclass"] < len(row) else "").lower()
        count_raw = row[idx["count"]] if idx["count"] < len(row) else 0
        if "bus" in vehicle_class:
            bucket = "bus"
        elif "car" in vehicle_class:
            bucket = "car"
        elif "motor" in vehicle_class or "bike" in vehicle_class:
            bucket = "motorcycle"
        else:
            continue
        try:
            count_val = max(0.0, float(count_raw))
        except (TypeError, ValueError):
            count_val = 0.0
        out[(direction, row_interval)][bucket] += count_val

    return [
        {
            "Direction": direction,
            "Interval": interval,
            "bus": values["bus"],
            "car": values["car"],
            "motorcycle": values["motorcycle"],
        }
        for (direction, interval), values in out.items()
    ]


def extract_direction_counts(path: Path, fallback_interval: str | None) -> list[dict[str, object]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []

    sheet = get_sheet_case_insensitive(workbook, "Direction Counts")
    parsed = parse_direction_counts_sheet(sheet, fallback_interval) if sheet is not None else []
    if parsed:
        return parsed

    summary_sheet = get_sheet_case_insensitive(workbook, "Summary")
    parsed = parse_summary_sheet(summary_sheet, fallback_interval) if summary_sheet is not None else []
    return parsed


def approach_from_direction(direction: str) -> str:
    if "-to-" in direction:
        return direction.split("-to-", 1)[0].strip()
    return direction.strip()


def discover_raw_records(raw_roots: list[Path]) -> tuple[pd.DataFrame, dict[str, object]]:
    begin_sec = to_seconds("08:00:00") or 0
    end_sec = to_seconds("10:00:00") or 24 * 3600

    rows: list[dict[str, object]] = []
    included_files = 0
    skipped_bad_structure = 0
    skipped_outside_window = 0
    skipped_non_allowed = 0
    scanned_files = 0
    dropped_duplicate_candidates = 0

    grouped_candidates: dict[tuple[str, str, str], list[dict[str, object]]] = defaultdict(list)

    for root_idx, root in enumerate(raw_roots):
        if not root.exists():
            continue

        for path in root.rglob("*.xlsx"):
            if path.name.startswith("~$"):
                continue
            scanned_files += 1

            rel = path.relative_to(root)
            if len(rel.parts) != 4:
                skipped_bad_structure += 1
                continue

            folder, date_token, _interval_folder, _file = rel.parts
            location_label = resolve_location_label(folder)
            if not location_label:
                skipped_non_allowed += 1
                continue

            lower = path.name.lower()
            if any(marker in lower for marker in SUMMARY_MARKERS):
                continue

            date_norm = normalize_date_token(date_token)
            if not date_norm:
                continue

            interval_label = parse_interval_from_filename(path)

            dedupe_key = interval_label or rel.as_posix()
            key = (location_label, date_norm, dedupe_key)
            grouped_candidates[key].append(
                {
                    "root": root,
                    "root_idx": root_idx,
                    "path": path,
                    "rel": rel,
                    "interval_label": interval_label,
                }
            )

    for key, candidates in grouped_candidates.items():
        chosen = choose_best_candidate(candidates)
        dropped_duplicate_candidates += max(0, len(candidates) - 1)

        path = Path(str(chosen["path"]))
        rel = Path(str(chosen["rel"]))
        root = Path(str(chosen["root"]))
        location_label, date_norm, _dedupe_key = key
        interval_label = str(chosen.get("interval_label") or "")

        parsed = extract_direction_counts(path, interval_label if interval_label else None)
        if not parsed:
            continue

        included_files += 1
        if root.resolve() == RAW_ROOT.resolve():
            source_rel = rel.as_posix()
        else:
            source_rel = f"[extra:{root.as_posix()}]/" + rel.as_posix()

        file_has_in_window_rows = False
        for item in parsed:
            direction = str(item["Direction"])
            row_interval = str(item.get("Interval") or interval_label or "")
            row_interval = canonicalize_to_5min_interval(row_interval) if row_interval else None
            if not row_interval or not interval_in_window(row_interval, begin_sec, end_sec):
                continue
            file_has_in_window_rows = True
            approach = approach_from_direction(direction)
            if is_total_text(direction) or is_total_text(approach):
                continue
            rows.append(
                {
                    "Location": location_label,
                    "Date": date_norm,
                    "Approach": approach,
                    "Direction": direction,
                    "VehicleClass": "Bus",
                    "Count": float(item.get("bus", 0)),
                    "Interval": row_interval,
                    "SourceWorkbook": source_rel,
                }
            )
            rows.append(
                {
                    "Location": location_label,
                    "Date": date_norm,
                    "Approach": approach,
                    "Direction": direction,
                    "VehicleClass": "Car",
                    "Count": float(item.get("car", 0)),
                    "Interval": row_interval,
                    "SourceWorkbook": source_rel,
                }
            )
            rows.append(
                {
                    "Location": location_label,
                    "Date": date_norm,
                    "Approach": approach,
                    "Direction": direction,
                    "VehicleClass": "Motorcycle",
                    "Count": float(item.get("motorcycle", 0)),
                    "Interval": row_interval,
                    "SourceWorkbook": source_rel,
                }
            )
        if not file_has_in_window_rows:
            included_files -= 1
            skipped_outside_window += 1

    diagnostics = {
        "scanned_files": scanned_files,
        "raw_roots": [root.as_posix() for root in raw_roots],
        "included_files": included_files,
        "dropped_duplicate_candidates": dropped_duplicate_candidates,
        "skipped_bad_structure": skipped_bad_structure,
        "skipped_outside_window": skipped_outside_window,
        "skipped_non_allowed": skipped_non_allowed,
    }

    frame = pd.DataFrame.from_records(rows)
    return frame, diagnostics


def build_output_sheets(raw_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    key_cols = ["Location", "Date", "Approach", "Direction"]
    if raw_df.empty:
        empty = pd.DataFrame(columns=key_cols + ["SourceWorkbook"])
        return {
            "Bus_5MinCols": empty.copy(),
            "Car_5MinCols": empty.copy(),
            "Motorcycle_5MinCols": empty.copy(),
            "Total_Volume_5MinCols": empty.copy(),
        }

    work = raw_df.copy()
    work = work[(~work["Direction"].map(is_total_text)) & (~work["Approach"].map(is_total_text))]

    src = (
        work.groupby(key_cols + ["Interval"], as_index=False)["SourceWorkbook"]
        .agg(lambda values: " | ".join(sorted({str(v).strip() for v in values if str(v).strip()})))
    )

    intervals = sorted(work["Interval"].unique().tolist(), key=lambda s: (to_seconds(s.split("-")[0]) or 10**9, s))

    def make_sheet(vehicle_classes: set[str]) -> pd.DataFrame:
        subset = work[work["VehicleClass"].isin(vehicle_classes)].copy()
        grouped = (
            subset.groupby(key_cols + ["Interval"], as_index=False)["Count"]
            .sum()
        )

        pivot = grouped.pivot_table(
            index=key_cols,
            columns="Interval",
            values="Count",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()
        pivot.columns = [str(col) for col in pivot.columns]

        src_group = (
            src[src["SourceWorkbook"].astype(str).str.strip() != ""]
            .groupby(key_cols, as_index=False)["SourceWorkbook"]
            .agg(lambda values: " | ".join(sorted({v for v in values if v})))
        )

        out = pivot.merge(src_group, on=key_cols, how="left")
        out["SourceWorkbook"] = out["SourceWorkbook"].fillna("")
        for col in intervals:
            if col not in out.columns:
                out[col] = 0
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0).astype(int)

        out = out[key_cols + ["SourceWorkbook"] + intervals]
        out = out[(~out["Direction"].map(is_total_text)) & (~out["Approach"].map(is_total_text))]
        out = out.sort_values(
            by=["Location", "Date", "Approach", "Direction"],
            key=lambda series: series.map(lambda x: LOCATION_ORDER_IDX.get(str(x), 999) if series.name == "Location" else str(x).lower()),
        )
        return out

    bus = make_sheet({"Bus"})
    car = make_sheet({"Car"})
    motorcycle = make_sheet({"Motorcycle"})
    total = make_sheet({"Bus", "Car", "Motorcycle"})

    return {
        "Bus_5MinCols": bus,
        "Car_5MinCols": car,
        "Motorcycle_5MinCols": motorcycle,
        "Total_Volume_5MinCols": total,
    }


def write_workbook(sheets: dict[str, pd.DataFrame]) -> tuple[Path, Path | None]:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    fallback_path: Path | None = None
    try:
        with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
            for sheet_name, frame in sheets.items():
                frame.to_excel(writer, sheet_name=sheet_name, index=False)
        return OUTPUT_PATH, None
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = OUTPUT_PATH.with_name(f"{OUTPUT_PATH.stem}_unlocked_{stamp}{OUTPUT_PATH.suffix}")
        with pd.ExcelWriter(fallback_path, engine="openpyxl") as writer:
            for sheet_name, frame in sheets.items():
                frame.to_excel(writer, sheet_name=sheet_name, index=False)
        try:
            shutil.copy2(fallback_path, OUTPUT_PATH)
        except Exception:
            pass
        return OUTPUT_PATH, fallback_path


def run() -> None:
    args = parse_args()
    extra_roots = [Path(text).expanduser().resolve() for text in args.extra_root if str(text).strip()]
    raw_roots = [RAW_ROOT.resolve()] + extra_roots

    raw_df, diagnostics = discover_raw_records(raw_roots)
    sheets = build_output_sheets(raw_df)
    canonical_path, fallback_path = write_workbook(sheets)

    bus = sheets["Bus_5MinCols"]
    car = sheets["Car_5MinCols"]
    motorcycle = sheets["Motorcycle_5MinCols"]
    total = sheets["Total_Volume_5MinCols"]

    interval_cols = [
        col
        for col in total.columns
        if isinstance(col, str) and "-" in col and to_seconds(col.split("-")[0]) is not None and to_seconds(col.split("-")[1]) is not None
    ]

    total_ok = True
    for col in interval_cols:
        if int(total[col].sum()) != int(bus[col].sum() + car[col].sum() + motorcycle[col].sum()):
            total_ok = False
            break

    has_905 = raw_df[raw_df["SourceWorkbook"].astype(str).str.endswith("/905.xlsx")]
    has_905_interval = bool((has_905["Interval"] == "09:15:00-09:20:00").any()) if not has_905.empty else False

    source_non_empty = True
    if interval_cols:
        nonzero_rows = total[total[interval_cols].sum(axis=1) > 0]
        source_non_empty = bool((nonzero_rows["SourceWorkbook"].astype(str).str.strip() != "").all())

    total_row_present = False
    for frame in sheets.values():
        if (frame["Direction"].astype(str).str.strip().str.lower() == "total").any() or (frame["Approach"].astype(str).str.strip().str.lower() == "total").any():
            total_row_present = True
            break

    focus_locations = {"Kupondole Busstop", "Kandevtasthan", "Patan Dhoka Road"}
    focus_nonzero = {}
    for location in focus_locations:
        if interval_cols and not total.empty:
            loc_df = total[total["Location"].astype(str) == location]
            focus_nonzero[location] = int(loc_df[interval_cols].sum().sum())
        else:
            focus_nonzero[location] = 0

    print("canonical_output_path:", canonical_path)
    if fallback_path is not None:
        print("fallback_output_path:", fallback_path)

    print("raw_roots:", diagnostics["raw_roots"])
    print("scanned_files:", diagnostics["scanned_files"])
    print("included_files:", diagnostics["included_files"])
    print("dropped_duplicate_candidates:", diagnostics["dropped_duplicate_candidates"])
    print("skipped_bad_structure:", diagnostics["skipped_bad_structure"])
    print("skipped_outside_window:", diagnostics["skipped_outside_window"])
    print("skipped_non_allowed:", diagnostics["skipped_non_allowed"])

    print("validation_total_volume_equals_sum:", total_ok)
    print("validation_905_interval_present:", has_905_interval)
    print("validation_sourceworkbook_nonempty:", source_non_empty)
    print("validation_no_total_rows:", not total_row_present)
    print("focus_location_nonzero_sums:", focus_nonzero)

    for sheet_name, frame in sheets.items():
        print(f"sheet_rows_{sheet_name}:", len(frame))


if __name__ == "__main__":
    run()
