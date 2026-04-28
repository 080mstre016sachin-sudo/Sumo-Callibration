from pathlib import Path
from collections import defaultdict
from datetime import datetime, timedelta
import re
import shutil
from openpyxl import load_workbook, Workbook

BASE = Path(r"C:/Users/gupta/SumoCallibration")
ROOT = BASE / "ProcessedVideoOutput"
CANONICAL_OUTPUT = ROOT / "EdgeWise_DirectionalCounts_5MinIntervalColumns_Datewise.xlsx"

ALLOWED_FOLDERS = {
    "Thapathali",
    "Jwagal",
    "Kandevtasthan",
    "Kesharmahal",
    "Krishna Marg (Kupondole Busstop)",
    "Krishnagalli",
    "Maitri Marg (Bhakundole)",
    "Patan Dhoka Road",
    "Pulchowk North",
    "Pulchowk South",
}

LOCATION_MAP = {
    "Thapathali": "Thapathali",
    "Krishna Marg (Kupondole Busstop)": "Kupondole Busstop",
    "Kandevtasthan": "Kandevtasthan",
    "Jwagal": "Jwagal",
    "Patan Dhoka Road": "Patan Dhoka Road",
    "Maitri Marg (Bhakundole)": "Bhakundole",
    "Krishnagalli": "Krishnagalli",
    "Kesharmahal": "Kesharmahal",
    "Pulchowk North": "Pulchowk North",
    "Pulchowk South": "Pulchowk South",
}

LOCATION_ORDER = [
    "Thapathali",
    "Kupondole Busstop",
    "Kandevtasthan",
    "Jwagal",
    "Patan Dhoka Road",
    "Bhakundole",
    "Krishnagalli",
    "Kesharmahal",
    "Pulchowk North",
    "Pulchowk South",
]
LOC_IDX = {v: i for i, v in enumerate(LOCATION_ORDER)}

SHEETS = {
    "Bus": "Bus_5MinCols",
    "Car": "Car_5MinCols",
    "Motorcycle": "Motorcycle_5MinCols",
    "Total": "Total_Volume_5MinCols",
}

SUMMARY_MARKERS = ["compile+summary", "edgewise", "datewise", "compact"]
RANGE_RE = re.compile(r"(?<!\d)(\d{1,2})[.:](\d{2})[.:](\d{2})\s*[-_]\s*(\d{1,2})[.:](\d{2})[.:](\d{2})(?!\d)")
NUMERIC_STEM_RE = re.compile(r"^(\d{3,4})$")

def normalize_date(s: str):
    raw = s.strip()
    cleaned = re.sub(r"[._/\\]", "-", raw)
    fmts = [
        "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y",
        "%Y-%d-%m", "%d-%m-%y", "%m-%d-%y",
        "%Y%m%d", "%d%m%Y"
    ]
    for fmt in fmts:
        try:
            d = datetime.strptime(cleaned, fmt).date()
            return d.isoformat()
        except Exception:
            pass
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 8:
        try:
            if digits.startswith("20"):
                d = datetime.strptime(digits, "%Y%m%d").date()
            else:
                d = datetime.strptime(digits, "%d%m%Y").date()
            return d.isoformat()
        except Exception:
            return None
    return None

def parse_interval_from_filename(filename: str):
    stem = Path(filename).stem
    m = RANGE_RE.search(stem)
    if m:
        h1, m1, s1, h2, m2, s2 = map(int, m.groups())
        try:
            t1 = datetime(2000, 1, 1, h1, m1, s1)
            t2 = datetime(2000, 1, 1, h2, m2, s2)
            return f"{t1.strftime('%H:%M:%S')}-{t2.strftime('%H:%M:%S')}"
        except Exception:
            pass
    m2 = NUMERIC_STEM_RE.match(stem)
    if m2:
        val = m2.group(1)
        if len(val) == 3:
            hh = int(val[0])
            mm = int(val[1:])
        else:
            hh = int(val[:2])
            mm = int(val[2:])
        try:
            base = datetime(2000, 1, 1, hh, mm, 0)
            start = base + timedelta(minutes=10)
            end = start + timedelta(minutes=5)
            return f"{start.strftime('%H:%M:%S')}-{end.strftime('%H:%M:%S')}"
        except Exception:
            return None
    return None

def normalize_direction(d):
    if d is None:
        return None
    x = str(d).strip()
    if not x:
        return None
    x = re.sub(r"\s+", " ", x.replace("_", "-"))
    return x

def get_approach(direction: str):
    m = re.split(r"-to-", direction, flags=re.IGNORECASE)
    if len(m) > 1:
        return m[0].strip()
    return direction

def as_number(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return 0
    try:
        return float(s)
    except Exception:
        return 0

def class_bucket(v):
    s = str(v).strip().lower()
    if "bus" in s:
        return "Bus"
    if "car" in s:
        return "Car"
    if "motor" in s or "bike" in s:
        return "Motorcycle"
    return None

def find_sheet(wb, name):
    for ws in wb.worksheets:
        if ws.title.strip().lower() == name:
            return ws
    return None

def parse_direction_counts(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        hv = ws.cell(row=1, column=c).value
        if hv is not None:
            headers[str(hv).strip().lower()] = c
    dcol = headers.get("direction")
    ccol = headers.get("class")
    ncol = headers.get("count")
    if not (dcol and ccol and ncol):
        return None
    out = defaultdict(lambda: {"Bus": 0.0, "Car": 0.0, "Motorcycle": 0.0})
    found = False
    for r in range(2, ws.max_row + 1):
        direction = normalize_direction(ws.cell(row=r, column=dcol).value)
        cls = ws.cell(row=r, column=ccol).value
        cnt = as_number(ws.cell(row=r, column=ncol).value)
        if direction is None:
            continue
        bucket = class_bucket(cls)
        if bucket is None:
            continue
        out[direction][bucket] += cnt
        found = True
    return out if found else None

def parse_summary(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        hv = ws.cell(row=1, column=c).value
        if hv is not None:
            headers[str(hv).strip().lower()] = c
    dcol = headers.get("direction")
    bcol = headers.get("bus")
    ccol = headers.get("car")
    mcol = headers.get("motorcycle")
    if not (dcol and bcol and ccol and mcol):
        return None
    out = defaultdict(lambda: {"Bus": 0.0, "Car": 0.0, "Motorcycle": 0.0})
    found = False
    for r in range(2, ws.max_row + 1):
        direction = normalize_direction(ws.cell(row=r, column=dcol).value)
        if direction is None:
            continue
        out[direction]["Bus"] += as_number(ws.cell(row=r, column=bcol).value)
        out[direction]["Car"] += as_number(ws.cell(row=r, column=ccol).value)
        out[direction]["Motorcycle"] += as_number(ws.cell(row=r, column=mcol).value)
        found = True
    return out if found else None

all_intervals = set()
data = {k: defaultdict(lambda: defaultdict(float)) for k in SHEETS.keys()}
sources = {k: defaultdict(set) for k in SHEETS.keys()}

included_files = 0
excluded_summary_like = 0

for path in ROOT.rglob("*.xlsx"):
    if path.name.startswith("~$"):
        continue
    rel = path.relative_to(ROOT)
    parts = rel.parts
    if len(parts) != 4:
        continue

    folder = parts[0]
    if folder not in ALLOWED_FOLDERS:
        continue

    lower_name = path.name.lower()
    if any(m in lower_name for m in SUMMARY_MARKERS) or ("compile" in lower_name and "summary" in lower_name):
        excluded_summary_like += 1
        continue

    date_norm = normalize_date(parts[1])
    if not date_norm:
        continue

    interval = parse_interval_from_filename(path.name)
    if not interval:
        continue

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        continue

    parsed = None
    ws_dc = find_sheet(wb, "direction counts")
    if ws_dc is not None:
        parsed = parse_direction_counts(ws_dc)
    if parsed is None:
        ws_sum = find_sheet(wb, "summary")
        if ws_sum is not None:
            parsed = parse_summary(ws_sum)

    if not parsed:
        continue

    included_files += 1
    all_intervals.add(interval)

    loc_label = LOCATION_MAP[folder]
    rel_str = rel.as_posix()

    for direction, cls_map in parsed.items():
        approach = get_approach(direction)
        row_key = (loc_label, date_norm, approach, direction)
        b = cls_map["Bus"]
        c = cls_map["Car"]
        m = cls_map["Motorcycle"]

        data["Bus"][row_key][interval] += b
        data["Car"][row_key][interval] += c
        data["Motorcycle"][row_key][interval] += m
        data["Total"][row_key][interval] += (b + c + m)

        sources["Bus"][row_key].add(rel_str)
        sources["Car"][row_key].add(rel_str)
        sources["Motorcycle"][row_key].add(rel_str)
        sources["Total"][row_key].add(rel_str)

interval_cols = sorted(all_intervals)

wb_out = Workbook()
def_sheet = wb_out.active
wb_out.remove(def_sheet)

sheet_row_counts = {}
nonempty_source_ok = True
has_905_mapped = False

all_row_keys = set()
for k in data:
    all_row_keys.update(data[k].keys())

def row_sort_key(rk):
    loc, d, a, direction = rk
    return (LOC_IDX.get(loc, 999), d, a.lower(), direction.lower())

for cls_key, sheet_name in SHEETS.items():
    ws = wb_out.create_sheet(sheet_name)
    header = ["Location", "Date", "Approach", "Direction", "SourceWorkbook"] + interval_cols
    ws.append(header)

    rows_written = 0
    for rk in sorted(data[cls_key].keys(), key=row_sort_key):
        loc, d, app, direction = rk
        src = " | ".join(sorted(sources[cls_key][rk]))
        vals = []
        row_nonzero = False
        for col in interval_cols:
            v = data[cls_key][rk].get(col, 0)
            iv = int(round(v))
            vals.append(iv)
            if iv != 0:
                row_nonzero = True
        if row_nonzero and not src.strip():
            nonempty_source_ok = False
        if "905.xlsx" in src and "09:15:00-09:20:00" in interval_cols:
            idx = interval_cols.index("09:15:00-09:20:00")
            if vals[idx] != 0:
                has_905_mapped = True
        ws.append([loc, d, app, direction, src] + vals)
        rows_written += 1

    sheet_row_counts[sheet_name] = rows_written

# Validation 1: total equals sum of bus+car+motorcycle
is_total_consistent = True
for rk in all_row_keys:
    for col in interval_cols:
        b = int(round(data["Bus"][rk].get(col, 0)))
        c = int(round(data["Car"][rk].get(col, 0)))
        m = int(round(data["Motorcycle"][rk].get(col, 0)))
        t = int(round(data["Total"][rk].get(col, 0)))
        if t != (b + c + m):
            is_total_consistent = False
            break
    if not is_total_consistent:
        break

fallback_path = None
lock_occurred = False

try:
    wb_out.save(CANONICAL_OUTPUT)
except PermissionError:
    lock_occurred = True
except Exception:
    lock_occurred = True

if lock_occurred:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fallback_path = ROOT / f"EdgeWise_DirectionalCounts_5MinIntervalColumns_Datewise_unlocked_{stamp}.xlsx"
    wb_out.save(fallback_path)
    try:
        shutil.copy2(fallback_path, CANONICAL_OUTPUT)
    except Exception:
        pass

print(f"Validation_TotalVolume_Equals_Sum: {is_total_consistent}")
print(f"Validation_905_To_091500_092000: {has_905_mapped}")
print(f"Excluded_SummaryLike_Files: {excluded_summary_like}")
print(f"Included_Files: {included_files}")
print(f"Validation_SourceWorkbook_NonEmpty_When_NonZero: {nonempty_source_ok}")
print(f"canonical_output_path: {CANONICAL_OUTPUT}")
if fallback_path is not None:
    print(f"fallback_output_path: {fallback_path}")
for sname, rcnt in sheet_row_counts.items():
    print(f"SheetRows_{sname}: {rcnt}")
