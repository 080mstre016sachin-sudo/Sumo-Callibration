from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


SOURCE_PATTERN = "zone_map_*.xlsx"
TARGET_WORKBOOK = "9.xlsx"


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 24)


def write_sheet(workbook: Workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    sheet = workbook.create_sheet(title=sheet_name)
    sheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        sheet.append(list(row))

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    autosize_columns(sheet)
    sheet.freeze_panes = "A2"


def write_compiled_direction_sheet(workbook: Workbook, dataframe: pd.DataFrame, source_count: int) -> None:
    sheet_name = "Compiled Directional"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    sheet = workbook.create_sheet(title=sheet_name)
    now = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")

    sheet.append(["Report", "All zone_map directional counts compiled"])
    sheet.append(["Generated At", now])
    sheet.append(["Source Files", source_count])
    sheet.append(["Date/Time Basis", "Dominant date and dominant 5-minute time from Frame Events per file"])
    sheet.append([])

    headers = [
        "dominant_date",
        "start_time",
        "end_time",
        "source_file",
        "approach",
        "direction",
        "direction_count",
        "bus",
        "car",
        "motorcycle",
        "summary_5min_total",
    ]
    sheet.append(headers)
    for row in dataframe[headers].itertuples(index=False, name=None):
        sheet.append(list(row))

    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for cell in sheet[6]:
        cell.font = Font(bold=True)

    autosize_columns(sheet)
    sheet.freeze_panes = "A7"


def read_summary_total(workbook) -> int:
    sheet = workbook["Summary"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        if row[0] is not None and str(row[0]).strip().upper() == "TOTAL":
            try:
                return int(row[1] or 0)
            except (TypeError, ValueError) as exc:
                logging.warning("Could not parse Summary TOTAL value %r: %s", row[1], exc)
                return 0
    return 0


def read_approach_rows(workbook) -> list[dict[str, object]]:
    sheet = workbook["Approach-Wise"]
    rows: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            logging.warning("Skipping short row in Approach-Wise sheet (expected 6+ columns, got %d)", len(row))
            continue
        if not row[1]:
            continue
        try:
            rows.append(
                {
                    "approach": str(row[0]).strip() if row[0] else "",
                    "direction": str(row[1]).strip(),
                    "direction_count": int(row[2] or 0),
                    "bus": int(row[3] or 0),
                    "car": int(row[4] or 0),
                    "motorcycle": int(row[5] or 0),
                }
            )
        except (TypeError, ValueError) as exc:
            logging.warning("Skipping malformed Approach-Wise row %r: %s", row[:6], exc)
    return rows


def dominant_window_from_frame_events(workbook) -> dict[str, object]:
    sheet = workbook["Frame Events"]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {str(value).strip(): idx for idx, value in enumerate(header) if value is not None}
    timestamp_idx = header_map.get("Timestamp")
    if timestamp_idx is None:
        return {
            "dominant_date": "",
            "start_time": "",
            "end_time": "",
            "frame_total": 0,
            "frame_on_dominant_date": 0,
            "frame_in_dominant_5min": 0,
        }

    timestamps = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_ts = row[timestamp_idx]
        if not raw_ts:
            continue
        parsed = pd.to_datetime(raw_ts, errors="coerce")
        if pd.isna(parsed):
            continue
        timestamps.append(parsed)

    if not timestamps:
        return {
            "dominant_date": "",
            "start_time": "",
            "end_time": "",
            "frame_total": 0,
            "frame_on_dominant_date": 0,
            "frame_in_dominant_5min": 0,
        }

    series = pd.Series(timestamps)
    dominant_date = series.dt.date.mode().iloc[0]
    on_dominant_date = series[series.dt.date == dominant_date]

    # Use the most frequent 5-minute slot on the dominant date as the file's dominant time.
    dominant_start = on_dominant_date.dt.floor("5min").mode().iloc[0]
    dominant_end = dominant_start + pd.Timedelta(minutes=5) - pd.Timedelta(seconds=1)
    in_dominant_window = ((on_dominant_date >= dominant_start) & (on_dominant_date < dominant_start + pd.Timedelta(minutes=5))).sum()

    return {
        "dominant_date": str(dominant_date),
        "start_time": dominant_start.strftime("%H:%M:%S"),
        "end_time": dominant_end.strftime("%H:%M:%S"),
        "frame_total": int(series.count()),
        "frame_on_dominant_date": int(on_dominant_date.count()),
        "frame_in_dominant_5min": int(in_dominant_window),
    }


def main() -> None:
    base_dir = Path(r"c:\Users\gupta\SumoCallibration\Kandevtasthan\2026 02 22\9")
    target_path = base_dir / TARGET_WORKBOOK

    source_files = sorted(base_dir.glob(SOURCE_PATTERN))
    if not source_files:
        raise SystemExit("No zone_map files found.")

    direction_rows: list[dict[str, object]] = []
    file_overview_rows: list[dict[str, object]] = []

    for source_path in source_files:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        if not {"Summary", "Approach-Wise", "Frame Events"}.issubset(set(workbook.sheetnames)):
            continue

        summary_total = read_summary_total(workbook)
        approach_rows = read_approach_rows(workbook)
        time_info = dominant_window_from_frame_events(workbook)
        approach_total = sum(int(item["direction_count"]) for item in approach_rows)

        file_overview_rows.append(
            {
                "source_file": source_path.name,
                "dominant_date": time_info["dominant_date"],
                "start_time": time_info["start_time"],
                "end_time": time_info["end_time"],
                "summary_5min_total": summary_total,
                "approach_total": approach_total,
                "match_summary": "YES" if summary_total == approach_total else "NO",
                "frame_total": time_info["frame_total"],
                "frame_on_dominant_date": time_info["frame_on_dominant_date"],
                "frame_in_dominant_5min": time_info["frame_in_dominant_5min"],
            }
        )

        for item in approach_rows:
            direction_rows.append(
                {
                    "source_file": source_path.name,
                    "dominant_date": time_info["dominant_date"],
                    "start_time": time_info["start_time"],
                    "end_time": time_info["end_time"],
                    "summary_5min_total": summary_total,
                    "approach": item["approach"],
                    "direction": item["direction"],
                    "direction_count": item["direction_count"],
                    "bus": item["bus"],
                    "car": item["car"],
                    "motorcycle": item["motorcycle"],
                }
            )

    if not direction_rows:
        raise SystemExit("No usable rows found in zone_map files.")

    direction_df = pd.DataFrame(direction_rows).sort_values(["dominant_date", "start_time", "source_file", "direction"])
    overview_df = pd.DataFrame(file_overview_rows).sort_values(["dominant_date", "start_time", "source_file"])

    slot_direction_df = (
        direction_df.groupby(["dominant_date", "start_time", "end_time", "direction"], as_index=False)["direction_count"]
        .sum()
        .sort_values(["dominant_date", "start_time", "direction"])
    )

    summary_totals_by_slot = (
        overview_df.groupby(["dominant_date", "start_time", "end_time"], as_index=False)["summary_5min_total"]
        .sum()
        .rename(columns={"summary_5min_total": "summary_total"})
    )
    direction_totals_by_slot = (
        direction_df.groupby(["dominant_date", "start_time", "end_time"], as_index=False)["direction_count"]
        .sum()
        .rename(columns={"direction_count": "direction_total"})
    )
    slot_total_df = summary_totals_by_slot.merge(
        direction_totals_by_slot,
        on=["dominant_date", "start_time", "end_time"],
        how="outer",
    ).sort_values(["dominant_date", "start_time"])

    workbook = load_workbook(target_path) if target_path.exists() else Workbook()
    if workbook.sheetnames == ["Sheet"] and workbook["Sheet"].max_row == 1 and workbook["Sheet"].max_column == 1 and workbook["Sheet"]["A1"].value is None:
        del workbook["Sheet"]
    if workbook.sheetnames == ["Sheet1"] and workbook["Sheet1"].max_row == 1 and workbook["Sheet1"].max_column == 1 and workbook["Sheet1"]["A1"].value is None:
        del workbook["Sheet1"]

    write_sheet(workbook, "Direction Wise 5 Min", direction_df)
    write_sheet(workbook, "File Validation", overview_df)
    write_sheet(workbook, "Slot Direction Sum", slot_direction_df)
    write_sheet(workbook, "Slot Total Check", slot_total_df)
    write_compiled_direction_sheet(workbook, direction_df, len(source_files))

    if not workbook.sheetnames:
        workbook.create_sheet("5 Min Summary")

    workbook.save(target_path)

    print(f"Wrote direction-wise summaries to {target_path}")
    print(f"Source files used: {len(source_files)}")
    print(f"Direction rows: {len(direction_df)}")
    print(f"Slot-direction rows: {len(slot_direction_df)}")
    print("Compiled sheet: Compiled Directional")
    print(f"All summary totals matched approach totals: {bool((overview_df['match_summary'] == 'YES').all())}")


if __name__ == "__main__":
    main()