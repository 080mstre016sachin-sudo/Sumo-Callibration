from __future__ import annotations

from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROOT_DIR = Path(r"c:\Users\gupta\SumoCallibration\20260222")
DATE_TOKEN = "20260222"
DATE_VALUE = f"{DATE_TOKEN[0:4]}-{DATE_TOKEN[4:6]}-{DATE_TOKEN[6:8]}"
INTERVAL_NAME_RE = re.compile(r"^(\d{3,4})\.xlsx$", re.IGNORECASE)
REPORT_NAME = "20260222_summary_report.xlsx"


def parse_interval_start(time_code: str) -> str:
    padded = time_code.zfill(4)
    return f"{padded[0:2]}:{padded[2:4]}:00"


def parse_interval_end(start_time: str) -> str:
    start_dt = pd.to_datetime(f"{DATE_VALUE} {start_time}")
    end_dt = start_dt + pd.Timedelta(minutes=5)
    return end_dt.strftime("%H:%M:%S")


def to_int(value) -> int:
    if value is None:
        return 0
    if isinstance(value, str) and not value.strip():
        return 0
    return int(float(value))


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 26)


def write_dataframe(workbook: Workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    sheet = workbook.create_sheet(sheet_name)
    sheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        sheet.append(list(row))

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    autosize_columns(sheet)
    sheet.freeze_panes = "A2"


def remove_event_sheets(workbook_path: Path) -> bool:
    wb = load_workbook(workbook_path)
    event_sheets = [name for name in wb.sheetnames if "event" in name.lower()]
    if not event_sheets:
        return False

    for name in event_sheets:
        del wb[name]

    wb.save(workbook_path)
    return True


def rows_from_approach_wise(sheet) -> list[dict[str, int | str]]:
    rows: list[dict[str, int | str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        direction = row[1]
        if not direction:
            continue
        rows.append(
            {
                "direction": str(direction).strip(),
                "total": to_int(row[2]),
                "bus": to_int(row[3]),
                "car": to_int(row[4]),
                "motorcycle": to_int(row[5]),
            }
        )
    return rows


def rows_from_summary(sheet) -> list[dict[str, int | str]]:
    rows: list[dict[str, int | str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        direction = row[0]
        if not direction:
            continue
        direction_text = str(direction).strip()
        if direction_text.upper() == "TOTAL":
            continue
        rows.append(
            {
                "direction": direction_text,
                "total": to_int(row[1]),
                "bus": to_int(row[2]),
                "car": to_int(row[3]),
                "motorcycle": to_int(row[4]),
            }
        )
    return rows


def rows_from_direction_counts(sheet) -> list[dict[str, int | str]]:
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {str(value).strip().lower(): idx for idx, value in enumerate(header) if value is not None}

    direction_idx = header_map.get("direction")
    class_idx = header_map.get("class")
    count_idx = header_map.get("count")
    if direction_idx is None or class_idx is None or count_idx is None:
        return []

    grouped: dict[str, dict[str, int | str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_len = len(row) if row else 0
        if row_len == 0:
            continue
        if direction_idx >= row_len or class_idx >= row_len or count_idx >= row_len:
            continue
        if not row[direction_idx]:
            continue
        direction = str(row[direction_idx]).strip()
        class_name = str(row[class_idx]).strip().lower()
        count = to_int(row[count_idx])

        if direction not in grouped:
            grouped[direction] = {
                "direction": direction,
                "total": 0,
                "bus": 0,
                "car": 0,
                "motorcycle": 0,
            }

        grouped[direction]["total"] = int(grouped[direction]["total"]) + count
        if class_name in ("bus", "car", "motorcycle"):
            grouped[direction][class_name] = int(grouped[direction][class_name]) + count

    return list(grouped.values())


def extract_direction_rows(workbook_path: Path) -> list[dict[str, int | str]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    if "Approach-Wise" in wb.sheetnames:
        return rows_from_approach_wise(wb["Approach-Wise"])
    if "Direction Counts" in wb.sheetnames:
        return rows_from_direction_counts(wb["Direction Counts"])
    if "Summary" in wb.sheetnames:
        return rows_from_summary(wb["Summary"])

    return []


def build_folder_report(folder: Path, interval_files: list[Path]) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows: list[dict[str, int | str]] = []

    for path in sorted(interval_files):
        match = INTERVAL_NAME_RE.match(path.name)
        if not match:
            continue

        start_time = parse_interval_start(match.group(1))
        end_time = parse_interval_end(start_time)

        direction_rows = extract_direction_rows(path)
        for item in direction_rows:
            rows.append(
                {
                    "date": DATE_VALUE,
                    "start_time": start_time,
                    "end_time": end_time,
                    "direction": item["direction"],
                    "total": int(item["total"]),
                    "bus": int(item["bus"]),
                    "car": int(item["car"]),
                    "motorcycle": int(item["motorcycle"]),
                    "source_file": path.name,
                }
            )

    if not rows:
        empty_cols = ["date", "start_time", "end_time", "direction", "total", "bus", "car", "motorcycle", "source_file"]
        empty = pd.DataFrame(columns=empty_cols)
        return empty, empty

    five_min = pd.DataFrame(rows).sort_values(["date", "start_time", "direction", "source_file"])

    five_min_dt = five_min.copy()
    five_min_dt["start_dt"] = pd.to_datetime(five_min_dt["date"] + " " + five_min_dt["start_time"])
    five_min_dt["bucket_15m"] = five_min_dt["start_dt"].dt.floor("15min")

    fifteen_min = (
        five_min_dt.groupby(["date", "bucket_15m", "direction"], as_index=False)[["total", "bus", "car", "motorcycle"]]
        .sum()
        .sort_values(["date", "bucket_15m", "direction"])
    )
    fifteen_min["start_time"] = fifteen_min["bucket_15m"].dt.strftime("%H:%M:%S")
    fifteen_min["end_time"] = (fifteen_min["bucket_15m"] + pd.Timedelta(minutes=15)).dt.strftime("%H:%M:%S")
    fifteen_min = fifteen_min[["date", "start_time", "end_time", "direction", "total", "bus", "car", "motorcycle"]]

    return five_min, fifteen_min


def main() -> None:
    if not ROOT_DIR.exists():
        raise SystemExit(f"Folder not found: {ROOT_DIR}")

    # Remove event sheets from all Excel files under 20260222.
    modified_files = 0
    for excel_path in ROOT_DIR.rglob("*.xlsx"):
        if excel_path.name.lower().startswith("~$"):
            continue
        if remove_event_sheets(excel_path):
            modified_files += 1

    interval_by_folder: dict[Path, list[Path]] = {}
    for excel_path in ROOT_DIR.rglob("*.xlsx"):
        if excel_path.name.lower().startswith("~$"):
            continue
        if excel_path.name == REPORT_NAME:
            continue
        if INTERVAL_NAME_RE.match(excel_path.name):
            interval_by_folder.setdefault(excel_path.parent, []).append(excel_path)

    if not interval_by_folder:
        raise SystemExit("No interval files (e.g., 905.xlsx) found under 20260222.")

    report_count = 0
    for folder, interval_files in sorted(interval_by_folder.items(), key=lambda x: str(x[0])):
        five_min, fifteen_min = build_folder_report(folder, interval_files)

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        write_dataframe(wb, "5 Min Summary", five_min)
        write_dataframe(wb, "15 Min Summary", fifteen_min)

        report_path = folder / REPORT_NAME
        wb.save(report_path)
        report_count += 1

        print(f"Report written: {report_path}")
        print(f"  5-min rows: {len(five_min)} | 15-min rows: {len(fifteen_min)}")

    print(f"Event sheets removed from files: {modified_files}")
    print(f"Folders processed: {report_count}")


if __name__ == "__main__":
    main()