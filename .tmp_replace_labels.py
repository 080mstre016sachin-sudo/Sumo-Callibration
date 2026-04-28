from pathlib import Path
import re
from openpyxl import load_workbook

base = Path(r"C:/Users/gupta/SumoCallibration")
target_root = base / "ProcessedVideoOutput" / "Krishna Marg (Kupondole Busstop)"

sheet_targets = {"direction counts", "events"}
phrase_replacements = [
    ("East-to-West", "East-to-South"),
    ("North-to-West", "North-to_East"),
]
token_patterns = [
    (re.compile(r"\bEW\b"), "ES"),
    (re.compile(r"\bNW\b"), "NE"),
]

def transform_text(text: str) -> str:
    out = text
    for old, new in phrase_replacements:
        out = out.replace(old, new)
    for pat, rep in token_patterns:
        out = pat.sub(rep, out)
    return out

files_scanned = 0
files_changed = 0
total_cell_changes = 0
per_file = []
skipped_locked = []

if not target_root.exists():
    print(f"Target folder not found: {target_root}")
    raise SystemExit(1)

for path in sorted(target_root.rglob("*.xlsx")):
    if path.name.startswith("~$"):
        continue

    files_scanned += 1
    try:
        wb = load_workbook(path)
    except Exception:
        skipped_locked.append(path)
        continue

    wb_changed = False
    cell_changes = 0

    for ws in wb.worksheets:
        if ws.title.strip().lower() not in sheet_targets:
            continue
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    new_val = transform_text(val)
                    if new_val != val:
                        cell.value = new_val
                        cell_changes += 1
                        wb_changed = True

    if wb_changed:
        try:
            wb.save(path)
            files_changed += 1
            total_cell_changes += cell_changes
            per_file.append((path, cell_changes))
        except PermissionError:
            skipped_locked.append(path)
        except Exception:
            skipped_locked.append(path)

print(f"Total files scanned: {files_scanned}")
print(f"Files changed: {files_changed}")
print(f"Total cell changes: {total_cell_changes}")
print("Per-file changes:")
if per_file:
    for p, c in per_file:
        print(f"- {p.relative_to(base)}: {c}")
else:
    print("- none")
if skipped_locked:
    print(f"Skipped (locked/unreadable): {len(skipped_locked)}")
